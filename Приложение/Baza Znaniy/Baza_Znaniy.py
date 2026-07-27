import sys
import json
import re
import html
import sqlite3
import getpass
import uuid
import shutil
import time
from datetime import datetime
from pathlib import Path
from urllib.parse import quote, unquote

from shiboken6 import isValid

from PySide6.QtCore import Qt, QTimer, QUrl, QSize, QRect, QRectF, QPoint, QObject, QEvent
from PySide6.QtWidgets import QStyle
from PySide6.QtGui import (
    QDesktopServices, QFont, QPixmap, QPalette, QColor, QTextDocument, QTextCursor,
    QTextCharFormat, QAbstractTextDocumentLayout, QPainter, QPen, QCursor,
    QMouseEvent, QPainterPath,
)
from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, QLabel,
    QLineEdit, QPushButton, QTreeWidget, QTreeWidgetItem, QSplitter, QTabWidget,
    QScrollArea, QFrame, QToolButton, QListWidget, QListWidgetItem, QTextEdit,
    QTextBrowser, QCheckBox, QComboBox, QMessageBox, QSpinBox, QGroupBox,
    QSizePolicy, QAbstractItemView, QInputDialog, QDialog, QDialogButtonBox,
    QFormLayout, QPlainTextEdit, QGraphicsDropShadowEffect, QFileDialog,
    QStyledItemDelegate, QStyleOptionViewItem, QMenu,
)

# ================== ПУТИ И ВЕРСИОНИРОВАНИЕ ==================

APP_VERSION = "1.0.2"
VERSIONS_DIR_NAME = "versions"
VERSION_CONFIG_NAME = "version.json"
AUTO_REFRESH_INTERVAL_MS = 500
UPDATE_CHECK_INTERVAL_MS = 5 * 60 * 1000
BACKUP_MAX_COUNT = 10

NETWORK_APP_DIR = Path(
    r"\\fileserver\УТЗ\10 Служба технического директора\05 СКБт"
    r"\!Общая\Сканы для ГПЯ\!Обработка документов\База знаний по задачам"
)


def resolve_app_paths():
    """Корень развёртывания (БД, version.json) и папка текущего exe."""
    if getattr(sys, "frozen", False):
        exe_dir = Path(sys.executable).resolve().parent
    else:
        exe_dir = Path(__file__).resolve().parent

    if exe_dir.parent.name.lower() == VERSIONS_DIR_NAME:
        app_root = exe_dir.parent.parent
    else:
        app_root = exe_dir

    return app_root, exe_dir


def parse_version_parts(version_str):
    parts = []
    for part in re.split(r"[.\-]", str(version_str or "").strip()):
        if part.isdigit():
            parts.append(int(part))
        elif part:
            break
    return tuple(parts) if parts else (0,)


def is_version_newer(available_version, current_version):
    return parse_version_parts(available_version) > parse_version_parts(current_version)


def load_deployed_version_config():
    config_path = APP_ROOT / VERSION_CONFIG_NAME
    if not config_path.exists():
        return {}

    try:
        with config_path.open(encoding="utf-8-sig") as config_file:
            data = json.load(config_file)
        return data if isinstance(data, dict) else {}
    except (OSError, ValueError):
        return {}


def resolve_db_path(app_root):
    """БД в корне развёртывания; при локальной разработке — сетевая, если доступна."""
    deployed_db = app_root / "knowledge_base.db"

    if getattr(sys, "frozen", False):
        exe_dir = Path(sys.executable).resolve().parent
        if exe_dir.parent.name.lower() == VERSIONS_DIR_NAME:
            return deployed_db

    try:
        network_db = NETWORK_APP_DIR / "knowledge_base.db"
        if NETWORK_APP_DIR.exists():
            return network_db
    except OSError:
        pass
    return deployed_db


def cleanup_sqlite_sidecar_files(db_path):
    """Убирает WAL/SHM, чтобы файл БД можно было заменить на сетевой папке."""
    db_path = Path(db_path)
    wal_path = Path(f"{db_path}-wal")
    shm_path = Path(f"{db_path}-shm")

    if not wal_path.exists() and not shm_path.exists():
        return

    for attempt in range(2):
        try:
            temp_conn = sqlite3.connect(str(db_path), timeout=5)
            temp_conn.execute("PRAGMA journal_mode = DELETE")
            temp_conn.execute("PRAGMA wal_checkpoint(TRUNCATE)")
            temp_conn.commit()
            temp_conn.close()
            break
        except sqlite3.OperationalError:
            if attempt == 0:
                time.sleep(2)

    for sidecar in (shm_path, wal_path):
        try:
            if sidecar.exists():
                sidecar.unlink()
        except OSError:
            pass


def backup_database_file(db_path):
    """Создаёт резервную копию БД перед запуском."""
    db_path = Path(db_path)
    if not db_path.exists():
        return

    backup_dir = db_path.parent / "backups"
    try:
        backup_dir.mkdir(parents=True, exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_path = backup_dir / f"knowledge_base_{timestamp}.db"
        shutil.copy2(db_path, backup_path)
        existing = sorted(
            backup_dir.glob("knowledge_base_*.db"),
            key=lambda p: p.stat().st_mtime,
            reverse=True,
        )
        for old in existing[BACKUP_MAX_COUNT:]:
            try:
                old.unlink()
            except OSError:
                pass
    except OSError:
        pass


def ensure_db_dir(db_path):
    db_path = Path(db_path)
    db_dir = db_path.parent
    try:
        if not db_dir.exists():
            raise sqlite3.OperationalError(f"Папка базы данных не найдена: {db_dir}")
        if not db_path.exists():
            test_conn = sqlite3.connect(str(db_path), timeout=5)
            test_conn.close()
        else:
            test_conn = sqlite3.connect(str(db_path), timeout=5)
            test_conn.close()
    except (OSError, sqlite3.OperationalError) as exc:
        raise sqlite3.OperationalError(
            f"Нет доступа к базе данных: {db_path}\n"
            f"Проверьте права доступа к сетевой папке.\n"
            f"Ошибка: {exc}"
        ) from exc


APP_ROOT, APP_EXE_DIR = resolve_app_paths()
DB_PATH = resolve_db_path(APP_ROOT)

ROLE_KIND = Qt.UserRole + 1
ROLE_ID = Qt.UserRole + 2
ROLE_STEP_INDEX = Qt.UserRole + 3

ROLE_CHECKLIST_KIND = Qt.UserRole + 4
ROLE_CHECKLIST_TARGET_INSTRUCTION = Qt.UserRole + 5
ROLE_CHECKLIST_TARGET_SECTION = Qt.UserRole + 6
ROLE_CHECKLIST_TARGET_INSTRUCTION_ID = Qt.UserRole + 7
ROLE_CHECKLIST_TARGET_SECTION_ID = Qt.UserRole + 8
ROLE_CHECKLIST_SECTION_INDEX = Qt.UserRole + 9
ROLE_CHECKLIST_SECTION_COLLAPSED = Qt.UserRole + 10
ROLE_CHECKLIST_TARGET_TASK_ID = Qt.UserRole + 11

ADMIN_USERNAME = "nyagavrilova"

APP_QSS = """
QMainWindow {
    background: #f7f9fd;
    color: #263238;
}

QWidget {
    background: #f7f9fd;
    color: #263238;
}

QSplitter::handle:horizontal {
    width: 6px;
    margin: 0 2px;
    background: #e3ebf5;
    border-radius: 3px;
}

QSplitter::handle:horizontal:hover {
    background: #b8cbe3;
}

QTabWidget::pane {
    border: 1px solid #d9e3f0;
    background: #fcfdff;
    border-radius: 10px;
}

QTabBar::tab {
    background: #edf4fb;
    color: #263238;
    padding: 9px 14px;
    border: 1px solid #d9e3f0;
    border-bottom: none;
    border-top-left-radius: 8px;
    border-top-right-radius: 8px;
    margin-right: 2px;
}

QTabBar::tab:selected {
    background: #ffffff;
}

QLineEdit, QTextEdit, QTextBrowser, QTreeWidget, QListWidget, QComboBox, QSpinBox {
    background: #ffffff;
    color: #263238;
    border: 1px solid #d9e3f0;
    border-radius: 8px;
    padding: 7px;
    selection-background-color: #dbeafe;
    selection-color: #263238;
}

QLineEdit:focus, QTextEdit:focus, QTextBrowser:focus, QComboBox:focus, QSpinBox:focus {
    border: 1px solid #b9cde8;
}

QComboBox QAbstractItemView {
    background: #ffffff;
    color: #263238;
    selection-background-color: #dbeafe;
    selection-color: #263238;
    border: 1px solid #d9e3f0;
    outline: 0;
}

QComboBox::drop-down {
    border: none;
    background: transparent;
}

QPushButton {
    padding: 8px 12px;
    border-radius: 8px;
    background: #dceeff;
    color: #263238;
    border: 1px solid #c3d8f2;
}

QPushButton:hover {
    background: #cfe6ff;
}

QPushButton:pressed {
    background: #bfdcff;
}

QPushButton:disabled {
    background: #eef4fb;
    color: #90a4b8;
    border: 1px solid #d9e3f0;
}

QTreeWidget {
    background: #ffffff;
    alternate-background-color: #f7fafc;
}

QListWidget {
    background: #ffffff;
    alternate-background-color: #f7fafc;
}

QListWidget::item {
    background: #ffffff;
    border: 1px solid #e1e8f0;
    border-radius: 8px;
    margin: 3px 4px;
    padding: 6px 8px;
}

QListWidget::item:selected,
QTreeWidget::item:selected {
    background: #dbeafe;
    color: #263238;
}

QHeaderView::section {
    background: #eef5fb;
    color: #263238;
    padding: 6px 8px;
    border: 1px solid #d9e3f0;
    font-weight: 600;
}

QGroupBox {
    font-weight: 600;
    border: 1px solid #d9e3f0;
    border-radius: 8px;
    margin-top: 10px;
    padding-top: 10px;
    background: #ffffff;
}

QGroupBox::title {
    subcontrol-origin: margin;
    left: 12px;
    padding: 0 4px;
    color: #263238;
}

QScrollBar:vertical {
    background: #f4f7fb;
    width: 12px;
    margin: 0px;
}

QScrollBar::handle:vertical {
    background: #c9d9ea;
    min-height: 24px;
    border-radius: 5px;
}

QScrollBar::handle:vertical:hover {
    background: #b6c9e0;
}

QStatusBar {
    background: #f7f9fd;
    color: #263238;
}

QListWidget::indicator:unchecked {
    border: 2px solid #C9D9EA;
    background: white;
    border-radius: 4px;
}
QListWidget::indicator:checked {
    border: 2px solid #C9D9EA;
    background: #C9D9EA;
    border-radius: 4px;
}

* {
    font-size: 13px;
}
QGroupBox::title {
    font-size: 13px;
}
QTabBar::tab {
    font-size: 13px;
}
QTreeWidget {
    font-size: 12px;
}

/* Синие ссылки во всех rich‑text виджетах */
a {
    color: #0000FF;
}

QToolTip {
    background-color: #fff4cc;
    color: #263238;
    border: 2px solid #f59e0b;
    border-radius: 8px;
    padding: 6px 8px;
    font-size: 12px;
}

QToolTip::title {
    color: #b45309;
}
"""


def ts():
    return datetime.now().strftime("%d.%m.%Y %H:%M")


def apply_light_palette(app):
    palette = QPalette()
    palette.setColor(QPalette.Window, QColor("#f7f9fd"))
    palette.setColor(QPalette.WindowText, QColor("#263238"))
    palette.setColor(QPalette.Base, QColor("#ffffff"))
    palette.setColor(QPalette.AlternateBase, QColor("#f7fafc"))
    palette.setColor(QPalette.Text, QColor("#263238"))
    palette.setColor(QPalette.Button, QColor("#dceeff"))
    palette.setColor(QPalette.ButtonText, QColor("#263238"))
    palette.setColor(QPalette.Highlight, QColor("#dbeafe"))
    palette.setColor(QPalette.HighlightedText, QColor("#263238"))
    palette.setColor(QPalette.ToolTipBase, QColor("#ffffff"))
    palette.setColor(QPalette.ToolTipText, QColor("#263238"))
    app.setPalette(palette)


def p(text):
    return f"<p>{html.escape(str(text))}</p>"


def html_list(items):
    return "<ul>" + "".join(f"<li>{html.escape(str(item))}</li>" for item in items) + "</ul>"


def internal_link(target, text=None):
    label = html.escape(text or str(target))
    if isinstance(target, int):
        return f"<a href='instruction://id:{target}'>{label}</a>"
    return f"<a href='instruction://{quote(str(target))}'>{label}</a>"


def escape_html(text):
    return html.escape(str(text)).replace("\n", "<br>")


def render_markdown(text):
    """
    Преобразует Markdown-подобный текст в HTML.
    Поддерживаемые стили GitHub:
    - # Заголовок 1
    - ## Заголовок 2
    - ### Заголовок 3
    - #### Заголовок 4
    - **жирный**
    - *курсив*
    - ~~зачёркнутый~~
    - `код`
    - ```многострочный код```
    - - элемент списка
    - 1. нумерованный список
    - > цитата
    - --- горизонтальная линия
    - [текст](ссылка)
    - обычный текст с переносами строк
    """
    if not text:
        return ""

    lines = text.split("\n")
    result = []
    in_code_block = False
    code_lines = []
    in_paragraph = False
    paragraph_lines = []

    def flush_paragraph():
        nonlocal in_paragraph, paragraph_lines
        if paragraph_lines:
            result.append("<p>" + "<br>".join(paragraph_lines) + "</p>")
            paragraph_lines = []
            in_paragraph = False

    def process_inline(text):
        """Обработка inline-элементов: жирный, курсив, код, зачёркнутый, ссылки."""
        # Экранируем HTML
        text = html.escape(text)

        # ```код``` -> <code>код</code>
        text = re.sub(r'```(.+?)```', r'<pre><code>\1</code></pre>', text)

        # `код` -> <code>код</code>
        text = re.sub(r'`([^`]+)`', r'<code>\1</code>', text)

        # **жирный**
        text = re.sub(r'\*\*(.+?)\*\*', r'<b>\1</b>', text)

        # *курсив* (но не **)
        text = re.sub(r'(?<!\*)\*([^*\n]+)\*(?!\*)', r'<i>\1</i>', text)

        # ~~зачёркнутый~~
        text = re.sub(r'~~(.+?)~~', r'<s>\1</s>', text)

        # [текст](ссылка) – синие ссылки
        text = re.sub(r'\[([^\]]+)\]\(([^)]+)\)', r'<a href="\2" style="color: #0000FF;">\1</a>', text)

        return text

    i = 0
    while i < len(lines):
        line = lines[i]

        # Многострочный код
        if line.strip().startswith("```"):
            if not in_code_block:
                flush_paragraph()
                in_code_block = True
                code_lines = []
            else:
                # Завершаем код
                result.append("<pre><code>" + "\n".join(code_lines) + "</code></pre>")
                in_code_block = False
                code_lines = []
            i += 1
            continue

        if in_code_block:
            code_lines.append(line)
            i += 1
            continue

        stripped = line.strip()

        # Горизонтальная линия
        if re.match(r'^[-*_]{3,}$', stripped):
            flush_paragraph()
            result.append("<hr>")
            i += 1
            continue

        # Заголовки
        if stripped.startswith("#### "):
            flush_paragraph()
            result.append("<h4>" + process_inline(stripped[5:]) + "</h4>")
            i += 1
            continue
        if stripped.startswith("### "):
            flush_paragraph()
            result.append("<h3>" + process_inline(stripped[4:]) + "</h3>")
            i += 1
            continue
        if stripped.startswith("## "):
            flush_paragraph()
            result.append("<h2>" + process_inline(stripped[3:]) + "</h2>")
            i += 1
            continue
        if stripped.startswith("# "):
            flush_paragraph()
            result.append("<h1>" + process_inline(stripped[2:]) + "</h1>")
            i += 1
            continue

        # Цитата
        if stripped.startswith("> "):
            flush_paragraph()
            result.append("<blockquote>" + process_inline(stripped[2:]) + "</blockquote>")
            i += 1
            continue

        # Ненумерованный список
        if re.match(r'^[\-\*\+]\s', stripped):
            flush_paragraph()
            list_items = []
            while i < len(lines) and re.match(r'^[\-\*\+]\s', lines[i].strip()):
                list_items.append("<li>" + process_inline(re.sub(r'^[\-\*\+]\s+', '', lines[i].strip())) + "</li>")
                i += 1
            result.append("<ul>" + "".join(list_items) + "</ul>")
            continue

        # Нумерованный список
        if re.match(r'^\d+\.\s', stripped):
            flush_paragraph()
            list_items = []
            while i < len(lines) and re.match(r'^\d+\.\s', lines[i].strip()):
                list_items.append("<li>" + process_inline(re.sub(r'^\d+\.\s+', '', lines[i].strip())) + "</li>")
                i += 1
            result.append("<ol>" + "".join(list_items) + "</ol>")
            continue

        # Пустая строка — завершаем параграф
        if not stripped:
            flush_paragraph()
            i += 1
            continue

        # Обычный текст — накапливаем в параграф
        in_paragraph = True
        paragraph_lines.append(process_inline(stripped))
        i += 1

    flush_paragraph()

    return "\n".join(result)


def strip_html_tags(text):
    text = re.sub(r"<[^>]+>", " ", str(text))
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def split_non_empty_lines(text):
    return [line.strip() for line in str(text).splitlines() if line.strip()]


def split_titles(text):
    parts = re.split(r"[,\n;]+", str(text))
    return [part.strip() for part in parts if part.strip()]

def short_button_text(text, limit=48):
    text = str(text).strip()
    return text if len(text) <= limit else text[:limit - 1].rstrip() + "…"

def instruction_display_label(instruction):
    """
    Красивое отображаемое имя инструкции:
    Раздел / Подзадача — Название инструкции
    """
    category = str(instruction.get("category_name", "")).strip()
    task = str(instruction.get("task_title", "")).strip()
    title = str(instruction.get("instruction_title") or instruction.get("title") or "").strip()

    prefix = " / ".join(part for part in (category, task) if part)
    return f"{prefix} — {title}" if prefix else title

def new_section_id():
    """Стабильный идентификатор секции инструкции для привязок чек-листа."""
    return f"sec_{uuid.uuid4().hex}"

def normalize_instruction_sections(raw):
    """Добавляет секциям инструкции стабильные id и сохраняет старые поля."""
    if not raw:
        return []

    if isinstance(raw, str):
        try:
            raw = json.loads(raw)
        except Exception:
            return []

    if not isinstance(raw, list):
        return []

    sections = []
    used_ids = set()
    for entry in raw:
        if not isinstance(entry, dict):
            continue

        section_data = dict(entry)
        section_id = str(section_data.get("id") or section_data.get("section_id") or "").strip()
        if not section_id or section_id in used_ids:
            section_id = new_section_id()

        section_data["id"] = section_id
        used_ids.add(section_id)
        sections.append(section_data)

    return sections

def section_index_by_id(instruction, section_id):
    if not instruction or not section_id:
        return None

    needle = str(section_id).strip()
    for idx, section_data in enumerate(instruction.get("sections", [])):
        if str(section_data.get("id") or "").strip() == needle:
            return idx
    return None

def normalize_checklist_sections(raw):
    """
    Приводит чек-лист к единому виду:
    [
        {
            "title": "Подзадача 1",
            "items": [
                {
                    "text": "Пункт",
                    "target_task_id": 12,
                    "target_instruction_id": 1,
                    "target_section_id": "sec_...",
                    "task_title": "Название подзадачи",
                    "instruction_title": "Название инструкции",
                    "section_title": "Название секции"
                }
            ]
        }
    ]

    Поддерживает старый формат:
    - список строк
    - список секций со строками
    """
    if not raw:
        return []

    if isinstance(raw, str):
        try:
            raw = json.loads(raw)
        except Exception:
            return []

    if not isinstance(raw, list):
        return []

    # Старый формат: просто список строк
    if raw and all(isinstance(item, str) for item in raw):
        items = []
        for item in raw:
            text = str(item).strip()
            if text:
                items.append({
                    "text": text,
                    "target_task_id": None,
                    "target_instruction_id": None,
                    "target_section_id": "",
                    "task_title": "",
                    "instruction_title": "",
                    "section_title": ""
                })
        return [{"title": "Чек-лист", "items": items}] if items else []

    sections = []
    for entry in raw:
        if not isinstance(entry, dict):
            continue

        section_title = str(entry.get("title") or entry.get("section_title") or "").strip()
        items_source = entry.get("items") or entry.get("checklist") or []

        items = []
        for item in items_source:
            if isinstance(item, str):
                text = item.strip()
                if text:
                    items.append({
                        "text": text,
                        "target_task_id": None,
                        "target_instruction_id": None,
                        "target_section_id": "",
                        "task_title": "",
                        "instruction_title": "",
                        "section_title": ""
                    })
                continue

            if not isinstance(item, dict):
                continue

            text = str(item.get("text") or item.get("title") or "").strip()
            if not text:
                continue

            items.append({
                "text": text,
                "target_task_id": item.get("target_task_id"),
                "target_instruction_id": item.get("target_instruction_id") or item.get("instruction_id"),
                "target_section_id": str(item.get("target_section_id") or item.get("section_id") or "").strip(),
                "task_title": str(item.get("task_title") or "").strip(),
                "instruction_title": str(item.get("instruction_title") or item.get("instruction") or "").strip(),
                "section_title": str(item.get("section_title") or item.get("anchor") or "").strip(),
            })

        if section_title or items:
            sections.append({
                "title": section_title or "Секция",
                "target_task_id": entry.get("target_task_id"),
                "target_instruction_id": entry.get("target_instruction_id") or entry.get("instruction_id"),
                "target_section_id": str(entry.get("target_section_id") or entry.get("section_id") or "").strip(),
                "task_title": str(entry.get("task_title") or "").strip(),
                "instruction_title": str(entry.get("instruction_title") or entry.get("instruction") or "").strip(),
                "section_title": str(entry.get("section_title") or entry.get("anchor") or "").strip(),
                "items": items
            })

    return sections


def checklist_item_overrides_section(item_data, section_data, current_instruction_id=None, current_task_id=None):
    """True, если у пункта своя привязка, отличная от привязки секции."""
    if not item_data:
        return False

    def normalize_instruction_id(instruction_id):
        try:
            instruction_id = int(instruction_id) if instruction_id else None
        except (TypeError, ValueError):
            instruction_id = None
        if instruction_id == current_instruction_id:
            return None
        return instruction_id

    def normalize_task_id(task_id):
        try:
            task_id = int(task_id) if task_id else None
        except (TypeError, ValueError):
            task_id = None
        if task_id == current_task_id:
            return None
        return task_id

    item_section_id = str(item_data.get("target_section_id") or "").strip()
    item_section_title = str(item_data.get("section_title") or "").strip()
    item_instruction_title = str(item_data.get("instruction_title") or "").strip()
    item_task_title = str(item_data.get("task_title") or "").strip()
    item_instruction_id = normalize_instruction_id(item_data.get("target_instruction_id"))
    item_task_id = normalize_task_id(item_data.get("target_task_id"))

    has_item_target = bool(
        item_section_id
        or item_section_title
        or item_instruction_title
        or item_task_title
        or item_instruction_id is not None
        or item_task_id is not None
    )
    if not has_item_target:
        return False

    if not section_data:
        return True

    section_section_id = str(section_data.get("target_section_id") or "").strip()
    section_section_title = str(section_data.get("section_title") or "").strip()
    section_instruction_title = str(section_data.get("instruction_title") or "").strip()
    section_task_title = str(section_data.get("task_title") or "").strip()
    section_instruction_id = normalize_instruction_id(section_data.get("target_instruction_id"))
    section_task_id = normalize_task_id(section_data.get("target_task_id"))

    return not (
            item_instruction_id == section_instruction_id
            and item_task_id == section_task_id
            and item_section_id == section_section_id
            and item_instruction_title == section_instruction_title
            and item_task_title == section_task_title
            and item_section_title == section_section_title
    )


def flatten_checklist_sections(sections):
    """Собирает все пункты чек-листа в один список."""
    result = []
    for section in normalize_checklist_sections(sections):
        result.extend(section.get("items", []))
    return result


def build_demo_checklist_sections(instruction_title):
    """Демо-чек-лист с секциями и переходами к секциям инструкции."""
    return [
        {
            "title": "Подготовка",
            "items": [
                {
                    "text": "Открыть инструкцию",
                    "instruction_title": instruction_title,
                    "section_title": "Коротко"
                },
                {
                    "text": "Проверить, что всё нужное под рукой",
                    "instruction_title": instruction_title,
                    "section_title": "Коротко"
                }
            ]
        },
        {
            "title": "Выполнение",
            "items": [
                {
                    "text": "Выполнить основные шаги",
                    "instruction_title": instruction_title,
                    "section_title": "Порядок действий"
                },
                {
                    "text": "Проверить важные замечания",
                    "instruction_title": instruction_title,
                    "section_title": "Важно"
                }
            ]
        },
        {
            "title": "Проверка",
            "items": [
                {
                    "text": "Сверить результат",
                    "instruction_title": instruction_title,
                    "section_title": "Важно"
                }
            ]
        }
    ]


class HtmlTreeDelegate(QStyledItemDelegate):
    """Делегат для отображения HTML в QTreeWidget — подсветка поиска"""

    def paint(self, painter, option, index):
        text = index.data(Qt.DisplayRole) or ""
        has_html = "<span" in text

        if has_html:
            painter.save()

            # Фон выделения
            if option.state & QStyle.StateFlag.State_Selected:
                painter.fillRect(option.rect, option.palette.highlight())
            elif option.state & QStyle.StateFlag.State_MouseOver:
                painter.fillRect(option.rect, option.palette.light())

            # Рендерим HTML
            doc = QTextDocument()
            doc.setDefaultFont(option.font)
            doc.setHtml(text)

            # Фиксированная ширина с отступом
            text_width = max(option.rect.width() - 6, 50)
            doc.setTextWidth(text_width)

            # Рисуем с вертикальным выравниванием по центру, как стандартный текст
            text_height = option.fontMetrics.height()
            doc_height = doc.size().height()

            # Центрируем по вертикали относительно стандартной высоты строки
            y_offset = option.rect.top() + max(0, (option.rect.height() - text_height) / 2)

            painter.translate(option.rect.left() + 3, y_offset)
            doc.drawContents(painter)
            painter.restore()
        else:
            super().paint(painter, option, index)

    def sizeHint(self, option, index):
        text = index.data(Qt.DisplayRole) or ""
        has_html = "<span" in text

        if has_html:
            # Возвращаем стандартную высоту строки, игнорируя высоту HTML
            base_size = super().sizeHint(option, index)
            return QSize(base_size.width(), option.fontMetrics.height() + 4)

        return super().sizeHint(option, index)

class WrappedChecklistDelegate(QStyledItemDelegate):
    def __init__(self, parent=None, padding=24):
        super().__init__(parent)
        self.padding = padding

    def initStyleOption(self, option, index):
        super().initStyleOption(option, index)
        option.textElideMode = Qt.ElideNone
        option.features |= QStyleOptionViewItem.ViewItemFeature.WrapText
        option.displayAlignment = Qt.AlignVCenter | Qt.AlignLeft

    def paint(self, painter, option, index):
        kind = index.data(ROLE_CHECKLIST_KIND)

        # Заголовок секции рисуем как скруглённую кнопку со стрелкой
        if kind == "section":
            painter.save()
            painter.setRenderHint(QPainter.RenderHint.Antialiasing, True)

            # Полностью очищаем строку перед рисованием скруглённой кнопки,
            # иначе при смене состояния Qt может оставлять старый текст в полях.
            if option.state & QStyle.StateFlag.State_Selected:
                painter.fillRect(option.rect, option.palette.highlight())
            else:
                painter.fillRect(option.rect, option.palette.base())

            rect = option.rect.adjusted(4, 3, -4, -3)
            painter.setPen(QPen(QColor("#d9e3f0"), 1))
            painter.setBrush(QColor("#eef5fb"))
            painter.drawRoundedRect(rect, 8, 8)

            font = option.font
            font.setBold(True)
            painter.setFont(font)
            painter.setPen(QColor("#263238"))

            collapsed = bool(index.data(ROLE_CHECKLIST_SECTION_COLLAPSED))
            arrow = ">" if collapsed else "v"
            text = index.data(Qt.DisplayRole) or ""
            painter.drawText(rect.adjusted(10, 0, -10, 0), Qt.AlignVCenter | Qt.AlignLeft, f"{arrow}  {text}")

            painter.restore()
            return

        # Обычные пункты чек-листа рисуем стандартно
        super().paint(painter, option, index)

    def sizeHint(self, option, index):
        kind = index.data(ROLE_CHECKLIST_KIND)

        opt = QStyleOptionViewItem(option)
        self.initStyleOption(opt, index)

        view = self.parent()
        available_width = 220
        if view is not None and hasattr(view, "viewport"):
            available_width = max(120, view.viewport().width() - self.padding)

        text = index.data(Qt.DisplayRole) or ""
        text_rect = opt.fontMetrics.boundingRect(
            0, 0, available_width, 0, Qt.TextWordWrap, text
        )
        base = super().sizeHint(option, index)

        if kind == "section":
            return QSize(base.width(), max(base.height(), text_rect.height() + 18))

        return QSize(base.width(), max(base.height(), text_rect.height() + 12))


def parse_manual_sections(text):
    raw = str(text).strip()
    if not raw:
        return []

    blocks = re.split(r"\n\s*\n", raw)
    result = []

    for block in blocks:
        lines = [line.rstrip() for line in block.splitlines() if line.strip()]
        if not lines:
            continue

        title = lines[0].strip()
        body = "\n".join(lines[1:]).strip()
        body_html = escape_html(body) if body else "<p></p>"
        result.append(section(title, body_html))

    return result


def section(title, body_html, image_path=None):
    return {
        "id": new_section_id(),
        "title": title,
        "body": body_html,
        "image_path": image_path or ""
    }


def make_sections(intro, steps, note_html):
    return [
        section("Коротко", p(intro)),
        section("Порядок действий", html_list(steps)),
        section("Важно", note_html),
    ]

def make_card(category, task_title, instruction_title, short_desc, checklist_sections, steps, note_html, related):
    checklist_flat = [
        item["text"]
        for section in checklist_sections
        for item in section.get("items", [])
    ]

    return {
        "category": category,
        "task_title": task_title,
        "instruction_title": instruction_title,
        "short_desc": short_desc,
        "checklist": checklist_flat,
        "checklist_sections": checklist_sections,
        "sections": make_sections(short_desc, steps, note_html),
        "related": related,
    }

# ================== ДЕМО-ДАННЫЕ ==================
# TODO: сюда потом можно вставить твои реальные 12 инструкций из Word.

def template_card(category, task_title, instruction_title, related=None):
    return make_card(
        category=category,
        task_title=task_title,
        instruction_title=instruction_title,
        short_desc=f"ВСТАВЬ СЮДА КРАТКОЕ ОПИСАНИЕ ИЗ WORD ДЛЯ «{instruction_title}».",
        checklist_sections=build_demo_checklist_sections(instruction_title),
        steps=[
            "ВСТАВЬ СЮДА ОСНОВНОЙ ТЕКСТ / ПЕРВУЮ ГЛАВУ ИЗ WORD.",
            "ВСТАВЬ СЮДА СЛЕДУЮЩИЙ РАЗДЕЛ ИНСТРУКЦИИ.",
            "ВСТАВЬ СЮДА ЗАКЛЮЧЕНИЕ, ПРИМЕЧАНИЯ И ОСОБЫЕ СЛУЧАИ.",
        ],
        note_html=f"<p><b>Важно:</b> сюда можно вставить ссылки на связанные инструкции и предупреждения для «{instruction_title}».</p>",
        related=related or [],
    )


SAMPLE_CARDS = [
    template_card(
        category="PDF и документы",
        task_title="Извлечь содержимое PDF-файлов",
        instruction_title="Инструкция по работе с программой для извлечения содержимого PDF-файлов - PDF Contents Extractor",
        related=[
            "Инструкция по подготовке документов для выгрузки в WindChill",
            "Инструкция по разбору документов для выгрузки в WindChill",
            "Инструкция по работе с программой split_spreads_ui по разрезанию разворотов в книгах",
        ],
    ),
    template_card(
        category="WindChill",
        task_title="Загрузить документы в WindChill",
        instruction_title="Инструкция по загрузке документов в WindChill",
        related=[
            "Инструкция по подготовке документов для выгрузки в WindChill",
            "Инструкция по разбору документов для выгрузки в WindChill",
        ],
    ),
    template_card(
        category="Утилизация",
        task_title="Работа со шредером",
        instruction_title="Инструкция для шредера",
        related=[
            "Инструкция по обработке макулатуры",
        ],
    ),
    template_card(
        category="Утилизация",
        task_title="Обработать макулатуру",
        instruction_title="Инструкция по обработке макулатуры",
        related=[
            "Инструкция для шредера",
        ],
    ),
    template_card(
        category="WindChill",
        task_title="Подготовить документы для выгрузки в WindChill",
        instruction_title="Инструкция по подготовке документов для выгрузки в WindChill",
        related=[
            "Инструкция по разбору документов для выгрузки в WindChill",
            "Инструкция по загрузке документов в WindChill",
        ],
    ),
    template_card(
        category="WindChill",
        task_title="Разобрать документы для выгрузки в WindChill",
        instruction_title="Инструкция по разбору документов для выгрузки в WindChill",
        related=[
            "Инструкция по подготовке документов для выгрузки в WindChill",
            "Инструкция по загрузке документов в WindChill",
        ],
    ),
    template_card(
        category="PDF и документы",
        task_title="Разрезать развороты в книгах",
        instruction_title="Инструкция по работе с программой split_spreads_ui по разрезанию разворотов в книгах",
        related=[
            "Инструкция по работе с программой для извлечения содержимого PDF-файлов - PDF Contents Extractor",
        ],
    ),
    template_card(
        category="Журналы",
        task_title="Смотреть прогресс по журналам",
        instruction_title="Инструкция по работе с приложением прогресса",
        related=[
            "Инструкция по загрузке статей в научную библиотеку УТЗ в WNC",
            "Инструкция по разделению закладок на отдельные файлы",
            "Инструкция по скачиванию статей с сайта Elibrary",
            "Инструкция по созданию закладок в PDF файле для последующего разъединения статей на отдельные файлы",
        ],
    ),
    template_card(
        category="Журналы",
        task_title="Загрузить статьи в научную библиотеку УТЗ",
        instruction_title="Инструкция по загрузке статей в научную библиотеку УТЗ в WNC",
        related=[
            "Инструкция по скачиванию статей с сайта Elibrary",
            "Инструкция по созданию закладок в PDF файле для последующего разъединения статей на отдельные файлы",
            "Инструкция по разделению закладок на отдельные файлы",
        ],
    ),
    template_card(
        category="Журналы",
        task_title="Разделить закладки на отдельные файлы",
        instruction_title="Инструкция по разделению закладок на отдельные файлы",
        related=[
            "Инструкция по созданию закладок в PDF файле для последующего разъединения статей на отдельные файлы",
            "Инструкция по загрузке статей в научную библиотеку УТЗ в WNC",
        ],
    ),
    template_card(
        category="Журналы",
        task_title="Скачать статьи с eLibrary",
        instruction_title="Инструкция по скачиванию статей с сайта Elibrary",
        related=[
            "Инструкция по созданию закладок в PDF файле для последующего разъединения статей на отдельные файлы",
        ],
    ),
    template_card(
        category="Журналы",
        task_title="Создать закладки в PDF",
        instruction_title="Инструкция по созданию закладок в PDF файле для последующего разъединения статей на отдельные файлы",
        related=[
            "Инструкция по разделению закладок на отдельные файлы",
            "Инструкция по скачиванию статей с сайта Elibrary",
            "Инструкция по загрузке статей в научную библиотеку УТЗ в WNC",
        ],
    ),
]


# ================== БАЗА ДАННЫХ ==================

class KnowledgeBaseDB:
    CHANGE_LOG_TABLES = (
        "categories",
        "instructions",
        "tasks",
        "comments",
        "ratings",
        "settings",
    )

    def __init__(self, path):
        self.path = Path(path)
        cleanup_sqlite_sidecar_files(self.path)
        backup_database_file(self.path)
        self._open_connection()
        self.init_schema()
        self.migrate_categories_schema()
        self.migrate_instructions_schema()
        self.migrate_tasks_schema()
        self.repair_category_foreign_keys_schema()
        self.migrate_ratings_schema()
        self.migrate_ratings_anchor_schema()
        self.migrate_comments_rating_schema()
        self.repair_instruction_reference_foreign_keys()
        self.migrate_change_log_schema()
        self.seed_demo_data()
        self.migrate_json_ids()

    def _open_connection(self):
        ensure_db_dir(self.path)
        self.conn = sqlite3.connect(str(self.path), timeout=10)
        self.conn.row_factory = sqlite3.Row
        self.conn.execute("PRAGMA busy_timeout = 5000")
        self.conn.execute("PRAGMA journal_mode = DELETE")
        self.conn.execute("PRAGMA synchronous = NORMAL")
        self.conn.execute("PRAGMA foreign_keys = ON")

    def reconnect(self):
        self.close()
        cleanup_sqlite_sidecar_files(self.path)
        self._open_connection()

    def get_file_signature(self):
        try:
            stat = self.path.stat()
            return (stat.st_mtime_ns, stat.st_size)
        except OSError:
            return None

    def get_change_signature(self):
        row = self.conn.execute(
            "SELECT COALESCE(MAX(id), 0) FROM change_log"
        ).fetchone()
        return int(row[0]) if row else 0

    def close(self):
        try:
            self.conn.close()
        except Exception:
            pass

    def init_schema(self):
        cur = self.conn.cursor()

        cur.execute("""
            CREATE TABLE IF NOT EXISTS categories (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL UNIQUE,
                parent_id INTEGER DEFAULT NULL,
                sort_order INTEGER NOT NULL DEFAULT 0,
                FOREIGN KEY(parent_id) REFERENCES categories(id) ON DELETE CASCADE
            )
        """)

        cur.execute("""
            CREATE TABLE IF NOT EXISTS instructions (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                category_id INTEGER NOT NULL,
                title TEXT NOT NULL,
                short_desc TEXT NOT NULL DEFAULT '',
                sections_json TEXT NOT NULL DEFAULT '[]',
                related_ids_json TEXT NOT NULL DEFAULT '[]',
                updated_at TEXT NOT NULL,
                FOREIGN KEY(category_id) REFERENCES categories(id) ON DELETE CASCADE
            )
        """)

        cur.execute("""
            CREATE TABLE IF NOT EXISTS tasks (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                category_id INTEGER NOT NULL,
                title TEXT NOT NULL,
                instruction_id INTEGER,
                checklist_json TEXT NOT NULL DEFAULT '[]',
                sort_order INTEGER NOT NULL DEFAULT 0,
                FOREIGN KEY(category_id) REFERENCES categories(id) ON DELETE CASCADE,
                FOREIGN KEY(instruction_id) REFERENCES instructions(id) ON DELETE SET NULL,
                UNIQUE(category_id, title)
            )
        """)

        cur.execute("""
            CREATE TABLE IF NOT EXISTS comments (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                instruction_id INTEGER NOT NULL,
                anchor TEXT NOT NULL DEFAULT '',
                author TEXT NOT NULL,
                is_anonymous INTEGER NOT NULL DEFAULT 1,
                text TEXT NOT NULL,
                created_at TEXT NOT NULL,
                FOREIGN KEY(instruction_id) REFERENCES instructions(id) ON DELETE CASCADE
            )
        """)

        cur.execute("""
            CREATE TABLE IF NOT EXISTS ratings (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                instruction_id INTEGER NOT NULL,
                rating INTEGER NOT NULL CHECK(rating BETWEEN 1 AND 10),
                created_at TEXT NOT NULL,
                FOREIGN KEY(instruction_id) REFERENCES instructions(id) ON DELETE CASCADE
            )
        """)
        self.conn.commit()

    def migrate_change_log_schema(self):
        cur = self.conn.cursor()
        cur.execute("""
            CREATE TABLE IF NOT EXISTS change_log (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                table_name TEXT NOT NULL,
                row_id INTEGER,
                action TEXT NOT NULL,
                created_at TEXT NOT NULL
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS app_meta (
                key TEXT PRIMARY KEY,
                value TEXT NOT NULL
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS settings (
                key TEXT PRIMARY KEY,
                value TEXT NOT NULL DEFAULT ''
            )
        """)

        for table_name in self.CHANGE_LOG_TABLES:
            for action in ("ai", "au", "ad"):
                trigger_name = f"trg_{table_name}_{action}"
                row_ref = "NEW" if action != "ad" else "OLD"
                row_id_expr = f"{row_ref}.id" if table_name != "settings" else "NULL"
                cur.execute(f"""
                    CREATE TRIGGER IF NOT EXISTS {trigger_name}
                    AFTER {'INSERT' if action == 'ai' else 'UPDATE' if action == 'au' else 'DELETE'} ON {table_name}
                    BEGIN
                        INSERT INTO change_log(table_name, row_id, action, created_at)
                        VALUES (
                            '{table_name}',
                            {row_id_expr},
                            '{"I" if action == "ai" else "U" if action == "au" else "D"}',
                            datetime('now')
                        );
                    END;
                """)

        cur.execute("CREATE INDEX IF NOT EXISTS idx_change_log_id ON change_log(id)")
        cur.execute(
            "INSERT OR IGNORE INTO app_meta (key, value) VALUES ('db_version', ?)",
            (datetime.now().isoformat(),),
        )
        self.conn.commit()

    def migrate_categories_schema(self):
        rows = self.conn.execute("PRAGMA table_info(categories)").fetchall()
        columns = {row["name"] for row in rows}

        if "parent_id" not in columns:
            self.conn.execute("""
                ALTER TABLE categories
                ADD COLUMN parent_id INTEGER DEFAULT NULL
            """)
            self.conn.commit()

    def migrate_instructions_schema(self):
        row = self.conn.execute("""
            SELECT sql
            FROM sqlite_master
            WHERE type='table' AND name='instructions'
        """).fetchone()

        sql = row["sql"] if row and row["sql"] else ""
        if "title TEXT NOT NULL UNIQUE" not in sql and "UNIQUE(title)" not in sql:
            return

        cur = self.conn.cursor()
        cur.execute("PRAGMA foreign_keys = OFF")
        cur.execute("ALTER TABLE instructions RENAME TO instructions_old")
        cur.execute("""
            CREATE TABLE instructions (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                category_id INTEGER NOT NULL,
                title TEXT NOT NULL,
                short_desc TEXT NOT NULL DEFAULT '',
                sections_json TEXT NOT NULL DEFAULT '[]',
                related_ids_json TEXT NOT NULL DEFAULT '[]',
                updated_at TEXT NOT NULL,
                FOREIGN KEY(category_id) REFERENCES categories(id) ON DELETE CASCADE
            )
        """)
        cur.execute("""
            INSERT INTO instructions (
                id, category_id, title, short_desc,
                sections_json, related_ids_json, updated_at
            )
            SELECT
                id, category_id, title, short_desc,
                sections_json, related_ids_json, updated_at
            FROM instructions_old
        """)
        cur.execute("DROP TABLE instructions_old")
        cur.execute("PRAGMA foreign_keys = ON")
        self.conn.commit()

    def repair_category_foreign_keys_schema(self):
        self._rebuild_table_if_needed(
            table_name="instructions",
            expected_fk_marker="categories_old",
            create_sql="""
                CREATE TABLE instructions (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    category_id INTEGER NOT NULL,
                    title TEXT NOT NULL,
                    short_desc TEXT NOT NULL DEFAULT '',
                    sections_json TEXT NOT NULL DEFAULT '[]',
                    related_ids_json TEXT NOT NULL DEFAULT '[]',
                    updated_at TEXT NOT NULL,
                    FOREIGN KEY(category_id) REFERENCES categories(id) ON DELETE CASCADE
                )
            """,
            columns="id, category_id, title, short_desc, sections_json, related_ids_json, updated_at"
        )

        self._rebuild_table_if_needed(
            table_name="tasks",
            expected_fk_marker="categories_old",
            create_sql="""
                CREATE TABLE tasks (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    category_id INTEGER NOT NULL,
                    title TEXT NOT NULL UNIQUE,
                    instruction_id INTEGER,
                    checklist_json TEXT NOT NULL DEFAULT '[]',
                    sort_order INTEGER NOT NULL DEFAULT 0,
                    FOREIGN KEY(category_id) REFERENCES categories(id) ON DELETE CASCADE,
                    FOREIGN KEY(instruction_id) REFERENCES instructions(id) ON DELETE SET NULL
                )
            """,
            columns="id, category_id, title, instruction_id, checklist_json, sort_order"
        )

    def _rebuild_table_if_needed(self, table_name, expected_fk_marker, create_sql, columns):
        row = self.conn.execute("""
            SELECT sql
            FROM sqlite_master
            WHERE type='table' AND name=?
        """, (table_name,)).fetchone()

        sql = row["sql"] if row and row["sql"] else ""
        if expected_fk_marker not in sql:
            return

        cur = self.conn.cursor()
        cur.execute("PRAGMA foreign_keys = OFF")
        cur.execute(f"ALTER TABLE {table_name} RENAME TO {table_name}_old")
        cur.execute(create_sql)
        cur.execute(f"""
            INSERT INTO {table_name} ({columns})
            SELECT {columns}
            FROM {table_name}_old
        """)
        cur.execute(f"DROP TABLE {table_name}_old")
        cur.execute("PRAGMA foreign_keys = ON")
        self.conn.commit()

    def migrate_tasks_schema(self):
        row = self.conn.execute("""
            SELECT sql
            FROM sqlite_master
            WHERE type='table' AND name='tasks'
        """).fetchone()

        sql = row["sql"] if row and row["sql"] else ""
        if "UNIQUE(category_id, title)" in sql:
            return

        cur = self.conn.cursor()
        cur.execute("PRAGMA foreign_keys = OFF")
        cur.execute("ALTER TABLE tasks RENAME TO tasks_old")
        cur.execute("""
            CREATE TABLE tasks (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                category_id INTEGER NOT NULL,
                title TEXT NOT NULL,
                instruction_id INTEGER,
                checklist_json TEXT NOT NULL DEFAULT '[]',
                sort_order INTEGER NOT NULL DEFAULT 0,
                FOREIGN KEY(category_id) REFERENCES categories(id) ON DELETE CASCADE,
                FOREIGN KEY(instruction_id) REFERENCES instructions(id) ON DELETE SET NULL,
                UNIQUE(category_id, title)
            )
        """)
        cur.execute("""
            INSERT INTO tasks (id, category_id, title, instruction_id, checklist_json, sort_order)
            SELECT id, category_id, title, instruction_id, checklist_json, sort_order
            FROM tasks_old
        """)
        cur.execute("DROP TABLE tasks_old")
        cur.execute("PRAGMA foreign_keys = ON")
        self.conn.commit()

    def migrate_ratings_schema(self):
        row = self.conn.execute("""
            SELECT sql
            FROM sqlite_master
            WHERE type='table' AND name='ratings'
        """).fetchone()

        sql = row["sql"] if row and row["sql"] else ""
        if "BETWEEN 1 AND 10" in sql:
            return

        cur = self.conn.cursor()
        cur.execute("PRAGMA foreign_keys = OFF")
        cur.execute("ALTER TABLE ratings RENAME TO ratings_old")
        cur.execute("""
            CREATE TABLE ratings (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                instruction_id INTEGER NOT NULL,
                rating INTEGER NOT NULL CHECK(rating BETWEEN 1 AND 10),
                created_at TEXT NOT NULL,
                FOREIGN KEY(instruction_id) REFERENCES instructions(id) ON DELETE CASCADE
            )
        """)
        cur.execute("""
            INSERT INTO ratings (id, instruction_id, rating, created_at)
            SELECT id, instruction_id, rating, created_at
            FROM ratings_old
        """)
        cur.execute("DROP TABLE ratings_old")
        cur.execute("PRAGMA foreign_keys = ON")
        self.conn.commit()

    def migrate_ratings_anchor_schema(self):
        rows = self.conn.execute("PRAGMA table_info(ratings)").fetchall()
        columns = {row["name"] for row in rows}
        if "anchor" not in columns:
            self.conn.execute(
                "ALTER TABLE ratings ADD COLUMN anchor TEXT NOT NULL DEFAULT ''"
            )
            self.conn.commit()

    def migrate_comments_rating_schema(self):
        rows = self.conn.execute("PRAGMA table_info(comments)").fetchall()
        columns = {row["name"] for row in rows}
        if "rating" not in columns:
            self.conn.execute(
                "ALTER TABLE comments ADD COLUMN rating INTEGER"
            )
            self.conn.commit()
            self._backfill_comment_ratings()

    def _backfill_comment_ratings(self):
        """Связывает старые комментарии с оценками по блоку и времени."""
        rows = self.conn.execute("""
            SELECT c.id, c.instruction_id, c.anchor, c.created_at
            FROM comments c
            WHERE c.rating IS NULL
        """).fetchall()
        for row in rows:
            rating_row = self.conn.execute("""
                SELECT rating
                FROM ratings
                WHERE instruction_id=?
                  AND anchor=?
                  AND created_at=?
                ORDER BY id DESC
                LIMIT 1
            """, (row["instruction_id"], row["anchor"] or "", row["created_at"])).fetchone()
            if rating_row:
                self.conn.execute(
                    "UPDATE comments SET rating=? WHERE id=?",
                    (int(rating_row["rating"]), row["id"]),
                )
        self.conn.commit()

    def repair_instruction_reference_foreign_keys(self):
        """Чинит FK на instructions после прерванных миграций (instructions_old)."""
        repairs = (
            (
                "ratings",
                """
                CREATE TABLE ratings_new (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    instruction_id INTEGER NOT NULL,
                    rating INTEGER NOT NULL CHECK(rating BETWEEN 1 AND 10),
                    anchor TEXT NOT NULL DEFAULT '',
                    created_at TEXT NOT NULL,
                    FOREIGN KEY(instruction_id) REFERENCES instructions(id) ON DELETE CASCADE
                )
                """,
                lambda cols: f"""
                    INSERT INTO ratings_new (id, instruction_id, rating, anchor, created_at)
                    SELECT id, instruction_id, rating,
                           {("anchor" if "anchor" in cols else "''")},
                           created_at
                    FROM ratings
                """,
            ),
            (
                "comments",
                """
                CREATE TABLE comments_new (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    instruction_id INTEGER NOT NULL,
                    anchor TEXT NOT NULL DEFAULT '',
                    author TEXT NOT NULL,
                    is_anonymous INTEGER NOT NULL DEFAULT 1,
                    text TEXT NOT NULL,
                    rating INTEGER,
                    created_at TEXT NOT NULL,
                    FOREIGN KEY(instruction_id) REFERENCES instructions(id) ON DELETE CASCADE
                )
                """,
                lambda cols: f"""
                    INSERT INTO comments_new (
                        id, instruction_id, anchor, author, is_anonymous, text, rating, created_at
                    )
                    SELECT id, instruction_id, anchor, author, is_anonymous, text,
                           {("rating" if "rating" in cols else "NULL")},
                           created_at
                    FROM comments
                """,
            ),
        )

        for table, create_sql, copy_sql_builder in repairs:
            fk_rows = self.conn.execute(f"PRAGMA foreign_key_list({table})").fetchall()
            broken = any(
                row["from"] == "instruction_id" and row["table"] != "instructions"
                for row in fk_rows
            )
            if not broken:
                continue

            columns = {row["name"] for row in self.conn.execute(f"PRAGMA table_info({table})").fetchall()}
            cur = self.conn.cursor()
            cur.execute("PRAGMA foreign_keys = OFF")
            cur.execute(create_sql)
            cur.execute(copy_sql_builder(columns))
            cur.execute(f"DROP TABLE {table}")
            cur.execute(f"ALTER TABLE {table}_new RENAME TO {table}")
            cur.execute("PRAGMA foreign_keys = ON")
            self.conn.commit()

    def seed_demo_data(self):
        cur = self.conn.cursor()
        cur.execute("SELECT COUNT(*) AS cnt FROM instructions")
        if cur.fetchone()["cnt"] > 0:
            return

        categories = list(dict.fromkeys(card["category"] for card in SAMPLE_CARDS))
        category_ids = {}

        for idx, category_name in enumerate(categories, start=1):
            cur.execute(
                "INSERT INTO categories (name, sort_order) VALUES (?, ?)",
                (category_name, idx)
            )
            category_ids[category_name] = cur.lastrowid

        instruction_ids = {}

        for card in SAMPLE_CARDS:
            cur.execute("""
                INSERT INTO instructions
                    (category_id, title, short_desc, sections_json, related_ids_json, updated_at)
                VALUES (?, ?, ?, ?, ?, ?)
            """, (
                category_ids[card["category"]],
                card["instruction_title"],
                card["short_desc"],
                json.dumps(card["sections"], ensure_ascii=False),
                json.dumps([], ensure_ascii=False),
                ts(),
            ))
            instruction_ids[card["instruction_title"]] = cur.lastrowid

        # Обновляем связи между инструкциями
        for card in SAMPLE_CARDS:
            related_ids = [
                instruction_ids[title]
                for title in card["related"]
                if title in instruction_ids
            ]
            cur.execute(
                "UPDATE instructions SET related_ids_json=?, updated_at=? WHERE title=?",
                (json.dumps(related_ids, ensure_ascii=False), ts(), card["instruction_title"])
            )

        # Задачи
        for idx, card in enumerate(SAMPLE_CARDS, start=1):
            cur.execute("""
                INSERT INTO tasks
                    (category_id, title, instruction_id, checklist_json, sort_order)
                VALUES (?, ?, ?, ?, ?)
            """, (
                category_ids[card["category"]],
                card["task_title"],
                instruction_ids[card["instruction_title"]],
                json.dumps(card.get("checklist_sections") or normalize_checklist_sections(card.get("checklist", [])), ensure_ascii=False),
                idx,
            ))

        self.conn.commit()

    def migrate_json_ids(self):
        """Проставляет стабильные id в JSON-секциях и старых целях чек-листа."""
        instruction_rows = self.conn.execute("""
            SELECT id, title, sections_json
            FROM instructions
        """).fetchall()

        instruction_titles = {}
        instruction_sections = {}
        instruction_titles_by_id = {}

        for row in instruction_rows:
            instruction_id = row["id"]
            instruction_titles[str(row["title"] or "").casefold().strip()] = instruction_id
            instruction_titles_by_id[instruction_id] = row["title"]

            old_sections = json.loads(row["sections_json"] or "[]")
            sections = normalize_instruction_sections(old_sections)
            instruction_sections[instruction_id] = sections

            if sections != old_sections:
                self.conn.execute(
                    "UPDATE instructions SET sections_json=?, updated_at=? WHERE id=?",
                    (json.dumps(sections, ensure_ascii=False), ts(), instruction_id)
                )

        task_rows = self.conn.execute("""
            SELECT id, instruction_id, checklist_json
            FROM tasks
        """).fetchall()

        instruction_tasks = {}
        for task_row in self.conn.execute("""
            SELECT id, instruction_id
            FROM tasks
            WHERE instruction_id IS NOT NULL
        """):
            instruction_tasks.setdefault(task_row["instruction_id"], []).append(task_row["id"])

        def backfill_target_task(entry, owner_task_id, owner_instruction_id):
            if entry.get("target_task_id"):
                try:
                    if int(entry["target_task_id"]):
                        return False
                except (TypeError, ValueError):
                    pass

            try:
                target_instruction_id = int(entry.get("target_instruction_id") or 0) or None
            except (TypeError, ValueError):
                target_instruction_id = None

            if not target_instruction_id or target_instruction_id == owner_instruction_id:
                return False

            linked_tasks = instruction_tasks.get(target_instruction_id, [])
            if len(linked_tasks) == 1:
                entry["target_task_id"] = linked_tasks[0]
                return True
            return False

        for row in task_rows:
            old_checklist = json.loads(row["checklist_json"] or "[]")
            checklist_sections = normalize_checklist_sections(old_checklist)
            changed = checklist_sections != old_checklist

            for checklist_section in checklist_sections:
                if backfill_target_task(checklist_section, row["id"], row["instruction_id"]):
                    changed = True

                for item in checklist_section.get("items", []):
                    target_instruction_id = item.get("target_instruction_id")
                    try:
                        target_instruction_id = int(target_instruction_id) if target_instruction_id else None
                    except (TypeError, ValueError):
                        target_instruction_id = None

                    target_title = str(item.get("instruction_title") or "").casefold().strip()
                    if target_instruction_id is None and target_title:
                        target_instruction_id = instruction_titles.get(target_title)

                    if target_instruction_id is None:
                        target_instruction_id = row["instruction_id"]

                    if item.get("target_instruction_id") != target_instruction_id:
                        item["target_instruction_id"] = target_instruction_id
                        changed = True

                    if target_instruction_id and not item.get("instruction_title") and target_instruction_id != row["instruction_id"]:
                        item["instruction_title"] = instruction_titles_by_id.get(target_instruction_id, "")
                        changed = True

                    if not item.get("target_section_id"):
                        target_section_title = str(item.get("section_title") or "").casefold().strip()
                        if target_section_title:
                            for section_data in instruction_sections.get(target_instruction_id, []):
                                if str(section_data.get("title") or "").casefold().strip() == target_section_title:
                                    item["target_section_id"] = section_data.get("id", "")
                                    changed = True
                                    break

                    if backfill_target_task(item, row["id"], row["instruction_id"]):
                        changed = True

            if changed:
                self.conn.execute(
                    "UPDATE tasks SET checklist_json=? WHERE id=?",
                    (json.dumps(checklist_sections, ensure_ascii=False), row["id"])
                )

        self.conn.commit()

    def all_instruction_titles(self):
        """Возвращает список названий всех инструкций."""
        rows = self.conn.execute("SELECT title FROM instructions ORDER BY title").fetchall()
        return [row["title"] for row in rows]

    def all_instructions(self):
        rows = self.conn.execute("""
            SELECT i.id AS instruction_id,
                   i.category_id,
                   c.name AS category_name,
                   i.title AS instruction_title,
                   (SELECT t.title
                    FROM tasks t
                    WHERE t.instruction_id = i.id
                    LIMIT 1) AS task_title
            FROM instructions i
            JOIN categories c ON c.id = i.category_id
            ORDER BY c.sort_order, i.title, i.id
        """).fetchall()
        return [dict(row) for row in rows]

    def all_instructions(self):
        rows = self.conn.execute("""
            SELECT i.id AS instruction_id,
                   i.title,
                   c.name AS category_name
            FROM instructions i
            JOIN categories c ON c.id = i.category_id
            ORDER BY c.sort_order, i.title, i.id
        """).fetchall()
        return [dict(row) for row in rows]

    def all_instruction_targets(self):
        rows = self.conn.execute("""
            SELECT i.id AS instruction_id,
                   i.category_id,
                   c.name AS category_name,
                   i.title AS instruction_title,
                   (SELECT t.title
                    FROM tasks t
                    WHERE t.instruction_id = i.id
                    LIMIT 1) AS task_title,
                   i.sections_json
            FROM instructions i
            JOIN categories c ON c.id = i.category_id
            ORDER BY c.sort_order, i.title, i.id
        """).fetchall()

        result = []
        for row in rows:
            data = dict(row)
            data["sections"] = normalize_instruction_sections(json.loads(data.pop("sections_json") or "[]"))
            result.append(data)
        return result

    def all_checklist_targets(self):
        """Все подзадачи с инструкциями — для привязки пунктов чек-листа."""
        rows = self.conn.execute("""
            SELECT t.id AS task_id,
                   t.title AS task_title,
                   t.category_id,
                   c.name AS category_name,
                   i.id AS instruction_id,
                   i.title AS instruction_title,
                   i.sections_json
            FROM tasks t
            JOIN categories c ON c.id = t.category_id
            JOIN instructions i ON i.id = t.instruction_id
            ORDER BY c.sort_order, t.sort_order, t.id
        """).fetchall()

        result = []
        for row in rows:
            data = dict(row)
            data["category_path"] = self.category_path(data["category_id"])
            data["sections"] = normalize_instruction_sections(json.loads(data.pop("sections_json") or "[]"))
            result.append(data)
        return result

    def category_by_id(self, category_id):
        row = self.conn.execute("""
            SELECT c.id, c.name, c.parent_id, c.sort_order,
                   COALESCE(COUNT(t.id), 0) AS task_count,
                   (SELECT COUNT(*) FROM categories ch WHERE ch.parent_id = c.id) AS child_count
            FROM categories c
            LEFT JOIN tasks t ON t.category_id = c.id
            WHERE c.id=?
            GROUP BY c.id, c.name, c.parent_id, c.sort_order
        """, (category_id,)).fetchone()
        return dict(row) if row else None

    def categories_by_parent(self, parent_id=None):
        if parent_id is None:
            rows = self.conn.execute("""
                SELECT id, name, parent_id, sort_order
                FROM categories
                WHERE parent_id IS NULL
                ORDER BY sort_order, name, id
            """).fetchall()
        else:
            rows = self.conn.execute("""
                SELECT id, name, parent_id, sort_order
                FROM categories
                WHERE parent_id=?
                ORDER BY sort_order, name, id
            """, (parent_id,)).fetchall()

        return [dict(row) for row in rows]

    def category_has_children(self, category_id):
        row = self.conn.execute("""
            SELECT COUNT(*) AS cnt
            FROM categories
            WHERE parent_id=?
        """, (category_id,)).fetchone()
        return int(row["cnt"]) > 0 if row else False

    def category_path(self, category_id):
        parts = []
        current_id = category_id

        while current_id is not None:
            row = self.conn.execute("""
                SELECT id, name, parent_id
                FROM categories
                WHERE id=?
            """, (current_id,)).fetchone()
            if not row:
                break
            parts.append(row["name"])
            current_id = row["parent_id"]

        return list(reversed(parts))

    def count_tasks_in_category(self, category_id):
        row = self.conn.execute(
            "SELECT COUNT(*) AS cnt FROM tasks WHERE category_id=?",
            (category_id,)
        ).fetchone()
        return int(row["cnt"]) if row else 0

    def tasks_for_category(self, category_id):
        rows = self.conn.execute("""
            SELECT id AS task_id, title AS task_title
            FROM tasks
            WHERE category_id=?
            ORDER BY sort_order, id
        """, (category_id,)).fetchall()
        return [dict(row) for row in rows]

    def instruction_by_id(self, instruction_id):
        row = self.conn.execute("""
            SELECT i.id AS instruction_id,
                   i.category_id,
                   c.name AS category_name,
                   i.title AS instruction_title,
                   (SELECT t.title
                    FROM tasks t
                    WHERE t.instruction_id = i.id
                    LIMIT 1) AS task_title,
                   i.short_desc,
                   i.sections_json,
                   i.related_ids_json,
                   i.updated_at
            FROM instructions i
            JOIN categories c ON c.id = i.category_id
            WHERE i.id=?
        """, (instruction_id,)).fetchone()

        if not row:
            return None

        data = dict(row)
        task_row = self.conn.execute("""
            SELECT id AS task_id, category_id AS task_category_id
            FROM tasks
            WHERE instruction_id=?
            LIMIT 1
        """, (instruction_id,)).fetchone()
        if task_row:
            data.update(dict(task_row))

        data["sections"] = normalize_instruction_sections(json.loads(data["sections_json"] or "[]"))
        related_ids = json.loads(data["related_ids_json"] or "[]")
        data["related_ids"] = related_ids
        data["related_titles"] = []

        for rid in related_ids:
            r = self.conn.execute(
                "SELECT title FROM instructions WHERE id=?",
                (rid,)
            ).fetchone()
            if r:
                data["related_titles"].append(r["title"])

        return data

    def instruction_by_title(self, title):
        row = self.conn.execute("""
            SELECT id
            FROM instructions
            WHERE LOWER(title)=LOWER(?)
            LIMIT 1
        """, (title,)).fetchone()
        if not row:
            return None
        return self.instruction_by_id(row["id"])

    def tasks_for_instruction(self, instruction_id):
        rows = self.conn.execute("""
            SELECT id AS task_id, category_id, title AS task_title
            FROM tasks
            WHERE instruction_id=?
            ORDER BY sort_order, id
        """, (instruction_id,)).fetchall()
        return [dict(row) for row in rows]

    def task_by_instruction_id(self, instruction_id):
        row = self.conn.execute("""
            SELECT id
            FROM tasks
            WHERE instruction_id=?
            LIMIT 1
        """, (instruction_id,)).fetchone()
        if not row:
            return None
        return self.task_bundle(row["id"])

    def _row_to_task(self, row):
        data = dict(row)
        checklist_raw = json.loads(data["checklist_json"] or "[]")
        checklist_sections = normalize_checklist_sections(checklist_raw)

        data["checklist_sections"] = checklist_sections
        data["checklist"] = [item["text"] for item in flatten_checklist_sections(checklist_sections)]
        data["sections"] = normalize_instruction_sections(json.loads(data.get("sections_json") or "[]"))
        related_ids = json.loads(data.get("related_ids_json") or "[]")
        data["related_ids"] = related_ids
        data["related_titles"] = []

        for rid in related_ids:
            rr = self.conn.execute("SELECT title FROM instructions WHERE id=?", (rid,)).fetchone()
            if rr:
                data["related_titles"].append(rr["title"])

        category_path = ""
        if data.get("category_id"):
            category_path = " / ".join(self.category_path(data["category_id"]))

        search_parts = [
            category_path,
            data.get("category_name") or "",
            data.get("task_title") or "",
            data.get("instruction_title") or "",
            data.get("short_desc") or "",
            " ".join(data.get("checklist", [])),
            " ".join(
                (section.get("title") or "") + " " + strip_html_tags(section.get("body") or "")
                for section in data.get("sections", [])
            ),
            " ".join(
                (sec.get("title") or "") + " " + " ".join(
                    item.get("text", "") for item in sec.get("items", [])
                )
                for sec in data.get("checklist_sections", [])
            ),
            " ".join((data.get("related_titles") or [])),
        ]
        data["search_blob"] = " ".join(search_parts).casefold()
        return data

    _TASK_SELECT_SQL = """
            SELECT t.id AS task_id,
                   t.title AS task_title,
                   t.instruction_id,
                   t.checklist_json,
                   t.sort_order,
                   c.id AS category_id,
                   c.name AS category_name,
                   c.sort_order AS category_sort_order,
                   i.id AS instruction_real_id,
                   i.title AS instruction_title,
                   i.short_desc,
                   i.sections_json,
                   i.related_ids_json
            FROM tasks t
            JOIN categories c ON c.id = t.category_id
            LEFT JOIN instructions i ON i.id = t.instruction_id
    """

    def all_tasks(self):
        rows = self.conn.execute(
            self._TASK_SELECT_SQL + " ORDER BY c.sort_order, t.sort_order, t.id"
        ).fetchall()

        return [self._row_to_task(row) for row in rows]

    def task_bundle(self, task_id):
        row = self.conn.execute(
            self._TASK_SELECT_SQL + " WHERE t.id=?",
            (task_id,)
        ).fetchone()
        if not row:
            return None
        return self._row_to_task(row)

    def search_tasks(self, search_text):
        tasks = self.all_tasks()
        if not search_text:
            return tasks

        needle = search_text.casefold().strip()
        return [task for task in tasks if needle in task["search_blob"]]

    def comments_for_instruction(self, instruction_id):
        rows = self.conn.execute("""
            SELECT id, instruction_id, anchor, author, is_anonymous, text, rating, created_at
            FROM comments
            WHERE instruction_id=?
            ORDER BY id DESC
        """, (instruction_id,)).fetchall()
        comments = [dict(row) for row in rows]
        return self.enrich_comments_with_ratings(comments, instruction_id)

    def enrich_comments_with_ratings(self, comments, instruction_id):
        """Дополняет комментарии оценкой из поля rating или таблицы ratings."""
        if not comments:
            return comments

        ratings = self.ratings_for_instruction(instruction_id)
        ratings_by_key = {
            ((r.get("anchor") or ""), (r.get("created_at") or "")): int(r["rating"])
            for r in ratings
        }

        for comment in comments:
            if comment.get("rating") is not None:
                continue
            key = ((comment.get("anchor") or ""), (comment.get("created_at") or ""))
            if key in ratings_by_key:
                comment["rating"] = ratings_by_key[key]

        return comments

    def add_comment(self, instruction_id, anchor, author, is_anonymous, text, rating=None, created_at=None):
        try:
            created_at = created_at or ts()
            rating_value = int(rating) if rating is not None else None
            self.conn.execute("""
                INSERT INTO comments
                    (instruction_id, anchor, author, is_anonymous, text, rating, created_at)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            """, (
                instruction_id,
                anchor or "",
                author,
                1 if is_anonymous else 0,
                text,
                rating_value,
                created_at,
            ))
            self.conn.commit()
            return True, None
        except Exception as e:
            return False, str(e)

    def comment_by_id(self, comment_id):
        row = self.conn.execute("""
            SELECT id, instruction_id, anchor, author, is_anonymous, text, rating, created_at
            FROM comments
            WHERE id=?
        """, (comment_id,)).fetchone()
        return dict(row) if row else None

    def update_comment(self, comment_id, text):
        self.conn.execute("""
            UPDATE comments
            SET text=?
            WHERE id=?
        """, (text, comment_id))
        self.conn.commit()

    def delete_comment(self, comment_id):
        self.conn.execute(
            "DELETE FROM comments WHERE id=?",
            (comment_id,)
        )
        self.conn.commit()

    def comments_export_rows(self):
        rows = self.conn.execute("""
            SELECT c.id,
                   cat.name AS category_name,
                   i.title AS instruction_title,
                   c.anchor,
                   c.author,
                   c.is_anonymous,
                   c.text,
                   c.created_at
            FROM comments c
            JOIN instructions i ON i.id = c.instruction_id
            JOIN categories cat ON cat.id = i.category_id
            ORDER BY cat.sort_order, i.title, c.id
        """).fetchall()
        return [dict(row) for row in rows]

    def ratings_export_rows(self):
        rows = self.conn.execute("""
            SELECT r.id,
                   cat.name AS category_name,
                   i.title AS instruction_title,
                   r.rating,
                   r.anchor,
                   r.created_at
            FROM ratings r
            JOIN instructions i ON i.id = r.instruction_id
            JOIN categories cat ON cat.id = i.category_id
            ORDER BY cat.sort_order, i.title, r.id
        """).fetchall()
        return [dict(row) for row in rows]

    def export_feedback_xlsx(self, path):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font
            from openpyxl.utils import get_column_letter
        except ImportError as exc:
            raise RuntimeError(
                "Для экспорта в Excel нужен пакет openpyxl. "
                "Установите его командой: pip install openpyxl"
            ) from exc

        wb = Workbook()

        comments_rows = self.comments_export_rows()
        ratings_rows = self.ratings_export_rows()

        summary_ws = wb.active
        summary_ws.title = "Сводка"
        summary_ws.append(["Показатель", "Значение"])
        summary_ws["A1"].font = Font(bold=True)
        summary_ws["B1"].font = Font(bold=True)

        total_avg = round(
            sum(row["rating"] for row in ratings_rows) / len(ratings_rows), 2
        ) if ratings_rows else 0.0

        summary_data = [
            ("Всего комментариев", len(comments_rows)),
            ("Всего оценок", len(ratings_rows)),
            ("Средняя оценка по всем", total_avg),
        ]
        for item in summary_data:
            summary_ws.append(list(item))

        comments_ws = wb.create_sheet("Комментарии")
        comments_headers = [
            "id", "category_name", "instruction_title", "anchor",
            "author", "is_anonymous", "text", "created_at"
        ]
        comments_ws.append(comments_headers)
        for cell in comments_ws[1]:
            cell.font = Font(bold=True)

        for row in comments_rows:
            comments_ws.append([
                row["id"],
                row["category_name"],
                row["instruction_title"],
                row["anchor"],
                row["author"],
                row["is_anonymous"],
                row["text"],
                row["created_at"],
            ])

        ratings_ws = wb.create_sheet("Оценки")
        ratings_headers = [
            "id", "category_name", "instruction_title", "rating", "anchor", "created_at"
        ]
        ratings_ws.append(ratings_headers)
        for cell in ratings_ws[1]:
            cell.font = Font(bold=True)

        for row in ratings_rows:
            ratings_ws.append([
                row["id"],
                row["category_name"],
                row["instruction_title"],
                row["rating"],
                row.get("anchor", ""),
                row["created_at"],
            ])

        for ws in (summary_ws, comments_ws, ratings_ws):
            for col_idx, column_cells in enumerate(ws.iter_cols(1, ws.max_column), start=1):
                max_len = 0
                for cell in column_cells:
                    value = cell.value
                    if value is not None:
                        max_len = max(max_len, len(str(value)))
                ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 60)

        wb.save(str(path))

    def rating_stats(self, instruction_id):
        row = self.conn.execute("""
            SELECT ROUND(AVG(rating), 2) AS avg_rating,
                   COUNT(*) AS cnt
            FROM ratings
            WHERE instruction_id=?
        """, (instruction_id,)).fetchone()

        avg_rating = float(row["avg_rating"]) if row and row["avg_rating"] is not None else 0.0
        cnt = int(row["cnt"]) if row and row["cnt"] is not None else 0
        return avg_rating, cnt

    def add_rating(self, instruction_id, rating, anchor="", created_at=None):
        instruction_id = int(instruction_id)
        created_at = created_at or ts()
        self.conn.execute("""
            INSERT INTO ratings (instruction_id, rating, anchor, created_at)
            VALUES (?, ?, ?, ?)
        """, (instruction_id, int(rating), anchor or "", created_at))
        self.conn.commit()

    def ratings_for_instruction(self, instruction_id):
        rows = self.conn.execute("""
            SELECT id, rating, anchor, created_at
            FROM ratings
            WHERE instruction_id=?
            ORDER BY id DESC
        """, (instruction_id,)).fetchall()
        return [dict(row) for row in rows]

    def delete_rating(self, rating_id):
        self.conn.execute(
            "DELETE FROM ratings WHERE id=?",
            (rating_id,)
        )
        self.conn.commit()

    def add_category(self, name, parent_id=None):
        name = name.strip()

        if parent_id is None:
            row = self.conn.execute("""
                SELECT COALESCE(MAX(sort_order), 0) + 1 AS next_order
                FROM categories
                WHERE parent_id IS NULL
            """).fetchone()
        else:
            row = self.conn.execute("""
                SELECT COALESCE(MAX(sort_order), 0) + 1 AS next_order
                FROM categories
                WHERE parent_id=?
            """, (parent_id,)).fetchone()

        next_order = int(row["next_order"]) if row else 1

        cur = self.conn.execute("""
            INSERT INTO categories (name, parent_id, sort_order)
            VALUES (?, ?, ?)
        """, (name, parent_id, next_order))
        self.conn.commit()
        return cur.lastrowid

    def add_task(self, category_id, title, instruction_title="", instruction_id=None):
        title = title.strip()
        instruction_id = None

        exists = self.conn.execute("""
            SELECT id
            FROM tasks
            WHERE category_id=? AND LOWER(title)=LOWER(?)
            LIMIT 1
        """, (category_id, title)).fetchone()
        if exists:
            raise ValueError("Задача с таким названием уже существует в этой подкатегории.")

        if instruction_id is None and instruction_title.strip():
            instruction = self.instruction_by_title(instruction_title.strip())
            if not instruction:
                raise ValueError(f"Инструкция не найдена: {instruction_title}")
            instruction_id = instruction["instruction_id"]

        if self.category_has_children(category_id):
            raise ValueError("Задачу можно добавлять только в подкатегорию без вложенных разделов.")

        row = self.conn.execute(
            "SELECT COALESCE(MAX(sort_order), 0) + 1 AS next_order FROM tasks WHERE category_id=?",
            (category_id,)
        ).fetchone()
        next_order = int(row["next_order"]) if row else 1

        cur = self.conn.execute("""
            INSERT INTO tasks
                (category_id, title, instruction_id, checklist_json, sort_order)
            VALUES (?, ?, ?, ?, ?)
        """, (
            category_id,
            title,
            instruction_id,
            json.dumps([], ensure_ascii=False),
            next_order
        ))
        self.conn.commit()
        return cur.lastrowid

    def rename_category(self, category_id, new_name):
        self.conn.execute(
            "UPDATE categories SET name=? WHERE id=?",
            (new_name.strip(), category_id)
        )
        self.conn.commit()

    def rename_task(self, task_id, new_title):
        task = self.conn.execute("""
            SELECT category_id
            FROM tasks
            WHERE id=?
        """, (task_id,)).fetchone()

        if not task:
            raise ValueError("Задача не найдена.")

        new_title = new_title.strip()

        exists = self.conn.execute("""
            SELECT id
            FROM tasks
            WHERE category_id=? AND LOWER(title)=LOWER(?) AND id<>?
            LIMIT 1
        """, (task["category_id"], new_title, task_id)).fetchone()
        if exists:
            raise ValueError("Задача с таким названием уже существует в этой подкатегории.")

        self.conn.execute(
            "UPDATE tasks SET title=? WHERE id=?",
            (new_title, task_id)
        )
        self.conn.commit()

    def delete_category(self, category_id):
        self.conn.execute(
            "DELETE FROM categories WHERE id=?",
            (category_id,)
        )
        self.conn.commit()

    def delete_task(self, task_id):
        self.conn.execute(
            "DELETE FROM tasks WHERE id=?",
            (task_id,)
        )
        self.conn.commit()

    def move_category_up(self, category_id):
        current = self.conn.execute(
            "SELECT sort_order, parent_id FROM categories WHERE id=?",
            (category_id,)
        ).fetchone()
        if not current:
            return

        cur_order = current["sort_order"]
        parent_id = current["parent_id"]

        if parent_id is None:
            prev = self.conn.execute("""
                SELECT id, sort_order
                FROM categories
                WHERE parent_id IS NULL AND sort_order < ?
                ORDER BY sort_order DESC
                LIMIT 1
            """, (cur_order,)).fetchone()
        else:
            prev = self.conn.execute("""
                SELECT id, sort_order
                FROM categories
                WHERE parent_id=? AND sort_order < ?
                ORDER BY sort_order DESC
                LIMIT 1
            """, (parent_id, cur_order)).fetchone()

        if not prev:
            return

        self.conn.execute(
            "UPDATE categories SET sort_order=? WHERE id=?",
            (prev["sort_order"], category_id)
        )
        self.conn.execute(
            "UPDATE categories SET sort_order=? WHERE id=?",
            (cur_order, prev["id"])
        )
        self.conn.commit()

    def move_category_down(self, category_id):
        current = self.conn.execute(
            "SELECT sort_order, parent_id FROM categories WHERE id=?",
            (category_id,)
        ).fetchone()
        if not current:
            return

        cur_order = current["sort_order"]
        parent_id = current["parent_id"]

        if parent_id is None:
            next_cat = self.conn.execute("""
                SELECT id, sort_order
                FROM categories
                WHERE parent_id IS NULL AND sort_order > ?
                ORDER BY sort_order ASC
                LIMIT 1
            """, (cur_order,)).fetchone()
        else:
            next_cat = self.conn.execute("""
                SELECT id, sort_order
                FROM categories
                WHERE parent_id=? AND sort_order > ?
                ORDER BY sort_order ASC
                LIMIT 1
            """, (parent_id, cur_order)).fetchone()

        if not next_cat:
            return

        self.conn.execute(
            "UPDATE categories SET sort_order=? WHERE id=?",
            (next_cat["sort_order"], category_id)
        )
        self.conn.execute(
            "UPDATE categories SET sort_order=? WHERE id=?",
            (cur_order, next_cat["id"])
        )
        self.conn.commit()

    def move_task_up(self, task_id):
        """Перемещает задачу вверх в пределах её категории."""
        task = self.conn.execute(
            "SELECT category_id, sort_order FROM tasks WHERE id=?", (task_id,)
        ).fetchone()
        if not task:
            return
        cat_id, cur_order = task["category_id"], task["sort_order"]
        prev = self.conn.execute(
            "SELECT id, sort_order FROM tasks WHERE category_id=? AND sort_order < ? ORDER BY sort_order DESC LIMIT 1",
            (cat_id, cur_order)
        ).fetchone()
        if not prev:
            return
        self.conn.execute("UPDATE tasks SET sort_order=? WHERE id=?", (prev["sort_order"], task_id))
        self.conn.execute("UPDATE tasks SET sort_order=? WHERE id=?", (cur_order, prev["id"]))
        self.conn.commit()

    def move_task_down(self, task_id):
        """Перемещает задачу вниз в пределах её категории."""
        task = self.conn.execute(
            "SELECT category_id, sort_order FROM tasks WHERE id=?", (task_id,)
        ).fetchone()
        if not task:
            return
        cat_id, cur_order = task["category_id"], task["sort_order"]
        next_task = self.conn.execute(
            "SELECT id, sort_order FROM tasks WHERE category_id=? AND sort_order > ? ORDER BY sort_order ASC LIMIT 1",
            (cat_id, cur_order)
        ).fetchone()
        if not next_task:
            return
        self.conn.execute("UPDATE tasks SET sort_order=? WHERE id=?", (next_task["sort_order"], task_id))
        self.conn.execute("UPDATE tasks SET sort_order=? WHERE id=?", (cur_order, next_task["id"]))
        self.conn.commit()

    def update_task_view_data(self, task_id, short_desc, instruction_title, checklist):
        task = self.task_bundle(task_id)
        if not task:
            raise ValueError("Задача не найдена.")

        instruction_id = task.get("instruction_real_id")
        if not instruction_id:
            raise ValueError("У задачи нет привязанной инструкции.")

        short_desc = short_desc.strip()
        instruction_title = instruction_title.strip()

        if not short_desc:
            raise ValueError("Описание не может быть пустым.")

        if not instruction_title:
            raise ValueError("Название инструкции не может быть пустым.")

        other = self.conn.execute("""
            SELECT id
            FROM instructions
            WHERE LOWER(title)=LOWER(?) AND id<>?
            LIMIT 1
        """, (instruction_title, instruction_id)).fetchone()

        if other:
            raise ValueError("Инструкция с таким названием уже существует.")

        self.conn.execute("""
            UPDATE instructions
            SET title=?, short_desc=?, updated_at=?
            WHERE id=?
        """, (
            instruction_title,
            short_desc,
            ts(),
            instruction_id
        ))

        self.conn.execute("""
            UPDATE tasks
            SET checklist_json=?
            WHERE id=?
        """, (
            json.dumps(normalize_checklist_sections(checklist), ensure_ascii=False),
            task_id
        ))

        self.conn.commit()

    def add_instruction(self, category_id, task_id, instruction_title, short_desc, sections, related_ids):
        """
        Создаёт инструкцию и привязывает её к существующей задаче task_id.
        Больше не создаёт новую задачу.
        """
        instruction_title = instruction_title.strip()
        short_desc = short_desc.strip()

        if not instruction_title:
            raise ValueError("Название инструкции не может быть пустым.")
        if not task_id:
            raise ValueError("Не указана задача для привязки инструкции.")

        # Проверяем, что задача существует
        task = self.task_bundle(task_id)
        if not task:
            raise ValueError("Задача не найдена.")

        if self.category_has_children(category_id):
            raise ValueError("Инструкцию можно привязать только к подкатегории без вложенных разделов.")

        if not sections:
            sections = [{"title": "Коротко", "body": p(short_desc or instruction_title), "image_path": ""}]

        if self.category_has_children(category_id):
            raise ValueError("Инструкцию можно привязать только к подкатегории без вложенных разделов.")

        related_ids = [int(rid) for rid in (related_ids or []) if rid is not None]

        cur = self.conn.execute("""
            INSERT INTO instructions
                (category_id, title, short_desc, sections_json, related_ids_json, updated_at)
            VALUES (?, ?, ?, ?, ?, ?)
        """, (
            category_id,
            instruction_title,
            short_desc,
            json.dumps(normalize_instruction_sections(sections), ensure_ascii=False),
            json.dumps(related_ids, ensure_ascii=False),
            ts(),
        ))
        instruction_id = cur.lastrowid

        # Привязываем инструкцию к задаче
        self.conn.execute(
            "UPDATE tasks SET instruction_id = ? WHERE id = ?",
            (instruction_id, task_id)
        )
        self.conn.commit()

        return instruction_id, []

    def set_task_instruction(self, task_id, instruction_id):
        self.conn.execute("""
            UPDATE tasks
            SET instruction_id=?
            WHERE id=?
        """, (instruction_id, task_id))
        self.conn.commit()

    def get_banner_text(self):
        """Возвращает сохранённый текст баннера или значение по умолчанию."""
        self.conn.execute("""
            CREATE TABLE IF NOT EXISTS settings (
                key TEXT PRIMARY KEY,
                value TEXT NOT NULL DEFAULT ''
            )
        """)
        row = self.conn.execute(
            "SELECT value FROM settings WHERE key='banner_text'"
        ).fetchone()
        if row:
            return row["value"]
        default = (
            "Слева выбери задачу. Сначала смотри чек‑лист, потом подробную инструкцию. "
            "Комментарии и оценки — анонимные."
        )
        self.conn.execute(
            "INSERT OR IGNORE INTO settings (key, value) VALUES ('banner_text', ?)",
            (default,)
        )
        self.conn.commit()
        return default

    def save_banner_text(self, text):
        """Сохраняет текст баннера."""
        self.conn.execute("""
            CREATE TABLE IF NOT EXISTS settings (
                key TEXT PRIMARY KEY,
                value TEXT NOT NULL DEFAULT ''
            )
        """)
        self.conn.execute(
            "INSERT OR REPLACE INTO settings (key, value) VALUES ('banner_text', ?)",
            (text,)
        )
        self.conn.commit()

    def _onboarding_setting_key(self, username):
        safe = (username or "unknown").strip().casefold()
        safe = safe.replace("\\", "_").replace("/", "_")
        return f"onboarding_completed:{safe}"

    def _ensure_settings_table(self):
        self.conn.execute("""
            CREATE TABLE IF NOT EXISTS settings (
                key TEXT PRIMARY KEY,
                value TEXT NOT NULL DEFAULT ''
            )
        """)

    def is_onboarding_completed(self, username):
        """Пройдено ли интерактивное обучение для пользователя в текущей версии."""
        self._ensure_settings_table()
        row = self.conn.execute(
            "SELECT value FROM settings WHERE key=?",
            (self._onboarding_setting_key(username),),
        ).fetchone()
        if not row:
            return False
        return str(row["value"] or "").strip() == APP_VERSION

    def set_onboarding_completed(self, username, completed=True):
        """Отмечает обучение пройденным для пользователя в текущей версии."""
        self._ensure_settings_table()
        value = APP_VERSION if completed else ""
        self.conn.execute(
            "INSERT OR REPLACE INTO settings (key, value) VALUES (?, ?)",
            (self._onboarding_setting_key(username), value),
        )
        self.conn.commit()

    def update_instruction(self, instruction_id, category_id, title, short_desc, sections, related_ids, task_id=None):
        """Обновляет существующую инструкцию. Одна инструкция может быть привязана к нескольким подзадачам."""
        title = title.strip()
        short_desc = short_desc.strip()
        if not title:
            raise ValueError("Название инструкции не может быть пустым.")
        if not task_id:
            raise ValueError("Не указана подзадача для привязки инструкции.")

        existing = self.instruction_by_id(instruction_id)
        if not existing:
            raise ValueError("Инструкция не найдена.")

        task = self.task_bundle(task_id)
        if not task:
            raise ValueError("Подзадача не найдена.")

        if self.category_has_children(task["category_id"]):
            raise ValueError("Инструкцию можно привязать только к подзадаче без вложенных разделов.")

        linked_tasks = self.tasks_for_instruction(instruction_id)
        if len(linked_tasks) <= 1:
            category_id = task["category_id"]
        else:
            category_id = existing["category_id"]

        other = self.conn.execute("""
            SELECT id
            FROM instructions
            WHERE LOWER(title)=LOWER(?) AND id<>?
            LIMIT 1
        """, (title, instruction_id)).fetchone()
        if other:
            raise ValueError("Инструкция с таким названием уже существует.")

        related_ids = [int(rid) for rid in (related_ids or []) if rid is not None]

        self.conn.execute("""
            UPDATE instructions
            SET category_id=?,
                title=?,
                short_desc=?,
                sections_json=?,
                related_ids_json=?,
                updated_at=?
            WHERE id=?
        """, (
            category_id,
            title,
            short_desc,
            json.dumps(normalize_instruction_sections(sections), ensure_ascii=False),
            json.dumps(related_ids, ensure_ascii=False),
            ts(),
            instruction_id
        ))

        # Обновляем привязку только у выбранной подзадачи, не снимая её с остальных.
        self.conn.execute(
            "UPDATE tasks SET instruction_id = ? WHERE id = ?",
            (instruction_id, task_id)
        )
        self.conn.commit()
        return []


# ================== UI ==================
class LargeTextEditDialog(QDialog):
    """Большое окно для удобного редактирования длинного текста."""

    def __init__(self, title, text="", parent=None):
        super().__init__(parent)
        self.setWindowTitle(title)
        self.setModal(True)
        self.resize(920, 680)

        layout = QVBoxLayout(self)
        layout.setSpacing(8)

        self.text_edit = TypographicPlainTextEdit()
        self.text_edit.setPlainText(text or "")
        self.text_edit.setPlaceholderText("Введите текст...")
        layout.addWidget(self.text_edit, 1)

        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)

    def text(self):
        return self.text_edit.toPlainText()


class EditorSectionFrame(QFrame):
    """Сворачиваемая карточка секции для редакторов."""

    def __init__(self, title_getter=None, summary_getter=None, parent=None):
        super().__init__(parent)
        self.title_getter = title_getter
        self.summary_getter = summary_getter
        self.setFrameStyle(QFrame.Box | QFrame.Plain)
        self.setStyleSheet("""
            QFrame {
                border: 1px solid #cfd8e3;
                border-radius: 8px;
                background: #ffffff;
            }
        """)

        layout = QVBoxLayout(self)
        layout.setContentsMargins(8, 8, 8, 8)
        layout.setSpacing(8)

        header = QHBoxLayout()
        header.setSpacing(6)

        self.toggle_btn = QToolButton()
        self.toggle_btn.setCheckable(True)
        self.toggle_btn.setChecked(True)
        self.toggle_btn.setArrowType(Qt.DownArrow)
        self.toggle_btn.clicked.connect(self.set_expanded)
        header.addWidget(self.toggle_btn)

        self.title_label = QLabel("Секция")
        self.title_label.setStyleSheet("font-weight: 600; color: #263238; border: none;")
        header.addWidget(self.title_label, 1)

        self.summary_label = QLabel("")
        self.summary_label.setStyleSheet("color: #6b7280; font-size: 11px; border: none;")
        header.addWidget(self.summary_label)

        self.header_controls = QHBoxLayout()
        self.header_controls.setSpacing(4)
        header.addLayout(self.header_controls)
        layout.addLayout(header)

        self.content_widget = QWidget()
        self.content_layout = QVBoxLayout(self.content_widget)
        self.content_layout.setContentsMargins(0, 0, 0, 0)
        self.content_layout.setSpacing(8)
        layout.addWidget(self.content_widget)

        self.refresh_header()

    def set_expanded(self, expanded):
        self.content_widget.setVisible(expanded)
        self.toggle_btn.setChecked(expanded)
        self.toggle_btn.setArrowType(Qt.DownArrow if expanded else Qt.RightArrow)
        self.refresh_header()

    def refresh_header(self):
        title = self.title_getter() if self.title_getter else ""
        title = str(title or "").strip() or "Без названия"
        self.title_label.setText(title)

        summary = self.summary_getter() if self.summary_getter else ""
        self.summary_label.setText(str(summary or ""))


class NoWheelComboBox(QComboBox):
    """Комбобокс, который не меняет выбор при прокрутке колёсиком."""

    def wheelEvent(self, event):
        event.ignore()


def pick_typographic_quote(text_before: str) -> str:
    """Выбирает открывающую или закрывающую «ёлочку» по контексту."""
    if not text_before:
        return "«"

    stripped = text_before.rstrip()
    if not stripped:
        return "«"

    last = stripped[-1]
    if last.isspace() or last in "([{«":
        return "«"
    if last in ")]}»":
        return "«"

    opens = text_before.count("«")
    closes = text_before.count("»")
    if opens > closes:
        return "»"
    return "«"


def replace_straight_quotes(text: str) -> str:
    """Заменяет прямые кавычки на «ёлочки» с чередованием."""
    if not text or '"' not in text:
        return text

    result = []
    open_next = True
    for ch in text:
        if ch == '"':
            result.append("«" if open_next else "»")
            open_next = not open_next
        else:
            result.append(ch)
    return "".join(result)


class TypographicLineEdit(QLineEdit):
    """Поле ввода с автозаменой \" на «»."""

    def keyPressEvent(self, event):
        if event.text() == '"':
            cursor_pos = self.cursorPosition()
            quote = pick_typographic_quote(self.text()[:cursor_pos])
            self.insert(quote)
            return
        super().keyPressEvent(event)

    def insert(self, text):
        super().insert(replace_straight_quotes(text))


class TypographicPlainTextEdit(QPlainTextEdit):
    """Многострочное поле с автозаменой \" на «»."""

    def keyPressEvent(self, event):
        if event.text() == '"':
            cursor = self.textCursor()
            text_before = cursor.block().text()[:cursor.positionInBlock()]
            cursor.insertText(pick_typographic_quote(text_before))
            return
        super().keyPressEvent(event)

    def insertFromMimeData(self, source):
        if source and source.hasText():
            self.textCursor().insertText(replace_straight_quotes(source.text()))
            return
        super().insertFromMimeData(source)


def open_large_text_editor(parent, text_edit, title="Редактирование текста"):
    dialog = LargeTextEditDialog(title, text_edit.toPlainText(), parent)
    if dialog.exec() == QDialog.Accepted:
        text_edit.setPlainText(dialog.text())


class _ImagePanHandler(QObject):
    """Перетаскивание и масштабирование изображения в QScrollArea."""

    def __init__(self, scroll_area, viewer, parent=None):
        super().__init__(parent)
        self._scroll_area = scroll_area
        self._viewer = viewer
        self._viewport = scroll_area.viewport()
        self._dragging = False
        self._drag_start = QPoint()
        self._scroll_start = (0, 0)
        self._viewport.setCursor(QCursor(Qt.OpenHandCursor))
        self._viewport.installEventFilter(self)

    def cleanup(self):
        self._dragging = False
        if isValid(self._viewport):
            try:
                self._viewport.releaseMouse()
            except RuntimeError:
                pass
            try:
                self._viewport.removeEventFilter(self)
            except RuntimeError:
                pass

    def _global_pos(self, event: QMouseEvent) -> QPoint:
        return event.globalPosition().toPoint()

    def eventFilter(self, obj, event):
        if not isValid(self._viewport) or not isValid(self._scroll_area):
            return False
        if obj is not self._viewport:
            return False

        if event.type() == QEvent.MouseButtonPress and event.button() == Qt.LeftButton:
            self._dragging = True
            self._drag_start = self._global_pos(event)
            hbar = self._scroll_area.horizontalScrollBar()
            vbar = self._scroll_area.verticalScrollBar()
            self._scroll_start = (hbar.value(), vbar.value())
            self._viewport.setCursor(QCursor(Qt.ClosedHandCursor))
            self._viewport.grabMouse()
            return True

        if event.type() == QEvent.MouseMove and self._dragging:
            delta = self._global_pos(event) - self._drag_start
            hbar = self._scroll_area.horizontalScrollBar()
            vbar = self._scroll_area.verticalScrollBar()
            hbar.setValue(self._scroll_start[0] - delta.x())
            vbar.setValue(self._scroll_start[1] - delta.y())
            return True

        if event.type() == QEvent.MouseButtonRelease and event.button() == Qt.LeftButton and self._dragging:
            self._dragging = False
            self._viewport.setCursor(QCursor(Qt.OpenHandCursor))
            self._viewport.releaseMouse()
            return True

        if event.type() == QEvent.Wheel and self._viewer is not None:
            delta = event.angleDelta().y()
            if delta > 0:
                self._viewer.zoom_in()
            elif delta < 0:
                self._viewer.zoom_out()
            return True

        return False


class ImageViewerDialog(QDialog):
    """Просмотр скриншота в полном размере с масштабированием."""

    MIN_ZOOM = 0.25
    MAX_ZOOM = 4.0
    ZOOM_STEP = 1.15

    def __init__(self, image_path, parent=None):
        super().__init__(parent)
        self.image_path = Path(image_path)
        self._pixmap = QPixmap(str(image_path))
        self._zoom = 1.0

        self.setWindowTitle(self.image_path.name or "Изображение")
        self.resize(960, 720)

        layout = QVBoxLayout(self)
        layout.setSpacing(8)

        self.scroll_area = QScrollArea()
        self.scroll_area.setWidgetResizable(False)
        self.scroll_area.setAlignment(Qt.AlignCenter)
        self.scroll_area.setFrameShape(QFrame.NoFrame)

        self.image_label = QLabel()
        self.image_label.setAlignment(Qt.AlignCenter)
        self.image_label.setBackgroundRole(QPalette.Base)
        self.image_label.setAutoFillBackground(True)
        self.scroll_area.setWidget(self.image_label)
        self._pan_handler = _ImagePanHandler(self.scroll_area, viewer=self, parent=self)
        layout.addWidget(self.scroll_area, 1)

        controls = QHBoxLayout()
        zoom_out_btn = QPushButton("−")
        zoom_out_btn.setFixedWidth(36)
        zoom_out_btn.setToolTip("Уменьшить (колёсико мыши)")
        zoom_out_btn.clicked.connect(self.zoom_out)

        self.zoom_label = QLabel("100%")
        self.zoom_label.setMinimumWidth(52)
        self.zoom_label.setAlignment(Qt.AlignCenter)

        zoom_in_btn = QPushButton("+")
        zoom_in_btn.setFixedWidth(36)
        zoom_in_btn.setToolTip("Увеличить (колёсико мыши)")
        zoom_in_btn.clicked.connect(self.zoom_in)

        fit_width_btn = QPushButton("По ширине")
        fit_width_btn.setToolTip("Вписать изображение по ширине окна")
        fit_width_btn.clicked.connect(self.fit_width)

        original_btn = QPushButton("Оригинал")
        original_btn.setToolTip("Показать в исходном масштабе (100%)")
        original_btn.clicked.connect(self.reset_zoom)

        controls.addStretch()
        controls.addWidget(zoom_out_btn)
        controls.addWidget(self.zoom_label)
        controls.addWidget(zoom_in_btn)
        controls.addWidget(fit_width_btn)
        controls.addWidget(original_btn)
        controls.addStretch()
        layout.addLayout(controls)

        hint = QLabel("Перетаскивайте изображение мышью. Колёсико — масштаб.")
        hint.setStyleSheet("color: #5b6577; font-size: 11px;")
        hint.setAlignment(Qt.AlignCenter)
        layout.addWidget(hint)

        buttons = QDialogButtonBox(QDialogButtonBox.Close)
        buttons.rejected.connect(self.reject)
        buttons.accepted.connect(self.accept)
        layout.addWidget(buttons)

        if self._pixmap.isNull():
            self.image_label.setText("Не удалось загрузить изображение.")
            for btn in (zoom_out_btn, zoom_in_btn, fit_width_btn, original_btn):
                btn.setEnabled(False)
            return

        self.fit_width()

    def _apply_zoom(self):
        if self._pixmap.isNull():
            return
        width = max(1, int(self._pixmap.width() * self._zoom))
        height = max(1, int(self._pixmap.height() * self._zoom))
        scaled = self._pixmap.scaled(width, height, Qt.KeepAspectRatio, Qt.SmoothTransformation)
        self.image_label.setPixmap(scaled)
        self.image_label.resize(scaled.size())
        self.zoom_label.setText(f"{int(round(self._zoom * 100))}%")

    def _set_zoom(self, zoom):
        self._zoom = max(self.MIN_ZOOM, min(self.MAX_ZOOM, zoom))
        self._apply_zoom()

    def zoom_in(self):
        self._set_zoom(self._zoom * self.ZOOM_STEP)

    def zoom_out(self):
        self._set_zoom(self._zoom / self.ZOOM_STEP)

    def reset_zoom(self):
        self._set_zoom(1.0)

    def fit_width(self):
        if self._pixmap.isNull():
            return
        viewport_width = max(200, self.scroll_area.viewport().width() - 16)
        self._set_zoom(viewport_width / self._pixmap.width())

    def showEvent(self, event):
        super().showEvent(event)
        if not self._pixmap.isNull():
            QTimer.singleShot(0, self.fit_width)

    def closeEvent(self, event):
        if hasattr(self, "_pan_handler") and self._pan_handler is not None:
            self._pan_handler.cleanup()
        super().closeEvent(event)


class _ClickableImageLabel(QLabel):
    def __init__(self, on_click, parent=None):
        super().__init__(parent)
        self._on_click = on_click
        self.setCursor(QCursor(Qt.PointingHandCursor))

    def mousePressEvent(self, event):
        if event.button() == Qt.LeftButton and self._on_click:
            self._on_click()
        super().mousePressEvent(event)


class InstructionImageWidget(QFrame):
    """Превью скриншота в инструкции с кнопкой увеличения."""

    def __init__(self, image_path, preview_pixmap, parent=None):
        super().__init__(parent)
        self.image_path = str(image_path)

        self.setStyleSheet("""
            InstructionImageWidget {
                background: transparent;
                border: none;
            }
        """)

        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(4)

        self.preview_label = _ClickableImageLabel(self.open_viewer)
        self.preview_label.setPixmap(preview_pixmap)
        self.preview_label.setAlignment(Qt.AlignCenter)
        self.preview_label.setToolTip("Нажмите на изображение, чтобы рассмотреть его ближе")
        self.preview_label.setStyleSheet(
            "border: 1px solid #d9e3f0; border-radius: 6px; background: #ffffff;"
        )
        shadow = QGraphicsDropShadowEffect()
        shadow.setBlurRadius(12)
        shadow.setOffset(2, 2)
        shadow.setColor(QColor(0, 0, 0, 40))
        self.preview_label.setGraphicsEffect(shadow)
        layout.addWidget(self.preview_label)

        zoom_btn = QPushButton("🔍  Увеличить")
        zoom_btn.setCursor(QCursor(Qt.PointingHandCursor))
        zoom_btn.setToolTip("Открыть изображение в полном размере")
        zoom_btn.setStyleSheet("""
            QPushButton {
                background: #eef5fb;
                border: 1px solid #c8d9ef;
                border-radius: 6px;
                padding: 4px 10px;
                color: #1f4b99;
            }
            QPushButton:hover {
                background: #dce9f8;
            }
        """)
        zoom_btn.clicked.connect(self.open_viewer)
        layout.addWidget(zoom_btn, 0, Qt.AlignRight)

    def set_preview_pixmap(self, pixmap: QPixmap):
        self.preview_label.setPixmap(pixmap)

    def apply_source_width(self, source: QPixmap, display_width: int):
        if source.isNull():
            return
        display_width = max(80, int(display_width))
        if source.width() <= display_width:
            pixmap = source
        else:
            pixmap = source.scaledToWidth(display_width, Qt.SmoothTransformation)
        self.set_preview_pixmap(pixmap)

    def open_viewer(self):
        dialog = ImageViewerDialog(self.image_path, self.window())
        dialog.exec()


def set_all_editor_sections_expanded(layout, expanded):
    for i in range(layout.count()):
        widget = layout.itemAt(i).widget()
        if isinstance(widget, EditorSectionFrame):
            widget.set_expanded(expanded)


def lines_from_multiline_input(parent, title, label):
    text, ok = QInputDialog.getMultiLineText(parent, title, label, "")
    if not ok:
        return []
    return [line.strip() for line in text.splitlines() if line.strip()]


class InstructionEditorDialog(QDialog):
    def __init__(self, db: KnowledgeBaseDB, categories, default_category_id=None, default_task_id=None, instruction_data=None, parent=None):
        super().__init__(parent)
        self.db = db
        self.current_instruction_id = None
        if instruction_data is not None:
            self.current_instruction_id = instruction_data.get("instruction_id")
        self.current_instruction_id = (instruction_data or {}).get("instruction_id")
        self.setModal(True)
        self.setWindowTitle("Редактор инструкции" if instruction_data is None else "Редактирование инструкции")
        self.resize(900, 800)
        self.setModal(True)
        self.resize(900, 800)

        layout = QVBoxLayout(self)

        # Категория и задача
        form = QFormLayout()
        self.category_combo = QComboBox()
        for cat in categories:
            self.category_combo.addItem(cat["name"], cat["id"])
        if default_category_id is not None:
            idx = self.category_combo.findData(default_category_id)
            if idx >= 0:
                self.category_combo.setCurrentIndex(idx)

        self.task_combo = QComboBox()
        self.default_task_id = default_task_id
        self._loading_instruction_data = False
        self._populate_tasks()

        self.category_combo.currentIndexChanged.connect(self._on_category_changed)

        self.title_edit = TypographicLineEdit()
        self.title_edit.setPlaceholderText("Название инструкции")
        self.short_desc_edit = TypographicLineEdit()
        self.short_desc_edit.setPlaceholderText("Краткое описание (одна строка)")

        form.addRow("Категория:", self.category_combo)
        form.addRow("Подзадача:", self.task_combo)
        form.addRow("Название инструкции:", self.title_edit)
        form.addRow("Краткое описание:", self.short_desc_edit)
        layout.addLayout(form)

        # Секции (без изменений)
        sections_group = QGroupBox("Секции")
        sections_layout = QVBoxLayout(sections_group)

        self.sections_widget = QWidget()
        self.sections_layout = QVBoxLayout(self.sections_widget)
        self.sections_layout.setContentsMargins(0, 0, 0, 0)
        self.sections_layout.setSpacing(8)

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setWidget(self.sections_widget)
        sections_layout.addWidget(scroll)

        # Подсказка по markdown
        markdown_hint = QLabel(
            '<span style="color:#5b6577; font-size:11px;">'
            'Поддерживается Markdown: <b>#</b> Заголовок, <b>**</b>жирный<b>**</b>, '
            '<b>*</b>курсив<b>*</b>, <b>`</b>код<b>`</b>, '
            '<b>-</b> список, <b>1.</b> нумерованный список, '
            '<b>&gt;</b> цитата, <b>[текст](url)</b> ссылка'
            '</span>'
        )
        markdown_hint.setWordWrap(True)
        markdown_hint.setTextFormat(Qt.RichText)
        sections_layout.addWidget(markdown_hint)

        btn_row = QHBoxLayout()
        add_btn = QPushButton("Добавить секцию")
        add_btn.clicked.connect(lambda: self.add_section())
        btn_row.addWidget(add_btn)
        add_many_btn = QPushButton("Создать из списка")
        add_many_btn.setToolTip("Каждая строка станет новой секцией.")
        add_many_btn.clicked.connect(self.add_sections_from_list)
        btn_row.addWidget(add_many_btn)
        collapse_btn = QPushButton("Свернуть все")
        collapse_btn.clicked.connect(lambda: set_all_editor_sections_expanded(self.sections_layout, False))
        btn_row.addWidget(collapse_btn)
        expand_btn = QPushButton("Развернуть все")
        expand_btn.clicked.connect(lambda: set_all_editor_sections_expanded(self.sections_layout, True))
        btn_row.addWidget(expand_btn)
        btn_row.addStretch()
        sections_layout.addLayout(btn_row)

        layout.addWidget(sections_group)

        # Связанные инструкции (без изменений)
        related_group = QGroupBox("Связанные инструкции")
        related_layout = QVBoxLayout(related_group)

        self.related_combo = QComboBox()
        self.related_combo.setEditable(False)
        self._populate_related_combo()
        add_rel_btn = QPushButton("Добавить связь")
        add_rel_btn.clicked.connect(self.add_related)

        rel_combo_layout = QHBoxLayout()
        rel_combo_layout.addWidget(self.related_combo, 1)
        rel_combo_layout.addWidget(add_rel_btn)

        self.related_list = QListWidget()
        self.related_list.setAlternatingRowColors(True)
        del_rel_btn = QPushButton("Удалить выбранную связь")
        del_rel_btn.clicked.connect(self.remove_selected_related)

        related_layout.addLayout(rel_combo_layout)
        related_layout.addWidget(self.related_list)
        related_layout.addWidget(del_rel_btn)

        layout.addWidget(related_group)

        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        buttons.accepted.connect(self.validate_and_accept)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)

        # Заполнение при редактировании
        if instruction_data is not None:
            self._populate_from_data(instruction_data)
        else:
            self.add_section()

    def _on_category_changed(self):
        if not self._loading_instruction_data:
            self.default_task_id = None
        self._populate_tasks()

    def _populate_tasks(self):
        """Заполняет выпадающий список задач для выбранной категории."""
        self.task_combo.clear()
        cat_id = self.category_combo.currentData()
        if cat_id is None:
            return

        tasks = self.db.tasks_for_category(cat_id)
        for t in tasks:
            self.task_combo.addItem(t["task_title"], t["task_id"])

        selected_task_id = self.default_task_id
        if selected_task_id is not None:
            for i in range(self.task_combo.count()):
                if self.task_combo.itemData(i) == selected_task_id:
                    self.task_combo.setCurrentIndex(i)
                    return

            task = self.db.task_bundle(selected_task_id)
            if task:
                self.task_combo.insertItem(0, task["task_title"], task["task_id"])
                self.task_combo.setCurrentIndex(0)
                return

        if self.task_combo.count() == 0:
            self.task_combo.addItem("(нет подзадач)", None)

    def _section_blocks_for_editor(self, section_data):
        blocks = section_data.get("blocks") or []
        if blocks:
            return blocks

        legacy_blocks = []
        body = str(section_data.get("body") or "").strip()
        if body:
            legacy_blocks.append({"type": "text", "content": body})

        image_path = str(section_data.get("image_path") or "").strip()
        if image_path:
            legacy_blocks.append({
                "type": "images",
                "paths": [image_path],
                "image_width": section_data.get("image_width", 760),
            })

        return legacy_blocks

    def _populate_from_data(self, data: dict):
        """Заполняет поля редактора существующими данными инструкции."""
        self._loading_instruction_data = True
        try:
            task_id = data.get("task_id")
            category_id = data.get("task_category_id") or data.get("category_id")
            if task_id:
                self.default_task_id = task_id

            self.category_combo.blockSignals(True)
            idx = self.category_combo.findData(category_id)
            if idx >= 0:
                self.category_combo.setCurrentIndex(idx)
            self.category_combo.blockSignals(False)

            self._populate_tasks()

            self.task_combo.setEnabled(True)
            self.category_combo.setEnabled(True)

            self.title_edit.setText(data.get("title", ""))
            self.short_desc_edit.setText(data.get("short_desc", ""))

            sections = data.get("sections", [])
            for sec in sections:
                self.add_section(sec.get("id") or sec.get("section_id"), sec.get("title", ""))
                frame = self.sections_layout.itemAt(self.sections_layout.count() - 1).widget()

                blocks = self._section_blocks_for_editor(sec)
                if blocks and hasattr(frame, 'blocks_layout'):
                    while frame.blocks_layout.count() > 0:
                        item = frame.blocks_layout.takeAt(0)
                        if item.widget():
                            item.widget().deleteLater()

                    for block in blocks:
                        block_type = block.get("type")
                        if block_type == "text":
                            self._add_text_block(frame.blocks_layout)
                            last_block = frame.blocks_layout.itemAt(frame.blocks_layout.count() - 1).widget()
                            if hasattr(last_block, 'text_edit'):
                                content = block.get("content", "")
                                last_block.text_edit.setPlainText(content)

                        elif block_type == "images":
                            self._add_images_block(frame.blocks_layout)
                            last_block = frame.blocks_layout.itemAt(frame.blocks_layout.count() - 1).widget()
                            if hasattr(last_block, 'width_combo'):
                                saved_width = str(block.get("image_width", 760))
                                if saved_width == "0":
                                    last_block.width_combo.setCurrentText("Оригинал")
                                else:
                                    last_block.width_combo.setCurrentText(saved_width)
                            if hasattr(last_block, 'img_list'):
                                for img_path in block.get("paths", []):
                                    if img_path:
                                        last_block.img_list.addItem(img_path)

            if not sections:
                self.add_section()

            for rel_id in data.get("related_ids", []):
                for i in range(self.related_combo.count()):
                    if self.related_combo.itemData(i) == rel_id:
                        item = QListWidgetItem(self.related_combo.itemText(i))
                        item.setData(Qt.UserRole, rel_id)
                        self.related_list.addItem(item)
                        break
        finally:
            self._loading_instruction_data = False

    def _populate_related_combo(self):
        self.related_combo.clear()
        for inst in self.db.all_instructions():
            if self.current_instruction_id is not None and inst["instruction_id"] == self.current_instruction_id:
                continue
            label = f"#{inst['instruction_id']} • {inst['category_name']} — {inst['title']}"
            self.related_combo.addItem(label, inst["instruction_id"])

    def _instruction_section_summary(self, frame):
        count = frame.blocks_layout.count() if hasattr(frame, "blocks_layout") else 0
        return f"{count} блок(ов)"

    def add_sections_from_list(self):
        titles = lines_from_multiline_input(
            self,
            "Создать секции из списка",
            "Вставьте названия секций, каждое с новой строки:"
        )
        for title in titles:
            self.add_section(title=title)

    def add_section(self, section_id=None, title=""):
        section_title = str(title or "")
        if isinstance(section_id, bool):
            section_id = None
        elif isinstance(section_id, str) and not section_id.startswith("sec_"):
            section_title = section_id
            section_id = None

        frame = EditorSectionFrame(parent=self)
        layout = frame.content_layout

        # Заголовок секции
        title_edit = TypographicLineEdit()
        title_edit.setPlaceholderText("Заголовок секции")
        title_edit.setText(section_title)
        layout.addWidget(title_edit)

        # Список блоков контента (текст / изображения)
        blocks_widget = QWidget()
        blocks_layout = QVBoxLayout(blocks_widget)
        blocks_layout.setContentsMargins(0, 0, 0, 0)
        blocks_layout.setSpacing(6)
        layout.addWidget(blocks_widget)

        # Кнопки для добавления блоков
        add_block_layout = QHBoxLayout()
        add_text_btn = QPushButton("+ Текст")
        add_text_btn.clicked.connect(lambda: self._add_text_block(blocks_layout))
        add_images_btn = QPushButton("+ Изображения")
        add_images_btn.clicked.connect(lambda: self._add_images_block(blocks_layout))
        add_block_layout.addWidget(add_text_btn)
        add_block_layout.addWidget(add_images_btn)
        add_block_layout.addStretch()
        layout.addLayout(add_block_layout)

        # Сохраняем ссылки в frame для последующего сбора данных
        frame.section_id = section_id or new_section_id()
        frame.title_edit = title_edit
        frame.blocks_layout = blocks_layout
        frame.title_getter = lambda f=frame: f.title_edit.text()
        frame.summary_getter = lambda f=frame: self._instruction_section_summary(f)
        title_edit.textChanged.connect(frame.refresh_header)

        # Кнопки управления секцией
        up_btn = QPushButton("↑")
        up_btn.setToolTip("Переместить выше")
        up_btn.clicked.connect(lambda: self._move_section(frame, -1))
        down_btn = QPushButton("↓")
        down_btn.setToolTip("Переместить ниже")
        down_btn.clicked.connect(lambda: self._move_section(frame, 1))
        duplicate_btn = QPushButton("Дублировать")
        duplicate_btn.clicked.connect(lambda: self._duplicate_section(frame))
        del_btn = QPushButton("Удалить секцию")
        del_btn.clicked.connect(lambda: self._delete_section(frame))

        frame.header_controls.addWidget(up_btn)
        frame.header_controls.addWidget(down_btn)
        frame.header_controls.addWidget(duplicate_btn)
        frame.header_controls.addWidget(del_btn)

        self.sections_layout.addWidget(frame)

        # Добавляем один текстовый блок по умолчанию
        self._add_text_block(blocks_layout)
        frame.refresh_header()

    def _browse_image(self, line_edit: QLineEdit):
        file_path, _ = QFileDialog.getOpenFileName(
            self, "Выберите изображение", "",
            "Изображения (*.png *.jpg *.jpeg *.bmp *.gif);;Все файлы (*)"
        )
        if file_path:
            line_edit.setText(file_path)

    def _add_text_block(self, blocks_layout):
        """Добавляет текстовый блок в секцию."""
        block_frame = QFrame()
        block_frame.setFrameStyle(QFrame.StyledPanel | QFrame.Plain)
        block_frame.setStyleSheet("QFrame { background: #ffffff; border: 1px solid #e0e0e0; border-radius: 4px; }")
        block_layout = QVBoxLayout(block_frame)
        block_layout.setContentsMargins(6, 6, 6, 6)
        block_layout.setSpacing(4)

        # Метка типа блока
        header = QHBoxLayout()
        header.addWidget(QLabel("📝 Текст"))
        header.addStretch()
        expand_btn = QPushButton("Развернуть")
        expand_btn.setToolTip("Открыть текст в большом окне")
        header.addWidget(expand_btn)
        del_btn = QPushButton("✕")
        del_btn.setFixedSize(24, 24)
        del_btn.setToolTip("Удалить блок")
        del_btn.clicked.connect(lambda: self._delete_block(block_frame, blocks_layout))
        header.addWidget(del_btn)
        block_layout.addLayout(header)

        text_edit = TypographicPlainTextEdit()
        text_edit.setPlaceholderText("Текст блока...")
        text_edit.setMinimumHeight(160)
        block_layout.addWidget(text_edit)
        expand_btn.clicked.connect(lambda: open_large_text_editor(self, text_edit, "Редактирование текстового блока"))

        # Сохраняем тип блока
        block_frame.block_type = "text"
        block_frame.text_edit = text_edit

        blocks_layout.addWidget(block_frame)

        parent_section = blocks_layout.parentWidget().parentWidget()
        if hasattr(parent_section, "refresh_header"):
            parent_section.refresh_header()

    def _add_images_block(self, blocks_layout):
        """Добавляет блок изображений в секцию."""
        block_frame = QFrame()
        block_frame.setFrameStyle(QFrame.StyledPanel | QFrame.Plain)
        block_frame.setStyleSheet("QFrame { background: #ffffff; border: 1px solid #e0e0e0; border-radius: 4px; }")
        block_layout = QVBoxLayout(block_frame)
        block_layout.setContentsMargins(6, 6, 6, 6)
        block_layout.setSpacing(4)

        # Метка типа блока
        header = QHBoxLayout()
        header.addWidget(QLabel("🖼 Изображения"))
        header.addStretch()
        del_btn = QPushButton("✕")
        del_btn.setFixedSize(24, 24)
        del_btn.setToolTip("Удалить блок")
        del_btn.clicked.connect(lambda: self._delete_block(block_frame, blocks_layout))
        header.addWidget(del_btn)
        block_layout.addLayout(header)

        # Список изображений
        img_list = QListWidget()
        img_list.setAlternatingRowColors(True)
        img_list.setMaximumHeight(100)
        block_layout.addWidget(img_list)

        # Кнопки управления изображениями
        img_btn_layout = QHBoxLayout()
        add_img_btn = QPushButton("Добавить изображение")
        add_img_btn.clicked.connect(lambda _, lst=img_list: self._add_image_to_list(lst))
        remove_img_btn = QPushButton("Удалить")
        remove_img_btn.clicked.connect(lambda _, lst=img_list: self._remove_selected_image(lst))
        up_img_btn = QPushButton("↑")
        up_img_btn.clicked.connect(lambda _, lst=img_list: self._move_image(lst, -1))
        down_img_btn = QPushButton("↓")
        down_img_btn.clicked.connect(lambda _, lst=img_list: self._move_image(lst, 1))

        img_btn_layout.addWidget(add_img_btn)
        img_btn_layout.addWidget(remove_img_btn)
        img_btn_layout.addWidget(up_img_btn)
        img_btn_layout.addWidget(down_img_btn)
        img_btn_layout.addStretch()
        block_layout.addLayout(img_btn_layout)

        # Выбор ширины отображения
        width_layout = QHBoxLayout()
        width_layout.addWidget(QLabel("Ширина:"))
        width_combo = QComboBox()
        width_combo.addItems(["400", "600", "760", "1000", "Оригинал"])
        width_combo.setCurrentText("760")
        width_layout.addWidget(width_combo)
        width_layout.addStretch()
        block_layout.addLayout(width_layout)

        # Сохраняем тип блока и ссылки
        block_frame.block_type = "images"
        block_frame.img_list = img_list
        block_frame.width_combo = width_combo

        blocks_layout.addWidget(block_frame)
        parent_section = blocks_layout.parentWidget().parentWidget()
        if hasattr(parent_section, "refresh_header"):
            parent_section.refresh_header()

    def _delete_block(self, block_frame, blocks_layout):
        """Удаляет блок из секции."""
        blocks_layout.removeWidget(block_frame)
        block_frame.deleteLater()
        parent_section = blocks_layout.parentWidget().parentWidget()
        if hasattr(parent_section, "refresh_header"):
            parent_section.refresh_header()

    def _add_image_to_list(self, img_list):
        file_path, _ = QFileDialog.getOpenFileName(
            self, "Выберите изображение", "",
            "Изображения (*.png *.jpg *.jpeg *.bmp *.gif);;Все файлы (*)"
        )
        if file_path:
            img_list.addItem(file_path)

    def _remove_selected_image(self, img_list):
        row = img_list.currentRow()
        if row >= 0:
            img_list.takeItem(row)

    def _move_image(self, img_list, direction):
        row = img_list.currentRow()
        if row < 0:
            return
        new_row = row + direction
        if 0 <= new_row < img_list.count():
            item = img_list.takeItem(row)
            img_list.insertItem(new_row, item)
            img_list.setCurrentRow(new_row)

    def _move_section(self, frame, direction: int):
        idx = self.sections_layout.indexOf(frame)
        if idx == -1:
            return
        new_idx = idx + direction
        if 0 <= new_idx < self.sections_layout.count():
            self.sections_layout.removeWidget(frame)
            self.sections_layout.insertWidget(new_idx, frame)

    def _section_blocks_from_frame(self, frame):
        blocks = []
        if not hasattr(frame, 'blocks_layout'):
            return blocks

        for j in range(frame.blocks_layout.count()):
            block_frame = frame.blocks_layout.itemAt(j).widget()
            if block_frame is None:
                continue

            block_type = getattr(block_frame, 'block_type', None)
            if block_type == "text":
                text = block_frame.text_edit.toPlainText().strip() if hasattr(block_frame, 'text_edit') else ""
                if text:
                    blocks.append({"type": "text", "content": text})
            elif block_type == "images":
                image_paths = []
                if hasattr(block_frame, 'img_list'):
                    for k in range(block_frame.img_list.count()):
                        path = block_frame.img_list.item(k).text().strip()
                        if path:
                            image_paths.append(path)

                image_width = 760
                if hasattr(block_frame, 'width_combo'):
                    w = block_frame.width_combo.currentText()
                    image_width = 0 if w == "Оригинал" else int(w)

                if image_paths:
                    blocks.append({
                        "type": "images",
                        "paths": image_paths,
                        "image_width": image_width
                    })

        return blocks

    def _duplicate_section(self, frame):
        title = frame.title_edit.text().strip() if hasattr(frame, "title_edit") else "Секция"
        blocks = self._section_blocks_from_frame(frame)
        self.add_section(title=f"{title} копия")
        new_frame = self.sections_layout.itemAt(self.sections_layout.count() - 1).widget()

        while new_frame.blocks_layout.count() > 0:
            item = new_frame.blocks_layout.takeAt(0)
            if item.widget():
                item.widget().deleteLater()

        for block in blocks:
            if block.get("type") == "text":
                self._add_text_block(new_frame.blocks_layout)
                last_block = new_frame.blocks_layout.itemAt(new_frame.blocks_layout.count() - 1).widget()
                last_block.text_edit.setPlainText(block.get("content", ""))
            elif block.get("type") == "images":
                self._add_images_block(new_frame.blocks_layout)
                last_block = new_frame.blocks_layout.itemAt(new_frame.blocks_layout.count() - 1).widget()
                if hasattr(last_block, "width_combo"):
                    width = block.get("image_width", 760)
                    last_block.width_combo.setCurrentText("Оригинал" if width == 0 else str(width))
                if hasattr(last_block, "img_list"):
                    for path in block.get("paths", []):
                        last_block.img_list.addItem(path)

        new_frame.refresh_header()

    def _delete_section(self, frame):
        if self.sections_layout.count() <= 1:
            QMessageBox.warning(self, "Внимание", "Должна остаться хотя бы одна секция.")
            return
        self.sections_layout.removeWidget(frame)
        frame.deleteLater()

    def add_related(self):
        instruction_id = self.related_combo.currentData()
        title = self.related_combo.currentText().strip()
        if instruction_id is None or not title:
            return

        for i in range(self.related_list.count()):
            if self.related_list.item(i).data(Qt.UserRole) == instruction_id:
                QMessageBox.information(self, "Внимание", "Эта инструкция уже добавлена.")
                return

        item = QListWidgetItem(title)
        item.setData(Qt.UserRole, instruction_id)
        self.related_list.addItem(item)

    def remove_selected_related(self):
        row = self.related_list.currentRow()
        if row >= 0:
            self.related_list.takeItem(row)

    def _gather_sections(self):
        sections = []
        for i in range(self.sections_layout.count()):
            frame = self.sections_layout.itemAt(i).widget()
            if frame is None:
                continue

            title = frame.title_edit.text().strip() if hasattr(frame, 'title_edit') else ""
            blocks = []

            if hasattr(frame, 'blocks_layout'):
                for j in range(frame.blocks_layout.count()):
                    block_frame = frame.blocks_layout.itemAt(j).widget()
                    if block_frame is None:
                        continue

                    block_type = getattr(block_frame, 'block_type', None)

                    if block_type == "text":
                        text = block_frame.text_edit.toPlainText().strip() if hasattr(block_frame, 'text_edit') else ""
                        if text:
                            blocks.append({
                                "type": "text",
                                "content": text  # сохраняем как есть, markdown применится при отображении
                            })

                    elif block_type == "images":
                        image_paths = []
                        if hasattr(block_frame, 'img_list'):
                            for k in range(block_frame.img_list.count()):
                                path = block_frame.img_list.item(k).text().strip()
                                if path:
                                    image_paths.append(path)
                        # Собираем выбранную ширину
                        image_width = None
                        if hasattr(block_frame, 'width_combo'):
                            w = block_frame.width_combo.currentText()
                            if w == "Оригинал":
                                image_width = 0  # 0 – не масштабировать
                            else:
                                image_width = int(w)
                        else:
                            image_width = 760  # default
                        if image_paths:
                            blocks.append({
                                "type": "images",
                                "paths": image_paths,
                                "image_width": image_width
                            })

            if title or blocks:
                sections.append({
                    "id": getattr(frame, "section_id", None) or new_section_id(),
                    "title": title,
                    "blocks": blocks
                })
        return sections

    def validate_and_accept(self):
        if not self.title_edit.text().strip():
            QMessageBox.warning(self, "Ошибка", "Введите название инструкции.")
            return
        if self.task_combo.currentData() is None:
            QMessageBox.warning(self, "Ошибка", "Выберите подзадачу, к которой привязать инструкцию.")
            return
        sections = self._gather_sections()
        if not sections:
            QMessageBox.warning(self, "Ошибка", "Добавьте хотя бы одну секцию.")
            return
        self.accept()

    def get_data(self):
        sections = self._gather_sections()
        related_ids = []
        for i in range(self.related_list.count()):
            item = self.related_list.item(i)
            rel_id = item.data(Qt.UserRole)
            if rel_id is not None:
                related_ids.append(rel_id)

        return {
            "category_id": self.category_combo.currentData(),
            "task_id": self.task_combo.currentData(),
            "title": self.title_edit.text().strip(),
            "short_desc": self.short_desc_edit.text().strip(),
            "sections": sections,
            "related_ids": related_ids
        }

class CollapsibleSection(QFrame):
    def __init__(self, title, blocks=None, link_handler=None, section_index=None, feedback_handler=None):
        super().__init__()
        self.section_index = section_index
        self.link_handler = link_handler
        self.anchor_title = (title or "").strip() or f"Секция {(section_index or 0) + 1}"

        self.setObjectName("sectionCard")
        self._base_stylesheet = """
QFrame#sectionCard {
    background: #ffffff;
    border: 1px solid #d9e3f0;
    border-radius: 10px;
}

QFrame#sectionCard QToolButton {
    font-size: 15px;
    font-weight: bold;
    background: #f7f9fd;
    color: #263238;
    border-bottom: 1px solid #d9e3f0;
    border-top-left-radius: 10px;
    border-top-right-radius: 10px;
}

QFrame#sectionCard QToolButton:hover {
    background: #eef5fb;
}

QPushButton#sectionCommentButton {
    font-size: 11px;
    padding: 5px 12px;
    background: #FFE7B8;
    border: 1px solid #F5D78E;
    border-radius: 6px;
}

QPushButton#sectionCommentButton:hover {
    background: #FFD88A;
    border: 1px solid #F59E0B;
}

QFrame#sectionCard QWidget#qt_scrollarea_viewport,
QFrame#sectionCard QWidget {
    background: #ffffff;
}
        """

        self._active_stylesheet = """
QFrame#sectionCard {
    background: #fff8e8;
    border: 2px solid #f59e0b;
    border-radius: 10px;
}

QFrame#sectionCard QToolButton {
    font-size: 15px;
    font-weight: bold;
    background: #fff1d6;
    color: #263238;
    border-bottom: 1px solid #f5c46b;
    border-top-left-radius: 10px;
    border-top-right-radius: 10px;
}

QFrame#sectionCard QToolButton:hover {
    background: #ffe7b8;
}

QPushButton#sectionCommentButton {
    font-size: 11px;
    padding: 5px 12px;
    background: #FFE7B8;
    border: 1px solid #F5D78E;
    border-radius: 6px;
}

QPushButton#sectionCommentButton:hover {
    background: #FFD88A;
    border: 1px solid #F59E0B;
}

QFrame#sectionCard QWidget#qt_scrollarea_viewport,
QFrame#sectionCard QWidget {
    background: #fff8e8;
}
        """

        self.setStyleSheet(self._base_stylesheet)

        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(0)

        self.toggle = QToolButton(text=title)
        self.toggle.setCheckable(True)
        self.toggle.setChecked(False)
        self.toggle.setToolButtonStyle(Qt.ToolButtonTextBesideIcon)
        self.toggle.setArrowType(Qt.RightArrow)
        self.toggle.toggled.connect(self.on_toggled)

        layout.addWidget(self.toggle)

        self.content = QWidget()
        content_layout = QVBoxLayout(self.content)
        content_layout.setContentsMargins(12, 8, 12, 12)
        content_layout.setSpacing(8)
        self._responsive_images = []

        # Отображаем блоки
        for block in (blocks or []):
            block_type = block.get("type")
            if block_type == "text":
                raw_text = block.get("content", "")
                body_html = render_markdown(raw_text)
                body_label = QLabel()
                body_label.setWordWrap(True)
                body_label.setTextFormat(Qt.RichText)
                body_label.setTextInteractionFlags(Qt.TextBrowserInteraction)
                body_label.linkActivated.connect(self.on_link_activated)
                body_label.setSizePolicy(QSizePolicy.Ignored, QSizePolicy.Preferred)
                body_label.setText(body_html)
                content_layout.addWidget(body_label)

            elif block_type == "images":
                image_width = block.get("image_width", 760) or 760
                for image_path in block.get("paths", []):
                    source = QPixmap(str(image_path))
                    if source.isNull():
                        continue
                    image_widget = InstructionImageWidget(image_path, source)
                    self._responsive_images.append({
                        "widget": image_widget,
                        "source": source,
                        "preferred_width": 0 if image_width == 0 else int(image_width),
                    })
                    content_layout.addWidget(image_widget)

        self.content.setVisible(False)
        layout.addWidget(self.content)

        self.footer = None
        if feedback_handler is not None:
            self.footer = QWidget()
            footer_layout = QHBoxLayout(self.footer)
            footer_layout.setContentsMargins(12, 4, 12, 10)
            footer_layout.addStretch()
            self.comment_button = QPushButton("Комментарий")
            self.comment_button.setObjectName("sectionCommentButton")
            self.comment_button.setCursor(Qt.PointingHandCursor)
            self.comment_button.setToolTip("Оставить комментарий и оценку к этому блоку")
            self.comment_button.clicked.connect(
                lambda: feedback_handler(self.anchor_title)
            )
            footer_layout.addWidget(self.comment_button, 0, Qt.AlignRight)
            layout.addWidget(self.footer)
            self.footer.setVisible(False)

        self.set_active(False)

    def update_images_width(self, container_width):
        """Масштабирует картинки под доступную ширину центральной колонки."""
        available = max(120, int(container_width) - 48)
        for entry in self._responsive_images:
            preferred = entry.get("preferred_width") or 0
            target = available if preferred == 0 else min(preferred, available)
            entry["widget"].apply_source_width(entry["source"], target)

    def set_active(self, active: bool):
        self.setStyleSheet(self._active_stylesheet if active else self._base_stylesheet)

    def on_toggled(self, checked):
        self.content.setVisible(checked)
        self.toggle.setArrowType(Qt.DownArrow if checked else Qt.RightArrow)
        if self.footer is not None:
            self.footer.setVisible(checked)

    def on_link_activated(self, link):
        if self.link_handler:
            self.link_handler(link)


class TaskEditorDialog(QDialog):
    """Диалог редактирования задачи: описание, инструкция и секционный чек-лист."""

    def __init__(self, db: KnowledgeBaseDB, task, instruction, parent=None):
        super().__init__(parent)
        self.db = db
        self.current_instruction_id = None
        if instruction:
            self.current_instruction_id = instruction.get("instruction_id") or instruction.get("instruction_real_id")
        if self.current_instruction_id is None:
            self.current_instruction_id = task.get("instruction_real_id") or task.get("instruction_id")

        self.current_instruction_title = ""
        if instruction:
            self.current_instruction_title = instruction.get("instruction_title", "")
        if not self.current_instruction_title:
            self.current_instruction_title = task.get("instruction_title", "")

        self.current_task_id = task["task_id"]
        self.checklist_targets = self.db.all_checklist_targets()
        self.checklist_target_by_task_id = {
            int(target["task_id"]): target
            for target in self.checklist_targets
            if target.get("task_id") is not None
        }
        if self.current_task_id not in self.checklist_target_by_task_id and self.current_instruction_id:
            self.checklist_target_by_task_id[self.current_task_id] = {
                "task_id": self.current_task_id,
                "task_title": task.get("task_title", ""),
                "category_id": task.get("category_id"),
                "category_name": task.get("category_name", ""),
                "category_path": self.db.category_path(task.get("category_id")),
                "instruction_id": self.current_instruction_id,
                "instruction_title": self.current_instruction_title,
                "sections": normalize_instruction_sections(
                    instruction.get("sections", []) if instruction else task.get("sections", [])
                ),
            }
            self.checklist_targets.append(self.checklist_target_by_task_id[self.current_task_id])
        self._applying_section_target = False

        self.setWindowTitle("Редактирование подзадачи")
        self.setModal(True)
        self.resize(880, 760)

        layout = QVBoxLayout(self)
        layout.setSpacing(12)

        # Описание подзадачи
        layout.addWidget(QLabel("Описание подзадачи (поддерживается Markdown):"))
        self.desc_edit = TypographicPlainTextEdit()
        self.desc_edit.setPlaceholderText("Описание задачи...")
        self.desc_edit.setMinimumHeight(150)
        self.desc_edit.setPlainText(task.get("short_desc", ""))
        desc_btn_row = QHBoxLayout()
        desc_btn_row.addStretch()
        desc_expand_btn = QPushButton("Редактировать описание в большом окне")
        desc_expand_btn.clicked.connect(lambda: open_large_text_editor(self, self.desc_edit, "Описание подзадачи"))
        desc_btn_row.addWidget(desc_expand_btn)
        layout.addWidget(self.desc_edit)
        layout.addLayout(desc_btn_row)

        # Название инструкции
        form = QFormLayout()
        self.instruction_edit = TypographicLineEdit()
        self.instruction_edit.setPlaceholderText("Название инструкции")
        if instruction:
            self.instruction_edit.setText(instruction.get("instruction_title", ""))
        else:
            self.instruction_edit.setText(task.get("instruction_title", ""))
        form.addRow("Инструкция:", self.instruction_edit)
        layout.addLayout(form)

        # Чек-лист
        checklist_group = QGroupBox("Чек-лист задачи")
        checklist_layout = QVBoxLayout(checklist_group)

        self.checklist_sections_widget = QWidget()
        self.checklist_sections_layout = QVBoxLayout(self.checklist_sections_widget)
        self.checklist_sections_layout.setContentsMargins(0, 0, 0, 0)
        self.checklist_sections_layout.setSpacing(10)

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setWidget(self.checklist_sections_widget)
        checklist_layout.addWidget(scroll)

        hint = QLabel(
            "Секция = подзадача. Внутри секции — пункты чек-листа. "
            "У каждого пункта можно указать, куда вести переход: "
            "в текущую инструкцию или в другую инструкцию."
        )
        hint.setWordWrap(True)
        hint.setStyleSheet("color: #5b6577; font-size: 11px;")
        checklist_layout.addWidget(hint)

        btn_row = QHBoxLayout()
        add_section_btn = QPushButton("Добавить секцию")
        add_section_btn.clicked.connect(lambda: self.add_checklist_section())
        btn_row.addWidget(add_section_btn)
        add_many_sections_btn = QPushButton("Создать секции из списка")
        add_many_sections_btn.clicked.connect(self.add_checklist_sections_from_list)
        btn_row.addWidget(add_many_sections_btn)
        collapse_btn = QPushButton("Свернуть все")
        collapse_btn.clicked.connect(lambda: set_all_editor_sections_expanded(self.checklist_sections_layout, False))
        btn_row.addWidget(collapse_btn)
        expand_btn = QPushButton("Развернуть все")
        expand_btn.clicked.connect(lambda: set_all_editor_sections_expanded(self.checklist_sections_layout, True))
        btn_row.addWidget(expand_btn)
        btn_row.addStretch()
        checklist_layout.addLayout(btn_row)

        layout.addWidget(checklist_group)

        # Кнопки
        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        buttons.accepted.connect(self.validate_and_accept)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)

        # Заполнение существующими данными
        sections = normalize_checklist_sections(task.get("checklist_sections") or task.get("checklist") or [])
        if sections:
            for sec in sections:
                self.add_checklist_section(sec.get("title", ""), sec.get("items", []), sec)
        else:
            self.add_checklist_section("Подзадача 1")

    def _checklist_section_summary(self, frame):
        count = frame.items_layout.count() if hasattr(frame, "items_layout") else 0
        return f"{count} пункт(ов)"

    def add_checklist_sections_from_list(self):
        titles = lines_from_multiline_input(
            self,
            "Создать секции чек-листа",
            "Вставьте названия секций, каждое с новой строки:"
        )
        for title in titles:
            self.add_checklist_section(title)

    def add_checklist_section(self, title="", items=None, section_data=None):
        if isinstance(title, bool):
            title = ""

        frame = EditorSectionFrame(parent=self)
        layout = frame.content_layout

        title_edit = TypographicLineEdit()
        title_edit.setPlaceholderText("Название секции / подзадачи")
        title_edit.setText(str(title or ""))
        title_edit.textChanged.connect(frame.refresh_header)
        layout.addWidget(title_edit)

        section_target_row = QHBoxLayout()
        section_target_row.addWidget(QLabel("Привязать всю секцию к подзадаче:"))

        section_task_combo = NoWheelComboBox()
        section_task_combo.setToolTip(
            "Подзадача, к которой ведут пункты этой секции. "
            "При смене привязки обновляются все пункты без отдельной настройки."
        )
        self._fill_checklist_task_combo(section_task_combo)

        section_section_combo = NoWheelComboBox()
        section_section_combo.setToolTip(
            "Секция инструкции для всех пунктов этой секции. "
            "При смене привязки обновляются все пункты, у которых не задана отдельная привязка."
        )
        section_target_row.addWidget(section_task_combo, 1)
        section_target_row.addWidget(section_section_combo, 1)
        layout.addLayout(section_target_row)

        items_widget = QWidget()
        items_layout = QVBoxLayout(items_widget)
        items_layout.setContentsMargins(0, 0, 0, 0)
        items_layout.setSpacing(6)
        layout.addWidget(items_widget)

        item_btn_row = QHBoxLayout()
        add_item_btn = QPushButton("+ Пункт")
        add_item_btn.clicked.connect(lambda: self._add_checklist_item(items_layout))
        item_btn_row.addWidget(add_item_btn)
        add_many_items_btn = QPushButton("Создать пункты из списка")
        add_many_items_btn.clicked.connect(lambda: self.add_checklist_items_from_list(items_layout))
        item_btn_row.addWidget(add_many_items_btn)
        item_btn_row.addStretch()
        layout.addLayout(item_btn_row)

        up_btn = QPushButton("↑")
        up_btn.setToolTip("Переместить секцию выше")
        up_btn.clicked.connect(lambda: self._move_section(frame, -1))
        down_btn = QPushButton("↓")
        down_btn.setToolTip("Переместить секцию ниже")
        down_btn.clicked.connect(lambda: self._move_section(frame, 1))
        duplicate_btn = QPushButton("Дублировать")
        duplicate_btn.clicked.connect(lambda: self._duplicate_checklist_section(frame))
        del_btn = QPushButton("Удалить секцию")
        del_btn.clicked.connect(lambda: self._delete_section(frame))

        frame.header_controls.addWidget(up_btn)
        frame.header_controls.addWidget(down_btn)
        frame.header_controls.addWidget(duplicate_btn)
        frame.header_controls.addWidget(del_btn)

        frame.section_title_edit = title_edit
        frame.section_task_combo = section_task_combo
        frame.section_section_combo = section_section_combo
        frame.items_layout = items_layout
        frame.title_getter = lambda f=frame: f.section_title_edit.text()
        frame.summary_getter = lambda f=frame: self._checklist_section_summary(f)
        section_task_combo.currentIndexChanged.connect(
            lambda _, f=frame: self._on_section_task_changed(f)
        )
        section_section_combo.currentIndexChanged.connect(
            lambda _, f=frame: self._on_section_target_changed(f)
        )

        self.checklist_sections_layout.addWidget(frame)

        if section_data:
            self._resolve_task_combo_selection(section_task_combo, section_data)

        self._populate_checklist_section_combo(
            section_task_combo,
            section_section_combo,
            str((section_data or {}).get("target_section_id") or "").strip(),
            str((section_data or {}).get("section_title") or "").strip()
        )

        if items:
            for item in items:
                item_for_editor = dict(item)
                overridden = self._checklist_item_overrides_section(item_for_editor, section_data)
                self._add_checklist_item(
                    items_layout,
                    item_for_editor,
                    section_frame=frame,
                    target_overridden=overridden
                )
        else:
            self._add_checklist_item(items_layout, section_frame=frame)

        frame.refresh_header()

    def add_checklist_items_from_list(self, items_layout):
        lines = lines_from_multiline_input(
            self,
            "Создать пункты чек-листа",
            "Вставьте пункты чек-листа, каждый с новой строки:"
        )
        for line in lines:
            self._add_checklist_item(items_layout, {"text": line})

    def _checklist_target_label(self, target):
        path_parts = list(target.get("category_path") or [])
        task_title = str(target.get("task_title") or "").strip()
        if task_title:
            path_parts.append(task_title)
        label = " / ".join(part for part in path_parts if part)
        instruction_title = str(target.get("instruction_title") or "").strip()
        if instruction_title and instruction_title.casefold() != task_title.casefold():
            label = f"{label} — {instruction_title}" if label else instruction_title
        return label or f"Подзадача #{target.get('task_id')}"

    def _fill_checklist_task_combo(self, combo, selected_task_id=None):
        combo.blockSignals(True)
        combo.clear()
        combo.addItem("(эта подзадача)", self.current_task_id)
        for target in self.checklist_targets:
            task_id = target.get("task_id")
            if task_id == self.current_task_id:
                continue
            combo.addItem(self._checklist_target_label(target), task_id)

        if selected_task_id:
            try:
                selected_task_id = int(selected_task_id)
            except (TypeError, ValueError):
                selected_task_id = None
            if selected_task_id:
                idx = combo.findData(selected_task_id)
                if idx >= 0:
                    combo.setCurrentIndex(idx)
        combo.blockSignals(False)

    def _selected_target_task_id(self, task_combo):
        value = task_combo.currentData()
        if value in (None,):
            return self.current_task_id
        try:
            return int(value)
        except (TypeError, ValueError):
            return self.current_task_id

    def _is_current_task_selected(self, task_combo):
        return self._selected_target_task_id(task_combo) == self.current_task_id

    def _target_from_task_id(self, task_id):
        if task_id in (None, self.current_task_id):
            return {
                "target_task_id": None,
                "target_instruction_id": None,
                "instruction_title": "",
                "task_title": "",
            }
        target = self.checklist_target_by_task_id.get(task_id, {})
        return {
            "target_task_id": task_id,
            "target_instruction_id": target.get("instruction_id"),
            "instruction_title": str(target.get("instruction_title") or "").strip(),
            "task_title": str(target.get("task_title") or "").strip(),
        }

    def _resolve_task_combo_selection(self, task_combo, item_data):
        target_task_id = item_data.get("target_task_id")
        try:
            target_task_id = int(target_task_id) if target_task_id else None
        except (TypeError, ValueError):
            target_task_id = None

        if target_task_id:
            idx = task_combo.findData(target_task_id)
            if idx >= 0:
                task_combo.setCurrentIndex(idx)
                return

        target_instruction_id = item_data.get("target_instruction_id")
        try:
            target_instruction_id = int(target_instruction_id) if target_instruction_id else None
        except (TypeError, ValueError):
            target_instruction_id = None

        if target_instruction_id and target_instruction_id != self.current_instruction_id:
            for idx in range(task_combo.count()):
                task_id = task_combo.itemData(idx)
                target = self.checklist_target_by_task_id.get(task_id)
                if target and target.get("instruction_id") == target_instruction_id:
                    task_combo.setCurrentIndex(idx)
                    return

    def _populate_checklist_section_combo(
            self,
            task_combo,
            section_combo,
            selected_section_id="",
            selected_section_title=""
    ):
        section_combo.blockSignals(True)
        section_combo.clear()
        section_combo.addItem("(авто / первая подходящая)", "")

        task_id = self._selected_target_task_id(task_combo)
        target = self.checklist_target_by_task_id.get(task_id)
        selected_section_id = str(selected_section_id or "").strip()
        selected_title = str(selected_section_title or "").casefold().strip()
        selected_index = 0

        if target:
            for section_data in target.get("sections", []):
                section_id = str(section_data.get("id") or "").strip()
                section_title = str(section_data.get("title") or "").strip()
                if not section_title:
                    continue

                section_combo.addItem(section_title, section_id)
                row = section_combo.count() - 1
                if selected_section_id and section_id == selected_section_id:
                    selected_index = row
                elif not selected_section_id and selected_title and section_title.casefold().strip() == selected_title:
                    selected_index = row

        if selected_index == 0 and selected_title:
            fallback_title = str(selected_section_title or "").strip()
            section_combo.addItem(f"{fallback_title} (старое название)", "")
            selected_index = section_combo.count() - 1

        section_combo.setCurrentIndex(selected_index)
        section_combo.blockSignals(False)

    def _checklist_item_overrides_section(self, item_data, section_data):
        return checklist_item_overrides_section(
            item_data,
            section_data,
            self.current_instruction_id,
            self.current_task_id
        )

    def _on_section_task_changed(self, section_frame):
        if hasattr(section_frame, "section_task_combo") and hasattr(section_frame, "section_section_combo"):
            self._populate_checklist_section_combo(
                section_frame.section_task_combo,
                section_frame.section_section_combo
            )
        self._on_section_target_changed(section_frame)

    def _on_section_target_changed(self, section_frame):
        if self._applying_section_target or not hasattr(section_frame, "items_layout"):
            return

        section_target = self._checklist_section_target_from_frame(section_frame)
        self._applying_section_target = True
        try:
            for i in range(section_frame.items_layout.count()):
                item_frame = section_frame.items_layout.itemAt(i).widget()
                if item_frame is None or getattr(item_frame, "target_overridden", False):
                    continue
                self._apply_target_to_item_frame(item_frame, section_target)
        finally:
            self._applying_section_target = False

    def _mark_item_target_overridden(self, item_frame):
        if self._applying_section_target:
            return
        item_frame.target_overridden = True
        self._update_item_target_hint(item_frame)

    def _apply_target_to_item_frame(self, item_frame, target):
        task_combo = item_frame.task_combo
        section_combo = item_frame.section_combo

        task_combo.blockSignals(True)
        target_task_id = target.get("target_task_id")
        if target_task_id:
            idx = task_combo.findData(int(target_task_id))
            task_combo.setCurrentIndex(idx if idx >= 0 else 0)
        else:
            task_combo.setCurrentIndex(0)
        task_combo.blockSignals(False)

        self._populate_checklist_section_combo(
            task_combo,
            section_combo,
            str(target.get("target_section_id") or "").strip(),
            str(target.get("section_title") or "").strip()
        )
        self._update_item_target_hint(item_frame)

    def _update_item_target_hint(self, item_frame):
        if not hasattr(item_frame, "target_hint_label"):
            return
        if getattr(item_frame, "target_overridden", False):
            item_frame.target_hint_label.setText("Отдельная привязка пункта")
            item_frame.target_hint_label.setStyleSheet("color: #5b6577; font-size: 11px;")
        else:
            item_frame.target_hint_label.setText("Как у секции (меняется вместе с секцией)")
            item_frame.target_hint_label.setStyleSheet("color: #2f6fed; font-size: 11px;")

    def _add_checklist_item(self, items_layout, item_data=None, section_frame=None, target_overridden=None):
        frame = QFrame()
        frame.setFrameStyle(QFrame.StyledPanel | QFrame.Plain)
        frame.setStyleSheet("QFrame { background: #ffffff; border: 1px solid #e0e0e0; border-radius: 4px; }")

        layout = QVBoxLayout(frame)
        layout.setContentsMargins(6, 6, 6, 6)
        layout.setSpacing(4)

        header = QHBoxLayout()
        header.addWidget(QLabel("📝 Пункт"))
        header.addStretch()

        up_btn = QPushButton("↑")
        up_btn.setFixedSize(24, 24)
        up_btn.setToolTip("Переместить пункт выше")
        up_btn.clicked.connect(lambda _, f=frame, lay=items_layout: self._move_item(lay, f, -1))
        header.addWidget(up_btn)

        down_btn = QPushButton("↓")
        down_btn.setFixedSize(24, 24)
        down_btn.setToolTip("Переместить пункт ниже")
        down_btn.clicked.connect(lambda _, f=frame, lay=items_layout: self._move_item(lay, f, 1))
        header.addWidget(down_btn)

        del_btn = QPushButton("✕")
        del_btn.setFixedSize(24, 24)
        del_btn.setToolTip("Удалить пункт")
        del_btn.clicked.connect(lambda: self._delete_item(frame, items_layout))
        header.addWidget(del_btn)

        layout.addLayout(header)

        text_edit = TypographicPlainTextEdit()
        text_edit.setPlaceholderText("Текст пункта чек-листа")
        text_edit.setMinimumHeight(70)
        layout.addWidget(text_edit)

        text_btn_row = QHBoxLayout()
        text_btn_row.addStretch()
        expand_text_btn = QPushButton("Развернуть текст пункта")
        expand_text_btn.clicked.connect(lambda: open_large_text_editor(self, text_edit, "Пункт чек-листа"))
        text_btn_row.addWidget(expand_text_btn)
        layout.addLayout(text_btn_row)

        target_row = QHBoxLayout()

        target_row.addWidget(QLabel("Подзадача:"))

        task_combo = NoWheelComboBox()
        task_combo.setToolTip("Подзадача, к которой ведёт пункт чек-листа.")
        self._fill_checklist_task_combo(task_combo)

        section_combo = NoWheelComboBox()
        section_combo.setToolTip("Секция инструкции выбранной подзадачи.")

        target_row.addWidget(task_combo, 1)
        target_row.addWidget(section_combo, 1)
        layout.addLayout(target_row)

        target_hint_row = QHBoxLayout()
        reset_target_btn = QPushButton("Как у секции")
        reset_target_btn.setToolTip("Сбросить отдельную привязку и использовать привязку секции")
        target_hint_row.addWidget(reset_target_btn)
        target_hint_label = QLabel()
        target_hint_row.addWidget(target_hint_label, 1)
        layout.addLayout(target_hint_row)

        frame.text_edit = text_edit
        frame.task_combo = task_combo
        frame.section_combo = section_combo
        frame.target_hint_label = target_hint_label
        frame.target_overridden = bool(target_overridden)

        if section_frame is None:
            section_frame = items_layout.parentWidget().parentWidget()

        reset_target_btn.clicked.connect(
            lambda _, f=frame, s=section_frame: self._reset_item_target_to_section(f, s)
        )
        task_combo.currentIndexChanged.connect(
            lambda _, f=frame: self._mark_item_target_overridden(f)
        )
        section_combo.currentIndexChanged.connect(
            lambda _, f=frame: self._mark_item_target_overridden(f)
        )

        self._applying_section_target = True
        try:
            if item_data and target_overridden:
                text_edit.setPlainText(str(item_data.get("text", "")).strip())
                self._resolve_task_combo_selection(task_combo, item_data)
                selected_section_id = str(item_data.get("target_section_id") or "").strip()
                selected_section_title = str(item_data.get("section_title") or "").strip()
                self._populate_checklist_section_combo(
                    task_combo,
                    section_combo,
                    selected_section_id,
                    selected_section_title
                )
            elif item_data:
                text_edit.setPlainText(str(item_data.get("text", "")).strip())
                if section_frame is not None:
                    section_target = self._checklist_section_target_from_frame(section_frame)
                    self._apply_target_to_item_frame(frame, section_target)
                    frame.target_overridden = False
            elif section_frame is not None:
                section_target = self._checklist_section_target_from_frame(section_frame)
                self._apply_target_to_item_frame(frame, section_target)
                frame.target_overridden = False
        finally:
            self._applying_section_target = False

        task_combo.currentIndexChanged.connect(
            lambda _, combo=task_combo, sec_combo=section_combo:
            self._populate_checklist_section_combo(combo, sec_combo)
        )

        self._update_item_target_hint(frame)
        items_layout.addWidget(frame)
        parent_section = items_layout.parentWidget().parentWidget()
        if hasattr(parent_section, "refresh_header"):
            parent_section.refresh_header()

    def _reset_item_target_to_section(self, item_frame, section_frame):
        if section_frame is None:
            return
        item_frame.target_overridden = False
        section_target = self._checklist_section_target_from_frame(section_frame)
        self._applying_section_target = True
        try:
            self._apply_target_to_item_frame(item_frame, section_target)
        finally:
            self._applying_section_target = False
        self._update_item_target_hint(item_frame)

    def _move_section(self, frame, direction: int):
        idx = self.checklist_sections_layout.indexOf(frame)
        if idx == -1:
            return
        new_idx = idx + direction
        if 0 <= new_idx < self.checklist_sections_layout.count():
            self.checklist_sections_layout.removeWidget(frame)
            self.checklist_sections_layout.insertWidget(new_idx, frame)

    def _checklist_item_data_from_frame(self, item_frame):
        text = item_frame.text_edit.toPlainText().strip() if hasattr(item_frame, "text_edit") else ""
        if not getattr(item_frame, "target_overridden", False):
            return {
                "text": text,
                "target_task_id": None,
                "target_instruction_id": None,
                "target_section_id": "",
                "task_title": "",
                "instruction_title": "",
                "section_title": ""
            }

        task_id = None
        if hasattr(item_frame, "task_combo") and not self._is_current_task_selected(item_frame.task_combo):
            task_id = self._selected_target_task_id(item_frame.task_combo)

        base = self._target_from_task_id(task_id)

        target_section_id = ""
        section_title = ""
        if hasattr(item_frame, "section_combo"):
            target_section_id = str(item_frame.section_combo.currentData() or "").strip()
            if item_frame.section_combo.currentIndex() > 0:
                section_title = item_frame.section_combo.currentText().replace(" (старое название)", "").strip()

        return {
            "text": text,
            **base,
            "target_section_id": target_section_id,
            "section_title": section_title
        }

    def _checklist_section_target_from_frame(self, frame):
        task_id = None
        if hasattr(frame, "section_task_combo") and not self._is_current_task_selected(frame.section_task_combo):
            task_id = self._selected_target_task_id(frame.section_task_combo)

        base = self._target_from_task_id(task_id)

        target_section_id = ""
        section_title = ""
        if hasattr(frame, "section_section_combo"):
            target_section_id = str(frame.section_section_combo.currentData() or "").strip()
            if frame.section_section_combo.currentIndex() > 0:
                section_title = frame.section_section_combo.currentText().replace(" (старое название)", "").strip()

        return {
            **base,
            "target_section_id": target_section_id,
            "section_title": section_title
        }

    def _duplicate_checklist_section(self, frame):
        title = frame.section_title_edit.text().strip() if hasattr(frame, "section_title_edit") else "Секция"
        section_target = self._checklist_section_target_from_frame(frame)
        items = []
        if hasattr(frame, "items_layout"):
            for i in range(frame.items_layout.count()):
                item_frame = frame.items_layout.itemAt(i).widget()
                if item_frame is not None:
                    item_data = self._checklist_item_data_from_frame(item_frame)
                    if item_data.get("text"):
                        items.append(item_data)
        self.add_checklist_section(f"{title} копия", items, section_target)

    def _delete_section(self, frame):
        if self.checklist_sections_layout.count() <= 1:
            QMessageBox.warning(self, "Внимание", "Должна остаться хотя бы одна секция.")
            return
        self.checklist_sections_layout.removeWidget(frame)
        frame.deleteLater()

    def _move_item(self, items_layout, frame, direction: int):
        idx = items_layout.indexOf(frame)
        if idx == -1:
            return
        new_idx = idx + direction
        if 0 <= new_idx < items_layout.count():
            items_layout.removeWidget(frame)
            items_layout.insertWidget(new_idx, frame)

    def _delete_item(self, frame, items_layout):
        items_layout.removeWidget(frame)
        frame.deleteLater()
        parent_section = items_layout.parentWidget().parentWidget()
        if hasattr(parent_section, "refresh_header"):
            parent_section.refresh_header()

    def _gather_checklist_sections(self):
        sections = []

        for i in range(self.checklist_sections_layout.count()):
            frame = self.checklist_sections_layout.itemAt(i).widget()
            if frame is None:
                continue

            title = frame.section_title_edit.text().strip() if hasattr(frame, "section_title_edit") else ""
            items = []
            section_target = self._checklist_section_target_from_frame(frame)

            if hasattr(frame, "items_layout"):
                for j in range(frame.items_layout.count()):
                    item_frame = frame.items_layout.itemAt(j).widget()
                    if item_frame is None:
                        continue

                    item_data = self._checklist_item_data_from_frame(item_frame)
                    if item_data.get("text"):
                        items.append(item_data)

            if title or items:
                sections.append({
                    "title": title,
                    "target_task_id": section_target.get("target_task_id"),
                    "target_instruction_id": section_target.get("target_instruction_id"),
                    "target_section_id": section_target.get("target_section_id", ""),
                    "task_title": section_target.get("task_title", ""),
                    "instruction_title": section_target.get("instruction_title", ""),
                    "section_title": section_target.get("section_title", ""),
                    "items": items
                })

        return sections

    def validate_and_accept(self):
        desc = self.desc_edit.toPlainText().strip()
        instruction_title = self.instruction_edit.text().strip()
        checklist_sections = self._gather_checklist_sections()

        if not desc:
            QMessageBox.warning(self, "Ошибка", "Введите описание.")
            return
        if not instruction_title:
            QMessageBox.warning(self, "Ошибка", "Введите название инструкции.")
            return
        if not checklist_sections:
            QMessageBox.warning(self, "Ошибка", "Добавьте хотя бы одну секцию чек-листа.")
            return

        for sec in checklist_sections:
            if not sec.get("title", "").strip():
                QMessageBox.warning(self, "Ошибка", "У каждой секции должно быть название.")
                return
            if not sec.get("items"):
                QMessageBox.warning(self, "Ошибка", f"Секция «{sec['title']}» не может быть пустой.")
                return

        self.accept()

    def get_data(self):
        checklist_sections = self._gather_checklist_sections()
        return {
            "short_desc": self.desc_edit.toPlainText().strip(),
            "instruction_title": self.instruction_edit.text().strip(),
            "checklist_sections": checklist_sections,
            "checklist": [item["text"] for item in flatten_checklist_sections(checklist_sections)]
        }

class SectionFeedbackDialog(QDialog):
    """Комментарий и оценка к конкретному блоку инструкции."""

    def __init__(self, anchor_title, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Комментарий к блоку")
        self.setModal(True)
        self.resize(480, 360)

        layout = QVBoxLayout(self)
        layout.setSpacing(10)

        block_label = QLabel(f"Блок: {anchor_title}")
        block_label.setWordWrap(True)
        block_label.setStyleSheet("font-weight: 600; color: #92400e;")
        layout.addWidget(block_label)

        note = QLabel("Оценка анонимная. Для оценки ниже 7 нужен комментарий.")
        note.setWordWrap(True)
        note.setObjectName("hintBanner")
        layout.addWidget(note)

        rating_row = QHBoxLayout()
        rating_row.addWidget(QLabel("Оценка (1–10):"))
        self.rating_spin = QSpinBox()
        self.rating_spin.setRange(1, 10)
        self.rating_spin.setValue(10)
        self.rating_spin.setMaximumWidth(90)
        rating_row.addWidget(self.rating_spin)
        rating_row.addStretch()
        layout.addLayout(rating_row)

        layout.addWidget(QLabel("Комментарий (необязательно):"))
        self.comment_text_edit = QTextEdit()
        self.comment_text_edit.setPlaceholderText("Что непонятно? Что нужно поправить или доработать?")
        self.comment_text_edit.setMinimumHeight(120)
        layout.addWidget(self.comment_text_edit)

        author_row = QHBoxLayout()
        author_row.addWidget(QLabel("Имя:"))
        self.author_edit = QLineEdit()
        self.author_edit.setPlaceholderText("Необязательно")
        author_row.addWidget(self.author_edit, 1)
        layout.addLayout(author_row)

        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)

    def get_data(self):
        return {
            "rating": self.rating_spin.value(),
            "text": self.comment_text_edit.toPlainText().strip(),
            "author": self.author_edit.text().strip() or "Пользователь",
        }

    def accept(self):
        rating = self.rating_spin.value()
        text = self.comment_text_edit.toPlainText().strip()
        if rating < 7 and not text:
            QMessageBox.warning(
                self,
                "Требуется комментарий",
                "Для оценки ниже 7 необходимо оставить комментарий.\n"
                "Опиши, что именно непонятно или требует доработки.",
            )
            return
        super().accept()


class FeedbackManagerDialog(QDialog):
    """Диалог управления оценками и комментариями (для администратора)."""

    def __init__(self, db: KnowledgeBaseDB, instruction_id, parent=None):
        super().__init__(parent)
        self.db = db
        self.instruction_id = instruction_id
        self.setWindowTitle("Управление оценками и комментариями")
        self.setModal(True)
        self.resize(750, 550)

        layout = QVBoxLayout(self)
        layout.setSpacing(10)

        # Вкладки: Оценки | Комментарии
        tabs = QTabWidget()

        # --- Оценки ---
        ratings_tab = QWidget()
        ratings_layout = QVBoxLayout(ratings_tab)
        self.ratings_list = QListWidget()
        self.ratings_list.setAlternatingRowColors(True)
        ratings_layout.addWidget(self.ratings_list)

        del_rating_btn = QPushButton("Удалить выбранную оценку")
        del_rating_btn.clicked.connect(self._delete_rating)
        ratings_layout.addWidget(del_rating_btn)

        tabs.addTab(ratings_tab, "Оценки")

        # --- Комментарии ---
        comments_tab = QWidget()
        comments_layout = QVBoxLayout(comments_tab)
        self.comments_list = QListWidget()
        self.comments_list.setAlternatingRowColors(True)
        comments_layout.addWidget(self.comments_list)

        comm_btn_layout = QHBoxLayout()
        edit_comm_btn = QPushButton("Редактировать")
        edit_comm_btn.clicked.connect(self._edit_comment)
        del_comm_btn = QPushButton("Удалить")
        del_comm_btn.clicked.connect(self._delete_comment)
        comm_btn_layout.addWidget(edit_comm_btn)
        comm_btn_layout.addWidget(del_comm_btn)
        comm_btn_layout.addStretch()
        comments_layout.addLayout(comm_btn_layout)

        tabs.addTab(comments_tab, "Комментарии")

        layout.addWidget(tabs)

        # Закрыть
        close_btn = QPushButton("Закрыть")
        close_btn.clicked.connect(self.close)
        layout.addWidget(close_btn, alignment=Qt.AlignRight)

        self._load_data()

    def _load_data(self):
        # Оценки
        self.ratings_list.clear()
        ratings = self.db.ratings_for_instruction(self.instruction_id)
        for r in ratings:
            anchor = (r.get("anchor") or "").strip()
            anchor_suffix = f"  [{anchor}]" if anchor else "  [вся инструкция]"
            item = QListWidgetItem(f"{r['rating']} / 10{anchor_suffix}  —  {r['created_at']}")
            item.setData(Qt.UserRole, r["id"])
            self.ratings_list.addItem(item)

        # Комментарии
        self.comments_list.clear()
        comments = self.db.comments_for_instruction(self.instruction_id)
        for c in comments:
            text = c["text"][:100] + "…" if len(c["text"]) > 100 else c["text"]
            rating_suffix = ""
            if c.get("rating") is not None:
                rating_suffix = f"  [{c['rating']}/10]"
            item = QListWidgetItem(f"[{c['created_at']}]{rating_suffix} {c['author']}: {text}")
            item.setData(Qt.UserRole, c["id"])
            self.comments_list.addItem(item)

    def _delete_rating(self):
        item = self.ratings_list.currentItem()
        if not item:
            QMessageBox.warning(self, "Внимание", "Выберите оценку.")
            return
        if QMessageBox.question(self, "Подтверждение", "Удалить оценку?") == QMessageBox.Yes:
            self.db.delete_rating(item.data(Qt.UserRole))
            self._load_data()

    def _edit_comment(self):
        item = self.comments_list.currentItem()
        if not item:
            QMessageBox.warning(self, "Внимание", "Выберите комментарий.")
            return
        comment = self.db.comment_by_id(item.data(Qt.UserRole))
        if not comment:
            return
        text, ok = QInputDialog.getMultiLineText(self, "Редактировать", "Текст:", comment["text"])
        if ok and text.strip():
            self.db.update_comment(comment["id"], text.strip())
            self._load_data()

    def _delete_comment(self):
        item = self.comments_list.currentItem()
        if not item:
            QMessageBox.warning(self, "Внимание", "Выберите комментарий.")
            return
        if QMessageBox.question(self, "Подтверждение", "Удалить комментарий?") == QMessageBox.Yes:
            self.db.delete_comment(item.data(Qt.UserRole))
            self._load_data()


class OnboardingOverlay(QWidget):
    """Полноэкранная подсветка элементов интерфейса с пошаговыми подсказками."""

    def __init__(self, main_window, steps, on_completed=None, parent=None):
        super().__init__(parent or main_window)
        self._main_window = main_window
        self._steps = steps
        self._on_completed = on_completed
        self._step_index = 0
        self._highlight_rect = QRect()

        self.setAttribute(Qt.WA_TransparentForMouseEvents, False)
        self.setFocusPolicy(Qt.StrongFocus)

        self._card = QFrame(self)
        self._card.setObjectName("onboardingCard")
        self._card.setStyleSheet("""
            QFrame#onboardingCard {
                background: #ffffff;
                border: 1px solid #c6d8ff;
                border-radius: 10px;
            }
            QFrame#onboardingCard QLabel#onboardingTitle {
                font-size: 14px;
                font-weight: bold;
                color: #1f2937;
            }
            QFrame#onboardingCard QLabel#onboardingText {
                font-size: 12px;
                color: #374151;
            }
            QFrame#onboardingCard QLabel#onboardingCounter {
                font-size: 11px;
                color: #6b7280;
            }
        """)
        shadow = QGraphicsDropShadowEffect(self._card)
        shadow.setBlurRadius(24)
        shadow.setOffset(0, 4)
        shadow.setColor(QColor(0, 0, 0, 60))
        self._card.setGraphicsEffect(shadow)

        card_layout = QVBoxLayout(self._card)
        card_layout.setContentsMargins(16, 14, 16, 14)
        card_layout.setSpacing(10)

        self._title_label = QLabel()
        self._title_label.setObjectName("onboardingTitle")
        self._title_label.setWordWrap(True)
        card_layout.addWidget(self._title_label)

        self._text_label = QLabel()
        self._text_label.setObjectName("onboardingText")
        self._text_label.setWordWrap(True)
        card_layout.addWidget(self._text_label)

        buttons_row = QHBoxLayout()
        buttons_row.setSpacing(8)

        self._counter_label = QLabel()
        self._counter_label.setObjectName("onboardingCounter")
        buttons_row.addWidget(self._counter_label, 1)

        self._skip_button = QPushButton("Пропустить")
        self._skip_button.setFlat(True)
        self._skip_button.clicked.connect(self._finish)
        buttons_row.addWidget(self._skip_button)

        self._back_button = QPushButton("Назад")
        self._back_button.clicked.connect(self._go_back)
        buttons_row.addWidget(self._back_button)

        self._next_button = QPushButton("Далее")
        self._next_button.setDefault(True)
        self._next_button.clicked.connect(self._go_next)
        buttons_row.addWidget(self._next_button)

        card_layout.addLayout(buttons_row)
        self._card.setFixedWidth(360)

    def start(self):
        self._fit_to_parent()
        self.show()
        self.raise_()
        self.setFocus()
        self._show_step(0)

    def refresh_current_step(self):
        if self.isVisible():
            self._apply_step_geometry()

    def _fit_to_parent(self):
        parent = self.parentWidget()
        if parent:
            self.setGeometry(parent.rect())

    def _rect_in_overlay(self, global_rect):
        if global_rect is None or global_rect.isNull():
            return QRect()
        top_left = self.mapFromGlobal(global_rect.topLeft())
        return QRect(top_left, global_rect.size())

    def _widget_highlight_rect(self, widget):
        if widget is None or not widget.isVisible():
            return QRect()
        global_rect = QRect(widget.mapToGlobal(QPoint(0, 0)), widget.size())
        return self._rect_in_overlay(global_rect)

    def _resolve_highlight_rect(self, step):
        if step.get("rect_getter"):
            rect = step["rect_getter"](self._main_window, self)
            if rect is not None and not rect.isNull():
                return rect
        widget = step.get("widget_getter")
        target = widget(self._main_window) if callable(widget) else widget
        if target is None:
            return QRect(self.width() // 4, self.height() // 4, self.width() // 2, self.height() // 2)
        return self._widget_highlight_rect(target)

    def _apply_step_geometry(self):
        if self._step_index < 0 or self._step_index >= len(self._steps):
            return

        step = self._steps[self._step_index]
        self._fit_to_parent()

        padding = step.get("padding", 2)
        rect = self._resolve_highlight_rect(step)
        if rect.isNull():
            rect = QRect(self.width() // 4, self.height() // 4, self.width() // 2, self.height() // 2)
        self._highlight_rect = rect.adjusted(-padding, -padding, padding, padding)

        self._card.adjustSize()
        self._position_card(step.get("placement", "auto"))
        self.update()

    def _show_step(self, index):
        if index < 0 or index >= len(self._steps):
            self._finish()
            return

        self._step_index = index
        step = self._steps[index]

        on_enter = step.get("on_enter")
        if on_enter:
            on_enter(self._main_window)

        self._title_label.setText(step.get("title", ""))
        self._text_label.setText(step.get("text", ""))
        self._counter_label.setText(f"Шаг {index + 1} из {len(self._steps)}")
        self._back_button.setEnabled(index > 0)

        is_last = index >= len(self._steps) - 1
        self._next_button.setText("Готово" if is_last else "Далее")

        QApplication.processEvents()
        delay = step.get("geometry_delay", 50)
        QTimer.singleShot(delay, self._apply_step_geometry)

    def _position_card(self, placement):
        margin = 14
        card_w = self._card.width()
        card_h = self._card.height()
        highlight = self._highlight_rect

        candidates = []
        if placement == "right":
            candidates = ["right", "bottom", "left", "top"]
        elif placement == "left":
            candidates = ["left", "bottom", "right", "top"]
        elif placement == "bottom":
            candidates = ["bottom", "right", "left", "top"]
        elif placement == "top":
            candidates = ["top", "right", "bottom", "left"]
        else:
            candidates = ["right", "bottom", "left", "top"]

        def try_place(side):
            if side == "right":
                x = highlight.right() + margin
                y = highlight.center().y() - card_h // 2
            elif side == "left":
                x = highlight.left() - margin - card_w
                y = highlight.center().y() - card_h // 2
            elif side == "bottom":
                x = highlight.center().x() - card_w // 2
                y = highlight.bottom() + margin
            else:
                x = highlight.center().x() - card_w // 2
                y = highlight.top() - margin - card_h

            x = max(margin, min(x, self.width() - card_w - margin))
            y = max(margin, min(y, self.height() - card_h - margin))
            card_rect = QRect(x, y, card_w, card_h)
            if not card_rect.intersects(highlight.adjusted(-8, -8, 8, 8)):
                return card_rect
            return None

        placed = None
        for side in candidates:
            placed = try_place(side)
            if placed:
                break
        if not placed:
            placed = QRect(
                max(margin, (self.width() - card_w) // 2),
                max(margin, self.height() - card_h - margin - 8),
                card_w,
                card_h,
            )
        self._card.setGeometry(placed)

    def paintEvent(self, event):
        painter = QPainter(self)
        painter.setRenderHint(QPainter.Antialiasing)

        if self._highlight_rect.isNull():
            painter.fillRect(self.rect(), QColor(0, 0, 0, 165))
            return

        path = QPainterPath()
        path.setFillRule(Qt.OddEvenFill)
        path.addRect(QRectF(self.rect()))
        path.addRoundedRect(QRectF(self._highlight_rect), 8, 8)
        painter.fillPath(path, QColor(0, 0, 0, 165))

        pen = QPen(QColor("#3b82f6"), 2)
        painter.setPen(pen)
        painter.setBrush(Qt.NoBrush)
        painter.drawRoundedRect(self._highlight_rect, 8, 8)

    def mousePressEvent(self, event):
        if self._card.geometry().contains(event.position().toPoint()):
            super().mousePressEvent(event)
            return
        self._go_next()

    def keyPressEvent(self, event):
        if event.key() in (Qt.Key_Return, Qt.Key_Enter, Qt.Key_Space):
            self._go_next()
            return
        if event.key() == Qt.Key_Escape:
            self._finish()
            return
        if event.key() == Qt.Key_Backspace and self._step_index > 0:
            self._go_back()
            return
        super().keyPressEvent(event)

    def _go_next(self):
        if self._step_index >= len(self._steps) - 1:
            self._finish()
        else:
            self._show_step(self._step_index + 1)

    def _go_back(self):
        if self._step_index > 0:
            self._show_step(self._step_index - 1)

    def _finish(self):
        self.hide()
        callback = self._on_completed
        self._on_completed = None
        self.deleteLater()
        if callback:
            callback()


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle(f"База знаний по задачам (v{APP_VERSION})")
        self.resize(1500, 920)
        self.setMinimumSize(1250, 760)

        self.db = KnowledgeBaseDB(DB_PATH)

        self.current_category = None
        self.username = getpass.getuser()
        self.is_admin = self.username.casefold() == ADMIN_USERNAME.casefold()
        self._onboarding_overlay = None

        self.current_category = None
        self.current_task = None
        self.current_checklist_task = None
        self.current_instruction = None
        self.current_section_titles = []
        self._section_cards = []
        self.task_state_cache = {}
        self.checklist_collapsed_sections = {}
        self.selected_checklist_task_id = None
        self.selected_checklist_step_index = None
        self._rendered_checklist_task_id = None
        self._rendered_instruction_id = None
        self._nav_panel_user_hidden = False
        self._checklist_panel_user_hidden = False
        self._checklist_content_available = False
        self._saved_nav_width = None
        self._saved_checklist_width = None
        self._main_splitter_initialized = False
        self.last_db_change_id = None
        self.last_db_file_signature = None

        self.search_timer = QTimer(self)
        self.search_timer.setSingleShot(True)
        self.search_timer.setInterval(250)
        self.search_timer.timeout.connect(self.reload_nav_tree)

        self.refresh_timer = QTimer(self)
        self.refresh_timer.setInterval(AUTO_REFRESH_INTERVAL_MS)
        self.refresh_timer.timeout.connect(self.refresh_from_database)

        self.update_check_timer = QTimer(self)
        self.update_check_timer.setInterval(UPDATE_CHECK_INTERVAL_MS)
        self.update_check_timer.timeout.connect(self.check_for_updates)

        self.build_ui()
        self.apply_admin_mode()
        self.clear_views()
        self.reload_nav_tree()
        self.remember_db_state()
        self.refresh_timer.start()
        self.check_for_updates()
        self.update_check_timer.start()
        QTimer.singleShot(450, self._maybe_start_onboarding_on_first_run)

    def _maybe_start_onboarding_on_first_run(self):
        if not self.db.is_onboarding_completed(self.username):
            self.start_onboarding(mark_completed=True)

    def build_ui(self):
        self.setStyleSheet(APP_QSS)

        root = QWidget()
        root_layout = QVBoxLayout(root)
        root_layout.setContentsMargins(12, 12, 12, 12)
        root_layout.setSpacing(10)

        self.help_menu = QMenu("Справка", self)
        self.onboarding_menu_action = self.help_menu.addAction("Обучение")
        self.onboarding_menu_action.setToolTip("Интерактивный тур по интерфейсу программы")
        self.onboarding_menu_action.triggered.connect(lambda: self.start_onboarding(mark_completed=False))
        self.help_menu_bar_action = self.menuBar().addMenu(self.help_menu)

        self.update_banner = QFrame()
        self.update_banner.setObjectName("updateBanner")
        self.update_banner.setVisible(False)
        banner_layout = QHBoxLayout(self.update_banner)
        banner_layout.setContentsMargins(12, 8, 12, 8)
        self.update_banner_label = QLabel()
        self.update_banner_label.setWordWrap(True)
        banner_layout.addWidget(self.update_banner_label, 1)
        self.update_banner.setStyleSheet("""
            QFrame#updateBanner {
                background-color: #fff3cd;
                border: 1px solid #ffeeba;
                border-radius: 8px;
            }
        """)
        self.update_banner_label.setStyleSheet("color: #856404; background: transparent; border: none;")
        root_layout.addWidget(self.update_banner)

        # Левая панель
        self.nav_panel = QWidget()
        self.nav_panel.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        left_layout = QVBoxLayout(self.nav_panel)
        left_layout.setContentsMargins(0, 0, 0, 0)
        left_layout.setSpacing(8)

        self.search_edit = QLineEdit()
        self.search_edit.setPlaceholderText("Поиск по задачам, инструкциям и тексту блоков...")
        self.search_edit.textChanged.connect(self.search_timer.start)

        search_row = QHBoxLayout()
        search_row.setSpacing(6)
        search_row.addWidget(self.search_edit, 1)
        self.toggle_nav_panel_btn = QToolButton()
        self.toggle_nav_panel_btn.setText("◀")
        self.toggle_nav_panel_btn.setFixedSize(28, 28)
        self.toggle_nav_panel_btn.setToolTip("Скрыть дерево задач")
        self.toggle_nav_panel_btn.clicked.connect(self._collapse_nav_panel)
        search_row.addWidget(self.toggle_nav_panel_btn)

        self.tasks_group = QGroupBox("Задачи")
        tasks_layout = QVBoxLayout(self.tasks_group)
        tasks_layout.setContentsMargins(8, 8, 8, 8)
        tasks_layout.setSpacing(8)

        self.nav_tree = QTreeWidget()
        self.nav_tree.setHeaderHidden(True)
        self.nav_tree.setSelectionMode(QAbstractItemView.SingleSelection)
        self.nav_tree.setAlternatingRowColors(True)
        self.nav_tree.currentItemChanged.connect(self.on_nav_item_changed)
        self.nav_tree_delegate = HtmlTreeDelegate()
        self.nav_tree.setItemDelegate(self.nav_tree_delegate)
        self.nav_tree.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)

        tasks_layout.addLayout(search_row)
        tasks_layout.addWidget(self.nav_tree, 1)

        left_layout.addWidget(self.tasks_group, 1)

        self.tree_admin_group = QGroupBox("Управление деревом")
        self.tree_admin_group.setVisible(self.is_admin)
        tree_admin_layout = QVBoxLayout(self.tree_admin_group)
        tree_admin_layout.setSpacing(6)

        row1 = QHBoxLayout()

        self.add_category_button = QPushButton("Раздел")
        self.add_category_button.setToolTip("Добавить новый раздел верхнего уровня")
        self.add_category_button.clicked.connect(self.add_tree_category)
        row1.addWidget(self.add_category_button)

        self.add_subcategory_button = QPushButton("Задача")
        self.add_subcategory_button.setToolTip("Добавить задачу внутри выбранного раздела")
        self.add_subcategory_button.clicked.connect(self.add_tree_subcategory)
        row1.addWidget(self.add_subcategory_button)

        self.add_task_button = QPushButton("Подзадача")
        self.add_task_button.setToolTip("Добавить подзадачу в выбранную задачу")
        self.add_task_button.clicked.connect(self.add_tree_task)
        row1.addWidget(self.add_task_button)

        self.add_instruction_button = QPushButton("Добавить инструкцию")
        self.add_instruction_button.setToolTip("Создать инструкцию для выбранной задачи")
        self.add_instruction_button.clicked.connect(self.add_tree_instruction)
        row1.addWidget(self.add_instruction_button)

        tree_admin_layout.addLayout(row1)

        row2 = QHBoxLayout()
        self.edit_tree_button = QPushButton("Редактировать")
        self.edit_tree_button.clicked.connect(self.edit_tree_item)
        row2.addWidget(self.edit_tree_button)

        self.delete_tree_button = QPushButton("Удалить")
        self.delete_tree_button.clicked.connect(self.delete_tree_item)
        row2.addWidget(self.delete_tree_button)

        self.save_tree_button = QPushButton("Сохранить")
        self.save_tree_button.clicked.connect(self.save_tree_changes)
        row2.addWidget(self.save_tree_button)

        tree_admin_layout.addLayout(row2)
        row3 = QHBoxLayout()
        self.move_up_button = QPushButton("⬆ Вверх")
        self.move_up_button.setToolTip("Переместить выбранный элемент вверх")
        self.move_up_button.clicked.connect(self.move_tree_item_up)
        row3.addWidget(self.move_up_button)

        self.move_down_button = QPushButton("⬇ Вниз")
        self.move_down_button.setToolTip("Переместить выбранный элемент вниз")
        self.move_down_button.clicked.connect(self.move_tree_item_down)
        row3.addWidget(self.move_down_button)

        tree_admin_layout.addLayout(row3)
        left_layout.addWidget(self.tree_admin_group)

        # Основной контент: дерево 1/5, инструкция 3/5, чек-лист 1/5
        self.tabs = QTabWidget()
        self.tabs.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)

        self.build_task_tab()
        self.build_checklist_sidebar()
        self.build_instruction_tab()
        self.build_feedback_tab()

        self.tabs.addTab(self.instruction_tab, "📘 Инструкция")
        self.tabs.addTab(self.feedback_tab, "💬 Комментарии и оценка")

        self.nav_panel_expander = QToolButton()
        self.nav_panel_expander.setText("▶")
        self.nav_panel_expander.setFixedWidth(28)
        self.nav_panel_expander.setToolTip("Показать дерево задач")
        self.nav_panel_expander.setVisible(False)
        self.nav_panel_expander.clicked.connect(self._expand_nav_panel)
        self.nav_panel_expander.setStyleSheet("""
            QToolButton {
                background: #eef5fb;
                border: 1px solid #d9e3f0;
                border-radius: 6px;
                font-size: 14px;
            }
            QToolButton:hover { background: #dce9f8; }
        """)

        self.checklist_panel_expander = QToolButton()
        self.checklist_panel_expander.setText("◀")
        self.checklist_panel_expander.setFixedWidth(28)
        self.checklist_panel_expander.setToolTip("Показать чек-лист")
        self.checklist_panel_expander.setVisible(False)
        self.checklist_panel_expander.clicked.connect(self._expand_checklist_panel)
        self.checklist_panel_expander.setStyleSheet(self.nav_panel_expander.styleSheet())

        self.nav_panel.setMinimumWidth(180)
        self.checklist_sidebar.setMinimumWidth(180)

        self.main_splitter = QSplitter(Qt.Horizontal)
        self.main_splitter.setChildrenCollapsible(False)
        self.main_splitter.addWidget(self.nav_panel)
        self.main_splitter.addWidget(self.tabs)
        self.main_splitter.addWidget(self.checklist_sidebar)
        self.main_splitter.setStretchFactor(0, 1)
        self.main_splitter.setStretchFactor(1, 3)
        self.main_splitter.setStretchFactor(2, 1)
        self.main_splitter.splitterMoved.connect(self._on_main_splitter_moved)

        self.main_columns = QWidget()
        self.main_columns_layout = QHBoxLayout(self.main_columns)
        self.main_columns_layout.setContentsMargins(0, 0, 0, 0)
        self.main_columns_layout.setSpacing(8)
        self.main_columns_layout.addWidget(self.nav_panel_expander, 0)
        self.main_columns_layout.addWidget(self.main_splitter, 1)
        self.main_columns_layout.addWidget(self.checklist_panel_expander, 0)

        root_layout.addWidget(self.main_columns, 1)
        QTimer.singleShot(0, self._initialize_main_splitter_sizes)
        self.setCentralWidget(root)

        self.statusBar().showMessage("Выберите задачу слева")

    def check_for_updates(self):
        config = load_deployed_version_config()
        deployed_version = str(config.get("current", "")).strip()
        notes = str(config.get("notes", "")).strip()

        if not deployed_version or not is_version_newer(deployed_version, APP_VERSION):
            if hasattr(self, "update_banner"):
                self.update_banner.setVisible(False)
            return

        message = (
            f"Доступна новая версия {deployed_version} "
            f"(сейчас установлена {APP_VERSION}). "
            "Перезапустите приложение после обновления файлов в сетевой папке."
        )
        if notes:
            message += f" {notes}"
        self.update_banner_label.setText(message)
        self.update_banner.setVisible(True)

    def remember_db_state(self):
        try:
            self.last_db_change_id = self.db.get_change_signature()
            self.last_db_file_signature = self.db.get_file_signature()
        except sqlite3.OperationalError:
            self.last_db_change_id = None
            self.last_db_file_signature = None

    def refresh_from_database(self):
        if not self.isVisible():
            return

        try:
            file_signature = self.db.get_file_signature()
        except Exception:
            return

        if (
            self.last_db_file_signature is not None
            and file_signature is not None
            and file_signature != self.last_db_file_signature
        ):
            try:
                self.db.reconnect()
            except sqlite3.OperationalError:
                return
            self.remember_db_state()
            self.reload_all_views()
            self.statusBar().showMessage("База данных обновлена с сервера", 5000)
            return

        try:
            latest_change_id = self.db.get_change_signature()
        except sqlite3.OperationalError:
            return

        if latest_change_id == self.last_db_change_id:
            return

        self.last_db_change_id = latest_change_id
        self.last_db_file_signature = file_signature
        self.reload_all_views()
        self.statusBar().showMessage("Данные обновлены", 3000)

    def reload_all_views(self):
        task_id = self.current_task.get("task_id") if self.current_task else None
        category_id = self.current_category.get("id") if self.current_category else None

        self.reload_nav_tree()

        if task_id and self.db.task_bundle(task_id):
            self.show_task(task_id, force=True)
        elif category_id and self.db.category_by_id(category_id):
            self.show_category(category_id)
        else:
            self.clear_views()

    def _initialize_main_splitter_sizes(self):
        if not hasattr(self, "main_splitter"):
            return
        if self._main_splitter_initialized and self.main_splitter.width() > 0:
            return
        total = max(self.main_splitter.width(), 1000)
        nav_w = total // 5
        check_w = total // 5
        center_w = max(300, total - nav_w - check_w)
        self.main_splitter.setSizes([nav_w, center_w, check_w])
        self._saved_nav_width = nav_w
        self._saved_checklist_width = check_w
        self._main_splitter_initialized = True
        self._fit_instruction_content_width()

    def _on_main_splitter_moved(self, pos, index):
        sizes = self.main_splitter.sizes()
        if self.nav_panel.isVisible() and sizes[0] > 0:
            self._saved_nav_width = sizes[0]
        if self.checklist_sidebar.isVisible() and sizes[2] > 0:
            self._saved_checklist_width = sizes[2]
        QTimer.singleShot(0, self._fit_instruction_content_width)

    def _sync_main_splitter_layout(self):
        """Подстраивает ширины колонок при скрытии/показе боковых панелей."""
        if not hasattr(self, "main_splitter"):
            return

        total = self.main_splitter.width()
        if total <= 0:
            QTimer.singleShot(0, self._sync_main_splitter_layout)
            return

        sizes = list(self.main_splitter.sizes())
        nav_w, center_w, check_w = sizes[0], sizes[1], sizes[2]
        nav_open = self.nav_panel.isVisible()
        checklist_open = self.checklist_sidebar.isVisible()

        if nav_open and not checklist_open:
            check_w = 0
        if not nav_open and checklist_open:
            nav_w = 0
        if not nav_open and not checklist_open:
            nav_w = 0
            check_w = 0

        if nav_open:
            if nav_w <= 0:
                nav_w = self._saved_nav_width or max(180, total // 5)
            nav_w = max(180, min(nav_w, total - 360))
        else:
            if sizes[0] > 0:
                self._saved_nav_width = sizes[0]
            center_w += nav_w
            nav_w = 0

        if checklist_open:
            if check_w <= 0:
                check_w = self._saved_checklist_width or max(180, total // 5)
            check_w = max(180, min(check_w, total - nav_w - 300))
        else:
            if sizes[2] > 0:
                self._saved_checklist_width = sizes[2]
            center_w += check_w
            check_w = 0

        center_w = max(300, total - nav_w - check_w)
        self.main_splitter.setSizes([nav_w, center_w, check_w])
        QTimer.singleShot(0, self._fit_instruction_content_width)

    def _collapse_nav_panel(self):
        self._nav_panel_user_hidden = True
        self._update_nav_panel_visibility()

    def _expand_nav_panel(self):
        self._nav_panel_user_hidden = False
        self._update_nav_panel_visibility()

    def _update_nav_panel_visibility(self):
        if hasattr(self, "nav_panel"):
            self.nav_panel.setVisible(not self._nav_panel_user_hidden)
        if hasattr(self, "nav_panel_expander"):
            self.nav_panel_expander.setVisible(self._nav_panel_user_hidden)
        self._sync_main_splitter_layout()

    def _collapse_checklist_panel(self):
        if not self._checklist_content_available:
            return
        self._checklist_panel_user_hidden = True
        self._update_checklist_panel_visibility()

    def _expand_checklist_panel(self):
        self._checklist_panel_user_hidden = False
        self._update_checklist_panel_visibility()

    def _update_checklist_panel_visibility(self):
        show = self._checklist_content_available and not self._checklist_panel_user_hidden
        if hasattr(self, "checklist_sidebar"):
            self.checklist_sidebar.setVisible(show)
        if hasattr(self, "checklist_panel_expander"):
            self.checklist_panel_expander.setVisible(
                self._checklist_content_available and self._checklist_panel_user_hidden
            )
        if hasattr(self, "toggle_checklist_panel_btn"):
            self.toggle_checklist_panel_btn.setVisible(self._checklist_content_available)
            self.toggle_checklist_panel_btn.setText("▶" if self._checklist_panel_user_hidden else "◀")
            self.toggle_checklist_panel_btn.setToolTip(
                "Показать чек-лист" if self._checklist_panel_user_hidden else "Скрыть чек-лист"
            )
        self._sync_main_splitter_layout()

    def _instruction_content_width(self):
        if not hasattr(self, "instruction_scroll"):
            return 800
        return max(200, self.instruction_scroll.viewport().width())

    def _fit_instruction_content_width(self):
        width = self._instruction_content_width()
        if hasattr(self, "instruction_container"):
            self.instruction_container.setMaximumWidth(width)
        if hasattr(self, "instruction_desc_group"):
            self.instruction_desc_group.setMaximumWidth(width)
        for card in getattr(self, "_section_cards", []):
            if hasattr(card, "update_images_width"):
                card.update_images_width(width)

    def eventFilter(self, obj, event):
        if (
                hasattr(self, "instruction_scroll")
                and isValid(self.instruction_scroll)
                and obj is self.instruction_scroll.viewport()
                and event.type() == QEvent.Resize
        ):
            QTimer.singleShot(0, self._fit_instruction_content_width)
        return super().eventFilter(obj, event)

    def build_task_tab(self):
        self.task_tab = QWidget()
        layout = QVBoxLayout(self.task_tab)
        layout.setContentsMargins(10, 10, 10, 10)
        layout.setSpacing(10)

        header_row = QHBoxLayout()

        self.task_title_label = QLabel("Выберите задачу слева")
        self.task_title_label.setWordWrap(True)
        self.task_title_label.setFont(QFont("Segoe UI", 14, QFont.Bold))
        self.task_title_label.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)

        header_row.addWidget(self.task_title_label, 1)

        self.edit_task_button = QPushButton("Редактировать подзадачу")
        self.edit_task_button.setToolTip("Редактировать описание, инструкцию и чек-лист")
        self.edit_task_button.setVisible(self.is_admin)
        self.edit_task_button.clicked.connect(self.open_task_editor)
        header_row.addWidget(self.edit_task_button)

        layout.addLayout(header_row)

        self.task_category_label = QLabel("")
        self.task_category_label.setStyleSheet("color: #5b6577;")
        layout.addWidget(self.task_category_label)

        self.task_desc_group = QGroupBox("Описание")
        self.task_desc_group.setVisible(False)
        self.task_desc_group.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)
        desc_layout_inner = QVBoxLayout(self.task_desc_group)

        self.task_desc_label = QLabel("")
        self.task_desc_label.setWordWrap(True)
        self.task_desc_label.setTextFormat(Qt.RichText)
        self.task_desc_label.setOpenExternalLinks(True)
        self.task_desc_label.linkActivated.connect(self.handle_link_activated)
        self.task_desc_label.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)
        self.task_desc_label.setStyleSheet("color: #263238; background: transparent; border: none;")
        desc_layout_inner.addWidget(self.task_desc_label)

        layout.addWidget(self.task_desc_group)

        self.category_tasks_group = QGroupBox("Задачи")
        self.category_tasks_group.setVisible(False)
        self.category_tasks_layout = QVBoxLayout(self.category_tasks_group)
        self.category_tasks_layout.setContentsMargins(8, 8, 8, 8)
        self.category_tasks_layout.setSpacing(6)
        layout.addWidget(self.category_tasks_group)

        self.task_instruction_label = QLabel("")
        self.task_instruction_label.setWordWrap(True)
        self.task_instruction_label.setStyleSheet("font-weight: 600; color: #1f2937;")
        layout.addWidget(self.task_instruction_label)

        layout.addStretch()

    def build_checklist_sidebar(self):
        self.checklist_sidebar = QWidget()
        self.checklist_sidebar.setObjectName("checklistSidebar")
        self.checklist_sidebar.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self.checklist_sidebar.setStyleSheet("""
            QWidget#checklistSidebar {
                background: #ffffff;
                border-left: 1px solid #d9e3f0;
            }
        """)

        layout = QVBoxLayout(self.checklist_sidebar)
        layout.setContentsMargins(10, 10, 10, 10)
        layout.setSpacing(8)

        checklist_header = QHBoxLayout()
        checklist_title = QLabel("Чек-лист")
        checklist_title.setFont(QFont("Segoe UI", 11, QFont.Bold))
        checklist_header.addWidget(checklist_title, 1)
        self.toggle_checklist_panel_btn = QToolButton()
        self.toggle_checklist_panel_btn.setText("◀")
        self.toggle_checklist_panel_btn.setFixedSize(28, 28)
        self.toggle_checklist_panel_btn.setToolTip("Скрыть чек-лист")
        self.toggle_checklist_panel_btn.clicked.connect(self._collapse_checklist_panel)
        checklist_header.addWidget(self.toggle_checklist_panel_btn)
        layout.addLayout(checklist_header)

        self.task_hint_label = QLabel(
            "Секции — это подзадачи. Нажми на пункт чек-листа, "
            "и откроется нужный фрагмент инструкции."
        )
        self.task_hint_label.setWordWrap(True)
        self.task_hint_label.setStyleSheet("color: #5b6577; font-size: 11px;")
        layout.addWidget(self.task_hint_label)

        checklist_group = QGroupBox("Пункты")
        self.task_checklist_group = checklist_group
        self.task_checklist_group.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        checklist_layout = QVBoxLayout(checklist_group)
        checklist_layout.setContentsMargins(8, 8, 8, 8)
        checklist_layout.setSpacing(6)

        self.checklist_sticky_header = QToolButton()
        self.checklist_sticky_header.setVisible(False)
        self.checklist_sticky_header.setCheckable(True)
        self.checklist_sticky_header.setToolButtonStyle(Qt.ToolButtonTextBesideIcon)
        self.checklist_sticky_header.setArrowType(Qt.UpArrow)
        self.checklist_sticky_header.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        self.checklist_sticky_header.clicked.connect(self._toggle_all_checklist_sections)
        self.checklist_sticky_header.setStyleSheet("""
            QToolButton {
                font-size: 13px;
                font-weight: bold;
                background: #eef5fb;
                color: #263238;
                border: 1px solid #d9e3f0;
                border-radius: 8px;
                padding: 8px 10px;
                text-align: left;
            }
            QToolButton:hover {
                background: #e3eefb;
            }
        """)
        self.checklist_sticky_header.setText("Свернуть чек-лист")
        checklist_layout.addWidget(self.checklist_sticky_header)

        self.task_checklist = QListWidget()
        self.task_checklist.setItemDelegate(WrappedChecklistDelegate(self.task_checklist))
        self.task_checklist.itemChanged.connect(self.on_checklist_item_changed)
        self.task_checklist.itemClicked.connect(self.on_checklist_item_clicked)
        self.task_checklist.setMouseTracking(True)
        self.task_checklist.viewport().setMouseTracking(True)
        self.task_checklist.setSelectionMode(QAbstractItemView.SingleSelection)
        self.task_checklist.setAlternatingRowColors(True)
        self.task_checklist.setCursor(Qt.PointingHandCursor)
        self.task_checklist.setWordWrap(True)
        self.task_checklist.setTextElideMode(Qt.ElideNone)
        self.task_checklist.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.task_checklist.setUniformItemSizes(False)
        self.task_checklist.setSpacing(4)
        self.task_checklist.setToolTip("Нажми на пункт чек-листа, чтобы открыть нужный шаг в инструкции.")
        self.task_checklist.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self.task_checklist.setMinimumHeight(80)

        checklist_layout.addWidget(self.task_checklist)
        layout.addWidget(checklist_group, 1)

        button_row = QVBoxLayout()
        button_row.setSpacing(6)
        button_row.setAlignment(Qt.AlignTop | Qt.AlignLeft)

        self.prev_step_button = QPushButton("Предыдущий шаг")
        self.prev_step_button.setToolTip("Открыть предыдущий пункт чек-листа")
        self.prev_step_button.clicked.connect(self.go_to_previous_step)
        self.prev_step_button.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Fixed)

        self.next_step_button = QPushButton("Следующий шаг")
        self.next_step_button.setToolTip("Открыть следующий незавершённый шаг инструкции")
        self.next_step_button.clicked.connect(self.go_to_next_step)
        self.next_step_button.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Fixed)

        self.reset_checklist_button = QPushButton("Очистить отметки")
        self.reset_checklist_button.setToolTip("Снять галочки со всех пунктов текущего чек-листа")
        self.reset_checklist_button.clicked.connect(self.reset_checklist_states)
        self.reset_checklist_button.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Fixed)

        self.edit_checklist_button = QPushButton("Редактировать чек-лист")
        self.edit_checklist_button.setToolTip("Открыть редактор и изменить пункты чек-листа")
        self.edit_checklist_button.clicked.connect(self.open_checklist_editor)
        self.edit_checklist_button.setVisible(self.is_admin)
        self.edit_checklist_button.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Fixed)

        step_button_row = QHBoxLayout()
        step_button_row.setSpacing(6)
        step_button_row.setAlignment(Qt.AlignLeft)
        step_button_row.addWidget(self.prev_step_button)
        step_button_row.addWidget(self.next_step_button)
        button_row.addLayout(step_button_row)

        reset_button_row = QHBoxLayout()
        reset_button_row.setSpacing(6)
        reset_button_row.setAlignment(Qt.AlignLeft)
        reset_button_row.addWidget(self.reset_checklist_button)
        button_row.addLayout(reset_button_row)
        button_row.addWidget(self.edit_checklist_button, 0, Qt.AlignLeft)

        layout.addLayout(button_row)
        layout.addStretch()

    def build_instruction_tab(self):
        self.instruction_tab = QWidget()
        layout = QVBoxLayout(self.instruction_tab)
        layout.setContentsMargins(10, 10, 10, 10)
        layout.setSpacing(10)

        # Заголовок с кнопкой редактирования
        header_row = QHBoxLayout()
        self.instruction_title_label = QLabel("Инструкция")
        self.instruction_title_label.setWordWrap(True)
        self.instruction_title_label.setFont(QFont("Segoe UI", 14, QFont.Bold))
        self.instruction_title_label.setSizePolicy(QSizePolicy.Ignored, QSizePolicy.Preferred)

        header_row.addWidget(self.instruction_title_label, 1)
        header_row.addStretch()

        self.edit_instruction_btn = QPushButton("Редактировать инструкцию")
        self.edit_instruction_btn.setToolTip("Изменить содержание инструкции")
        self.edit_instruction_btn.setVisible(False)  # будет показана, когда загружена инструкция
        self.edit_instruction_btn.clicked.connect(self.edit_current_instruction)
        header_row.addWidget(self.edit_instruction_btn)

        layout.addLayout(header_row)

        self.instruction_desc_group = QGroupBox("Описание")
        self.instruction_desc_group.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)
        instruction_desc_layout = QVBoxLayout(self.instruction_desc_group)

        self.instruction_desc_label = QLabel("")
        self.instruction_desc_label.setWordWrap(True)
        self.instruction_desc_label.setTextFormat(Qt.RichText)
        self.instruction_desc_label.setOpenExternalLinks(True)
        self.instruction_desc_label.linkActivated.connect(self.handle_link_activated)
        self.instruction_desc_label.setSizePolicy(QSizePolicy.Ignored, QSizePolicy.Preferred)
        self.instruction_desc_label.setStyleSheet("color: #263238; background: transparent; border: none;")
        instruction_desc_layout.addWidget(self.instruction_desc_label)

        layout.addWidget(self.instruction_desc_group)
        self.instruction_nav_group = QGroupBox("Переходы")
        self.instruction_nav_group.setVisible(False)
        self.instruction_nav_layout = QVBoxLayout(self.instruction_nav_group)
        self.instruction_nav_layout.setContentsMargins(8, 8, 8, 8)
        self.instruction_nav_layout.setSpacing(6)
        layout.addWidget(self.instruction_nav_group)

        self.instruction_scroll = QScrollArea()
        self.instruction_scroll.setWidgetResizable(True)
        self.instruction_scroll.setFrameShape(QFrame.NoFrame)
        self.instruction_scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.instruction_scroll.viewport().installEventFilter(self)

        # --- Sticky header (приклеенный заголовок) ---
        self.sticky_header_section_index = None
        self.sticky_header_label = QToolButton()
        self.sticky_header_label.setVisible(False)
        self.sticky_header_label.setCheckable(True)
        self.sticky_header_label.setToolButtonStyle(Qt.ToolButtonTextBesideIcon)
        self.sticky_header_label.setArrowType(Qt.RightArrow)
        self.sticky_header_label.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        self.sticky_header_label.clicked.connect(self._toggle_sticky_header_section)
        self.sticky_header_label.setStyleSheet("""
            QToolButton {
                font-size: 15px;
                font-weight: bold;
                background: #f7f9fd;
                color: #263238;
                border: 1px solid #d9e3f0;
                border-radius: 10px;
                padding: 10px 12px;
                text-align: left;
            }
            QToolButton:hover {
                background: #eef5fb;
            }
        """)
        layout.addWidget(self.sticky_header_label)

        self.instruction_container = QWidget()
        self.instruction_container_layout = QVBoxLayout(self.instruction_container)
        self.instruction_container_layout.setContentsMargins(0, 0, 0, 0)
        self.instruction_container_layout.setSpacing(10)
        self.instruction_container_layout.setAlignment(Qt.AlignTop)

        self.instruction_scroll.setWidget(self.instruction_container)
        layout.addWidget(self.instruction_scroll)

        # Подключаемся к сигналу прокрутки
        self.instruction_scroll.verticalScrollBar().valueChanged.connect(self._update_sticky_header)

    def build_feedback_tab(self):
        self.feedback_tab = QWidget()
        layout = QVBoxLayout(self.feedback_tab)
        layout.setSpacing(10)

        # --- Заголовок ---
        title = QLabel("Комментарии и оценка к инструкции")
        title.setFont(QFont("Segoe UI", 14, QFont.Bold))
        layout.addWidget(title)

        note = QLabel("Оценка анонимная. Старайся быть честным — так инструкция станет лучше.")
        note.setStyleSheet("color: #5b6577;")
        note.setWordWrap(True)
        layout.addWidget(note)

        # --- Статистика ---
        self.feedback_stats_label = QLabel("Выберите инструкцию слева")
        self.feedback_stats_label.setFont(QFont("Segoe UI", 12, QFont.Bold))
        layout.addWidget(self.feedback_stats_label)

        # --- Оценка ---
        rating_layout = QHBoxLayout()
        rating_layout.addWidget(QLabel("Оценка (1–10):"))
        self.rating_spin = QSpinBox()
        self.rating_spin.setRange(1, 10)
        self.rating_spin.setValue(10)
        self.rating_spin.setMaximumWidth(90)
        rating_layout.addWidget(self.rating_spin)
        rating_layout.addStretch()
        layout.addLayout(rating_layout)

        hint_label = QLabel("Где 1 — очень плохо, а 10 — отличная и самодостаточная инструкция")
        hint_label.setWordWrap(True)
        hint_label.setStyleSheet("font-size: 11px; color: #5b6577; background: transparent; border: none;")
        layout.addWidget(hint_label)

        # --- Комментарий ---
        comment_title = QLabel("Комментарий к инструкции")
        comment_title.setFont(QFont("Segoe UI", 13, QFont.Bold))
        layout.addWidget(comment_title)

        top_row = QHBoxLayout()
        top_row.addWidget(QLabel("Блок:"))
        self.comment_anchor_box = QComboBox()
        self.comment_anchor_box.setMinimumWidth(280)
        top_row.addWidget(self.comment_anchor_box)
        self.comment_author_edit = QLineEdit()
        self.comment_author_edit.setPlaceholderText("Имя (необязательно)")
        top_row.addWidget(self.comment_author_edit)
        layout.addLayout(top_row)

        comment_note = QLabel("Можно оставить комментарий к целой инструкции или к конкретному блоку.")
        comment_note.setStyleSheet("color: #5b6577;")
        comment_note.setWordWrap(True)
        layout.addWidget(comment_note)

        self.comment_text_edit = QTextEdit()
        self.comment_text_edit.setPlaceholderText("Что непонятно? Что нужно поправить или доработать?")
        self.comment_text_edit.setMinimumHeight(100)
        layout.addWidget(self.comment_text_edit)

        # --- Кнопки ---
        btn_row = QHBoxLayout()

        self.submit_feedback_btn = QPushButton("Поставить оценку и добавить комментарий")
        self.submit_feedback_btn.clicked.connect(self.submit_feedback)
        btn_row.addWidget(self.submit_feedback_btn)

        self.manage_feedback_btn = QPushButton("Редактировать")
        self.manage_feedback_btn.setToolTip("Редактировать и удалять оценки и комментарии")
        self.manage_feedback_btn.setVisible(self.is_admin)
        self.manage_feedback_btn.clicked.connect(self.open_feedback_manager)
        btn_row.addWidget(self.manage_feedback_btn)

        btn_row.addStretch()
        layout.addLayout(btn_row)

        # --- Экспорт (только админ) ---
        self.export_feedback_button = QPushButton("Экспорт в Excel")
        self.export_feedback_button.clicked.connect(self.export_feedback_to_excel)
        self.export_feedback_button.setVisible(self.is_admin)
        layout.addWidget(self.export_feedback_button, alignment=Qt.AlignRight)

        # --- Список комментариев (видимый всем) ---
        self.feedback_comments_browser = QTextBrowser()
        self.feedback_comments_browser.setOpenExternalLinks(False)
        self.feedback_comments_browser.setStyleSheet("""
            QTextBrowser {
                background: white;
                border: 1px solid #d8dee9;
                border-radius: 8px;
            }
        """)
        layout.addWidget(self.feedback_comments_browser)

        layout.addStretch()

    def build_comments_tab(self):
        self.comments_tab = QWidget()
        layout = QVBoxLayout(self.comments_tab)
        layout.setSpacing(10)

        title = QLabel("Комментарии к инструкции")
        title.setFont(QFont("Segoe UI", 13, QFont.Bold))

        note = QLabel("Можно оставить комментарий к целой инструкции или к конкретному блоку.")
        note.setStyleSheet("color: #5b6577;")
        note.setWordWrap(True)

        layout.addWidget(title)
        layout.addWidget(note)

        self.comments_browser = QTextBrowser()
        self.comments_browser.setOpenExternalLinks(False)
        self.comments_browser.setStyleSheet("""
            QTextBrowser {
                background: white;
                border: 1px solid #d8dee9;
                border-radius: 8px;
            }
        """)
        layout.addWidget(self.comments_browser)

        self.comment_form_widget = QWidget()
        form_layout = QVBoxLayout(self.comment_form_widget)
        form_layout.setSpacing(8)

        top_row = QHBoxLayout()

        top_row.addWidget(QLabel("Блок:"))
        self.comment_anchor_box = QComboBox()
        self.comment_anchor_box.setMinimumWidth(280)
        top_row.addWidget(self.comment_anchor_box)

        top_row.addSpacing(14)
        self.comment_anonymous_check = QCheckBox("Анонимно")
        self.comment_anonymous_check.setChecked(True)
        self.comment_anonymous_check.toggled.connect(
            lambda checked: self.comment_author_edit.setEnabled(not checked)
        )
        top_row.addWidget(self.comment_anonymous_check)

        self.comment_author_edit = QLineEdit()
        self.comment_author_edit.setPlaceholderText("Имя (необязательно)")
        self.comment_author_edit.setEnabled(False)
        top_row.addWidget(self.comment_author_edit)

        form_layout.addLayout(top_row)

        self.comment_text_edit = QTextEdit()
        self.comment_text_edit.setPlaceholderText("Что непонятно? Что нужно поправить или доработать?")
        self.comment_text_edit.setMinimumHeight(90)
        form_layout.addWidget(self.comment_text_edit)

        btn_row = QHBoxLayout()
        btn_row.addStretch()
        self.add_comment_button = QPushButton("Добавить комментарий")
        self.add_comment_button.clicked.connect(self.add_comment)
        btn_row.addWidget(self.add_comment_button)
        form_layout.addLayout(btn_row)

        layout.addWidget(self.comment_form_widget)

    def build_rating_tab(self):
        self.rating_tab = QWidget()
        layout = QVBoxLayout(self.rating_tab)
        layout.setSpacing(10)

        title = QLabel("Анонимная оценка понятности")
        title.setFont(QFont("Segoe UI", 13, QFont.Bold))

        note = QLabel("Оценка анонимная. Старайся быть честным — так инструкция станет лучше.")
        note.setStyleSheet("color: #5b6577;")
        note.setWordWrap(True)

        layout.addWidget(title)
        layout.addWidget(note)

        self.rating_summary_label = QLabel("Выберите инструкцию слева")
        self.rating_summary_label.setWordWrap(True)
        self.rating_summary_label.setFont(QFont("Segoe UI", 12, QFont.Bold))
        layout.addWidget(self.rating_summary_label)

        self.rating_form_widget = QWidget()
        form_layout = QHBoxLayout(self.rating_form_widget)
        form_layout.setContentsMargins(0, 0, 0, 0)
        form_layout.setSpacing(10)

        form_layout.addWidget(QLabel("Оценка (1–5):"))
        self.rating_spin = QSpinBox()
        self.rating_spin.setRange(1, 5)
        self.rating_spin.setValue(5)
        self.rating_spin.setMaximumWidth(90)
        form_layout.addWidget(self.rating_spin)

        self.add_rating_button = QPushButton("Поставить оценку")
        self.add_rating_button.clicked.connect(self.add_rating)
        form_layout.addWidget(self.add_rating_button)

        form_layout.addStretch()
        layout.addWidget(self.rating_form_widget)
        layout.addStretch()

    # ================== Навигация ==================

    def reload_nav_tree(self):
        search = self.search_edit.text().strip().casefold()

        categories = [dict(row) for row in self.db.conn.execute("""
            SELECT id, name, parent_id, sort_order
            FROM categories
            ORDER BY sort_order, name, id
        """).fetchall()]

        categories_by_parent = {}
        parent_map = {}
        for cat in categories:
            categories_by_parent.setdefault(cat["parent_id"], []).append(cat)
            parent_map[cat["id"]] = cat["parent_id"]

        for items in categories_by_parent.values():
            items.sort(key=lambda c: (c["sort_order"], c["name"], c["id"]))

        tasks = self.db.search_tasks(search) if search else self.db.all_tasks()
        tasks_by_category = {}
        matched_category_ids = set()

        for task in tasks:
            tasks_by_category.setdefault(task["category_id"], []).append(task)

            cid = task["category_id"]
            while cid is not None:
                matched_category_ids.add(cid)
                cid = parent_map.get(cid)

        self.nav_tree.blockSignals(True)
        self.nav_tree.clear()

        item_map = {}

        def add_branch(parent_id, parent_item=None):
            for cat in categories_by_parent.get(parent_id, []):
                if search and cat["id"] not in matched_category_ids:
                    continue

                cat_item = QTreeWidgetItem([cat["name"]])
                cat_item.setData(0, ROLE_KIND, "category")
                cat_item.setData(0, ROLE_ID, cat["id"])

                font = cat_item.font(0)
                font.setBold(True)
                if parent_item is None:
                    font.setPointSize(10)   # Раздел
                else:
                    font.setPointSize(9)   # Подкатегория
                cat_item.setFont(0, font)

                item_map[cat["id"]] = cat_item

                if parent_item is None:
                    self.nav_tree.addTopLevelItem(cat_item)
                else:
                    parent_item.addChild(cat_item)

                # Сначала показываем подкатегории
                add_branch(cat["id"], cat_item)

                # Потом задачи этой категории
                for task in tasks_by_category.get(cat["id"], []):
                    task_item = QTreeWidgetItem([task["task_title"]])
                    task_item.setData(0, ROLE_KIND, "task")
                    task_item.setData(0, ROLE_ID, task["task_id"])

                    task_font = task_item.font(0)
                    task_font.setBold(False)
                    task_font.setPointSize(9)  # как сейчас, без выделения
                    task_item.setFont(0, task_font)

                    cat_item.addChild(task_item)
                    item_map[task["task_id"]] = task_item

                # По умолчанию все уровни свёрнуты: раздел → задача → подзадача
                # раскрываются только вручную или при навигации к выбранному элементу.
                cat_item.setExpanded(False)

        add_branch(None)

        if not item_map:
            empty = QTreeWidgetItem(["Ничего не найдено"])
            empty.setFlags(Qt.NoItemFlags)
            self.nav_tree.addTopLevelItem(empty)
            self.nav_tree.blockSignals(False)
            self.clear_views()
            self.statusBar().showMessage("Ничего не найдено")
            return

        self.nav_tree.blockSignals(False)

        target_task_id = self.current_task["task_id"] if self.current_task else None
        target_category_id = self.current_category["id"] if self.current_category else None

        if target_task_id and target_task_id in item_map:
            self._select_in_nav_tree(task_id=target_task_id)
        elif target_category_id and target_category_id in item_map:
            self._select_in_nav_tree(category_id=target_category_id)
        elif search:
            self._expand_nav_tree_for_search()
        else:
            self.nav_tree.clearSelection()
            self.nav_tree.setCurrentItem(None)

        # Подсвечиваем поисковый текст в дереве (всегда после построения)
        QTimer.singleShot(50, self._highlight_search_in_tree)

    def _expand_nav_tree_for_search(self):
        """При активном поиске раскрывает ветки, чтобы были видны найденные элементы."""
        def expand_if_has_children(item):
            for i in range(item.childCount()):
                expand_if_has_children(item.child(i))
            if item.childCount() > 0:
                item.setExpanded(True)

        for i in range(self.nav_tree.topLevelItemCount()):
            expand_if_has_children(self.nav_tree.topLevelItem(i))

    def _select_in_nav_tree(self, task_id=None, category_id=None):
        """Выделяет подзадачу или категорию в дереве без повторной загрузки контента."""
        if task_id is None and category_id is None:
            return

        def walk(item):
            kind = item.data(0, ROLE_KIND)
            entity_id = item.data(0, ROLE_ID)
            if task_id is not None and kind == "task" and entity_id == task_id:
                return item
            if category_id is not None and kind == "category" and entity_id == category_id:
                return item
            for i in range(item.childCount()):
                found = walk(item.child(i))
                if found:
                    return found
            return None

        target_item = None
        for i in range(self.nav_tree.topLevelItemCount()):
            top = self.nav_tree.topLevelItem(i)
            kind = top.data(0, ROLE_KIND)
            entity_id = top.data(0, ROLE_ID)
            if task_id is not None and kind == "task" and entity_id == task_id:
                target_item = top
                break
            if category_id is not None and kind == "category" and entity_id == category_id:
                target_item = top
                break
            found = walk(top)
            if found:
                target_item = found
                break

        if not target_item:
            return

        self.nav_tree.blockSignals(True)
        parent = target_item.parent()
        while parent:
            parent.setExpanded(True)
            parent = parent.parent()
        self.nav_tree.setCurrentItem(target_item)
        self.nav_tree.scrollToItem(target_item, QAbstractItemView.PositionAtCenter)
        self.nav_tree.blockSignals(False)

    def on_nav_item_changed(self, current, previous):
        if not current:
            self.clear_views()
            self.update_tree_admin_controls()
            return

        kind = current.data(0, ROLE_KIND)
        entity_id = current.data(0, ROLE_ID)

        if kind == "task":
            self.show_task(entity_id)
            self.tabs.setCurrentWidget(self.instruction_tab)

        elif kind == "category":
            self.show_category(entity_id)
            self.tabs.setCurrentWidget(self.instruction_tab)

        else:
            self.clear_views()

        self.update_tree_admin_controls()

    # ================== Отображение контекста ==================

    def _invalidate_render_cache(self):
        self._rendered_checklist_task_id = None
        self._rendered_instruction_id = None

    def _instruction_from_task(self, task):
        instruction_id = task.get("instruction_real_id") or task.get("instruction_id")
        if not instruction_id:
            return None

        return {
            "instruction_id": instruction_id,
            "category_id": task.get("category_id"),
            "category_name": task.get("category_name", ""),
            "instruction_title": task.get("instruction_title", ""),
            "short_desc": task.get("short_desc", ""),
            "sections": task.get("sections") or [],
            "related_ids": task.get("related_ids") or [],
            "related_titles": task.get("related_titles") or [],
        }

    def clear_views(self):
        self.current_category = None
        self.current_task = None
        self.current_checklist_task = None
        self.current_instruction = None
        self.current_section_titles = []
        self._rendered_checklist_task_id = None
        self._rendered_instruction_id = None

        self.task_title_label.setText("Выберите задачу слева")
        self.task_category_label.setText("")
        self.task_instruction_label.setText("Чек-лист появится после выбора задачи.")

        self.task_checklist.blockSignals(True)
        self.task_checklist.clear()
        self.task_checklist.blockSignals(False)
        self.task_checklist.setEnabled(False)
        if hasattr(self, "edit_checklist_button"):
            self.edit_checklist_button.setEnabled(False)
        if hasattr(self, "prev_step_button"):
            self.prev_step_button.setEnabled(False)
        if hasattr(self, "next_step_button"):
            self.next_step_button.setEnabled(False)
        if hasattr(self, "reset_checklist_button"):
            self.reset_checklist_button.setEnabled(False)

        if hasattr(self, "category_tasks_group"):
            self.category_tasks_group.setVisible(False)
        self.task_instruction_label.setVisible(False)
        if hasattr(self, "edit_checklist_button"):
            self.edit_checklist_button.setEnabled(False)
        if hasattr(self, "prev_step_button"):
            self.prev_step_button.setEnabled(False)
            self.prev_step_button.setVisible(False)
        if hasattr(self, "next_step_button"):
            self.next_step_button.setEnabled(False)
            self.next_step_button.setVisible(False)
        if hasattr(self, "reset_checklist_button"):
            self.reset_checklist_button.setEnabled(False)
            self.reset_checklist_button.setVisible(False)
        if hasattr(self, "task_desc_group"):
            self.task_desc_group.setVisible(False)
        if hasattr(self, 'sticky_header_label'):
            self._reset_sticky_header(reset_scroll=True)
        if hasattr(self, '_section_cards'):
            self._section_cards = []

        if hasattr(self, "instruction_nav_group"):
            self.instruction_nav_group.setVisible(False)
            clear_layout(self.instruction_nav_layout)

        self.instruction_title_label.setText("Инструкция")
        self.instruction_desc_label.setText("Выберите задачу или инструкцию слева.")
        clear_layout(self.instruction_container_layout)
        placeholder = QLabel("Здесь будет подробная инструкция с раскрывающимися блоками.")
        placeholder.setWordWrap(True)
        self.instruction_container_layout.addWidget(placeholder)

        self._hide_checklist_sidebar()

        if hasattr(self, 'feedback_stats_label'):
            self.feedback_stats_label.setText("Выберите инструкцию слева")
        if hasattr(self, 'submit_feedback_btn'):
            self.submit_feedback_btn.setEnabled(False)
        if hasattr(self, 'feedback_comments_browser'):
            self.feedback_comments_browser.setHtml("<i>Выберите инструкцию слева.</i>")

        self.update_tree_admin_controls()
        self.statusBar().showMessage("Выберите задачу слева")

    def show_category(self, category_id):
        category = self.db.category_by_id(category_id)
        if not category:
            return

        self.current_category = category
        self.current_task = None
        self.current_instruction = None
        self.current_section_titles = []
        self._invalidate_render_cache()

        count = self.db.count_tasks_in_category(category_id)

        self.task_title_label.setText(category["name"])
        self.task_category_label.setText(f"Раздел: {category['name']}")
        if hasattr(self, 'task_desc_group'):
            self.task_desc_group.setVisible(True)
            self.task_desc_label.setText(render_markdown(
                f"В этом разделе задач: {count}. Выбери конкретную задачу слева или ниже по кнопке."
            ))
        self.task_instruction_label.setText("Инструкция: —")

        # На уровне "Раздел" чек-лист не показываем
        if category.get("parent_id") is None:
            self._hide_checklist_sidebar()
        else:
            self.current_checklist_task = self._first_task_in_category(category_id)
            self._set_checklist_sidebar_visible(True)
            has_checklist = self._render_checklist_for_task(self.current_checklist_task)
            if hasattr(self, "prev_step_button"):
                self.prev_step_button.setVisible(has_checklist)
                self.prev_step_button.setEnabled(has_checklist)
            if hasattr(self, "next_step_button"):
                self.next_step_button.setVisible(has_checklist)
                self.next_step_button.setEnabled(has_checklist)
            if hasattr(self, "reset_checklist_button"):
                self.reset_checklist_button.setVisible(has_checklist)
                self.reset_checklist_button.setEnabled(has_checklist)
            if has_checklist:
                self._scroll_checklist_to_current_point(self.current_checklist_task)

        self.task_instruction_label.setVisible(False)
        self._invalidate_render_cache()
        self.refresh_instruction_tab()
        self.refresh_feedback_tab()
        self.tabs.setCurrentWidget(self.instruction_tab)
        self._select_in_nav_tree(category_id=category_id)

    def show_task(self, task_id, force=False):
        task = self.db.task_bundle(task_id)
        if not task:
            return

        if force:
            self._invalidate_render_cache()

        self.current_task = task
        self.current_category = self.db.category_by_id(task["category_id"])
        self.current_instruction = self._instruction_from_task(task)
        self.current_section_titles = [
            s.get("title", "")
            for s in (self.current_instruction["sections"] if self.current_instruction else [])
        ]

        # Задача
        self.task_title_label.setText(task["task_title"])
        self.task_category_label.setText(f"Раздел: {task['category_name']}")
        if hasattr(self, 'task_desc_group'):
            self.task_desc_group.setVisible(True)
            html_desc = render_markdown(task["short_desc"])
            self.task_desc_label.setText(html_desc)
        if self.current_instruction:
            self.task_instruction_label.setText(f"Инструкция: {self.current_instruction['instruction_title']}")
        else:
            self.task_instruction_label.setText("Инструкция: не привязана")

        common_checklist_task = self._first_task_in_category(task["category_id"]) or task
        self.current_checklist_task = common_checklist_task
        self._set_checklist_sidebar_visible(True)
        has_checklist = self._render_checklist_for_task(common_checklist_task, force=force)

        if hasattr(self, "prev_step_button"):
            self.prev_step_button.setVisible(has_checklist)
            self.prev_step_button.setEnabled(has_checklist)
        if hasattr(self, "next_step_button"):
            self.next_step_button.setVisible(has_checklist)
            self.next_step_button.setEnabled(has_checklist)
        if hasattr(self, "reset_checklist_button"):
            self.reset_checklist_button.setVisible(has_checklist)
            self.reset_checklist_button.setEnabled(has_checklist)

        if has_checklist:
            self._scroll_checklist_to_current_point(task)

        self.task_instruction_label.setVisible(True)
        if hasattr(self, "prev_step_button"):
            self.prev_step_button.setEnabled(self.current_instruction is not None and self.task_checklist.count() > 0)
        if hasattr(self, "next_step_button"):
            self.next_step_button.setEnabled(self.current_instruction is not None and self.task_checklist.count() > 0)
        if hasattr(self, "category_tasks_group"):
            self.category_tasks_group.setVisible(False)

        # Инструкция
        self.refresh_instruction_tab(force=force)
        # Комментарии и оценка
        self.refresh_feedback_tab()

        self.tabs.setCurrentIndex(0)
        self.statusBar().showMessage(f"Открыта задача: {task['task_title']}")

        # Подсвечиваем поисковый текст в описании задачи
        search_text = self.search_edit.text().strip()
        if search_text:
            QTimer.singleShot(100, lambda: self._highlight_search_in_widget(self.task_desc_label, search_text))

        self._select_in_nav_tree(task_id=task["task_id"])

    def _first_task_in_category(self, category_id):
        """
        Возвращает первую задачу с заполненным чек-листом в выбранной ветке.
        Это нужно для уровня "задача": чек-лист может лежать не прямо в
        выбранной категории, а в первой вложенной подзадаче.
        """
        fallback_task = None

        for task_row in self.db.tasks_for_category(category_id):
            task = self.db.task_bundle(task_row["task_id"])
            if not task:
                continue
            if fallback_task is None:
                fallback_task = task
            if self._task_has_checklist(task):
                return task

        for child_category in self.db.categories_by_parent(category_id):
            task = self._first_task_in_category(child_category["id"])
            if task:
                if self._task_has_checklist(task):
                    return task
                if fallback_task is None:
                    fallback_task = task

        return fallback_task

    def _task_has_checklist(self, task):
        if not task:
            return False

        checklist_sections = normalize_checklist_sections(
            task.get("checklist_sections") or task.get("checklist") or []
        )
        return any(section.get("items") for section in checklist_sections)

    def _render_checklist_for_task(self, task, force=False):
        task_id = task["task_id"] if task else None
        if (
                not force
                and task_id is not None
                and task_id == self._rendered_checklist_task_id
                and self.task_checklist.count() > 0
        ):
            self.task_checklist.setEnabled(True)
            if hasattr(self, "edit_checklist_button"):
                self.edit_checklist_button.setEnabled(True)
            return len(self._task_checklist_items()) > 0

        self.task_checklist.blockSignals(True)
        self.task_checklist.clear()
        self._rendered_checklist_task_id = task_id

        if not task:
            self._rendered_checklist_task_id = None
            self.task_checklist.blockSignals(False)
            self.task_checklist.setEnabled(False)
            self._reset_checklist_sticky_header()
            if hasattr(self, "edit_checklist_button"):
                self.edit_checklist_button.setEnabled(False)
            if hasattr(self, "prev_step_button"):
                self.prev_step_button.setEnabled(False)
                self.prev_step_button.setVisible(False)
            if hasattr(self, "next_step_button"):
                self.next_step_button.setEnabled(False)
                self.next_step_button.setVisible(False)
            if hasattr(self, "reset_checklist_button"):
                self.reset_checklist_button.setEnabled(False)
                self.reset_checklist_button.setVisible(False)
            return False

        checklist_sections = normalize_checklist_sections(
            task.get("checklist_sections") or task.get("checklist") or []
        )
        saved_states = self.task_state_cache.get(task["task_id"], [])
        collapsed_sections = self.checklist_collapsed_sections.setdefault(task["task_id"], set())
        selected_item = None

        step_index = 0
        for section_index, section in enumerate(checklist_sections):
            section_title = section.get("title", "").strip() or "Секция"
            section_target_task_id = section.get("target_task_id")
            section_target_instruction_id = section.get("target_instruction_id")
            section_target_section_id = str(section.get("target_section_id", "")).strip()
            section_target_task_title = str(section.get("task_title", "")).strip()
            section_target_instruction_title = str(section.get("instruction_title", "")).strip()
            section_target_section_title = str(section.get("section_title", "")).strip()

            section_item = QListWidgetItem(section_title)
            section_item.setData(ROLE_CHECKLIST_KIND, "section")
            section_item.setData(ROLE_CHECKLIST_SECTION_INDEX, section_index)
            section_item.setData(ROLE_CHECKLIST_SECTION_COLLAPSED, section_index in collapsed_sections)
            section_item.setFlags(Qt.ItemIsEnabled)
            section_font = section_item.font()
            section_font.setBold(True)
            section_item.setFont(section_font)
            section_item.setToolTip(section_title)
            self.task_checklist.addItem(section_item)
            section_item.setHidden(False)

            for item_data in section.get("items", []):
                item_text = str(item_data.get("text", "")).strip()
                if not item_text:
                    continue

                item = QListWidgetItem(item_text)
                item.setData(ROLE_CHECKLIST_KIND, "item")
                item.setData(ROLE_CHECKLIST_SECTION_INDEX, section_index)
                item.setData(ROLE_STEP_INDEX, step_index)
                item_target_task_id = item_data.get("target_task_id")
                item_target_instruction_id = item_data.get("target_instruction_id")
                item_target_section_id = str(item_data.get("target_section_id", "")).strip()
                item_target_task_title = str(item_data.get("task_title", "")).strip()
                item_target_instruction_title = str(item_data.get("instruction_title", "")).strip()
                item_target_section_title = str(item_data.get("section_title", "")).strip()

                task_instruction_id = task.get("instruction_real_id") or task.get("instruction_id")
                if not checklist_item_overrides_section(
                        item_data, section, task_instruction_id, task["task_id"]
                ):
                    item_target_task_id = section_target_task_id
                    item_target_instruction_id = section_target_instruction_id
                    item_target_section_id = section_target_section_id
                    item_target_task_title = section_target_task_title
                    item_target_instruction_title = section_target_instruction_title
                    item_target_section_title = section_target_section_title

                item.setData(
                    ROLE_CHECKLIST_TARGET_INSTRUCTION,
                    item_target_instruction_title
                )
                item.setData(
                    ROLE_CHECKLIST_TARGET_SECTION,
                    item_target_section_title
                )
                item.setData(
                    ROLE_CHECKLIST_TARGET_TASK_ID,
                    item_target_task_id
                )
                item.setData(
                    ROLE_CHECKLIST_TARGET_INSTRUCTION_ID,
                    item_target_instruction_id
                )
                item.setData(
                    ROLE_CHECKLIST_TARGET_SECTION_ID,
                    item_target_section_id
                )
                item.setFlags(item.flags() | Qt.ItemIsUserCheckable | Qt.ItemIsEnabled)

                if item_target_task_id and item_target_task_title:
                    target_instruction = item_target_task_title
                else:
                    target_instruction = item_target_instruction_title or "эта подзадача"
                target_section = item_target_section_title or "авто"
                if (
                        not item_target_task_id
                        and item_target_instruction_id
                        and item_target_instruction_id == task.get("instruction_real_id")
                ):
                    target_instruction = "эта подзадача"
                if self.is_admin:
                    item.setToolTip(f"{item_text}\nПереход: {target_instruction} / {target_section}")
                else:
                    item.setToolTip(item_text)
                item.setStatusTip(item_text)

                state = saved_states[step_index] if step_index < len(saved_states) else Qt.Unchecked
                item.setCheckState(state)

                self.task_checklist.addItem(item)
                item.setHidden(section_index in collapsed_sections)
                if (
                        self.selected_checklist_task_id == task["task_id"]
                        and self.selected_checklist_step_index == step_index
                ):
                    selected_item = item
                step_index += 1

        self.task_checklist.doItemsLayout()
        self.task_checklist.updateGeometries()
        self.task_checklist.blockSignals(False)
        self.task_checklist.setEnabled(True)

        if selected_item is not None:
            selected_step_index = int(selected_item.data(ROLE_STEP_INDEX) or 0)
            self.task_checklist.setCurrentItem(selected_item)
            selected_item.setSelected(True)
            QTimer.singleShot(
                0,
                lambda step=selected_step_index: self._scroll_checklist_to_step_index(step)
            )
        else:
            self.task_checklist.clearSelection()

        if hasattr(self, "edit_checklist_button"):
            self.edit_checklist_button.setEnabled(True)
        if hasattr(self, "prev_step_button"):
            self.prev_step_button.setEnabled(len(self._task_checklist_items()) > 0)
            self.prev_step_button.setVisible(True)
        if hasattr(self, "next_step_button"):
            self.next_step_button.setEnabled(len(self._task_checklist_items()) > 0)
            self.next_step_button.setVisible(True)
        if hasattr(self, "reset_checklist_button"):
            self.reset_checklist_button.setEnabled(len(self._task_checklist_items()) > 0)
            self.reset_checklist_button.setVisible(True)
        self._update_checklist_collapse_button()

        return len(self._task_checklist_items()) > 0

    def _scroll_checklist_to_current_point(self, task=None):
        items = self._task_checklist_items()
        if not items:
            return

        focus_item = None
        task_title = ""
        if task:
            task_title = str(task.get("task_title") or "").casefold().strip()

        # 1) Сначала пробуем найти секцию, совпадающую с названием подзадачи
        if task_title:
            for item in items:
                if item.data(ROLE_CHECKLIST_KIND) == "section":
                    title = item.text().casefold().strip()
                    if title and (title == task_title or task_title in title or title in task_title):
                        focus_item = item
                        break

        # 2) Если не нашли — берём первый незавершённый пункт
        if focus_item is None:
            for item in items:
                if item.data(ROLE_CHECKLIST_KIND) == "item" and item.checkState() != Qt.Checked:
                    focus_item = item
                    break

        # 3) Если всё отмечено — последний пункт
        if focus_item is None:
            for item in reversed(items):
                if item.data(ROLE_CHECKLIST_KIND) == "item":
                    focus_item = item
                    break

        if focus_item is None:
            focus_item = items[0]

        focus_row = self.task_checklist.row(focus_item)
        scroll_row = focus_row

        for row in range(focus_row, -1, -1):
            candidate = self.task_checklist.item(row)
            if candidate and candidate.data(ROLE_CHECKLIST_KIND) == "section":
                scroll_row = row
                break

        QTimer.singleShot(
            0,
            lambda row=scroll_row: self._scroll_checklist_row_to_top(row)
        )

    def _scroll_checklist_row_to_top(self, row):
        if not hasattr(self, "task_checklist"):
            return

        if row is None:
            return

        if row < 0 or row >= self.task_checklist.count():
            return

        item = self.task_checklist.item(row)
        if item is None:
            return

        self.task_checklist.scrollToItem(item, QAbstractItemView.PositionAtTop)

    def _find_checklist_item_by_step(self, step_index):
        if step_index is None:
            return None

        for row in range(self.task_checklist.count()):
            item = self.task_checklist.item(row)
            if not item:
                continue
            if (
                    item.data(ROLE_CHECKLIST_KIND) == "item"
                    and int(item.data(ROLE_STEP_INDEX) or -1) == step_index
            ):
                return item
        return None

    def _scroll_checklist_to_step_index(self, step_index, position=QAbstractItemView.PositionAtCenter):
        item = self._find_checklist_item_by_step(step_index)
        if item is None:
            return

        self.task_checklist.setCurrentItem(item)
        item.setSelected(True)
        self.task_checklist.scrollToItem(item, position)

    def _set_checklist_sticky_visible(self, visible: bool):
        if hasattr(self, "checklist_sticky_header"):
            self.checklist_sticky_header.setVisible(visible)

    def _reset_checklist_sticky_header(self):
        if hasattr(self, "checklist_sticky_header"):
            self.checklist_sticky_header.setText("Свернуть чек-лист")
            self.checklist_sticky_header.setArrowType(Qt.UpArrow)
            self.checklist_sticky_header.setChecked(False)
        self._set_checklist_sticky_visible(False)

    def _checklist_section_collapsed_set(self):
        active_task = self.current_checklist_task or self.current_task
        if not active_task:
            return set()
        return self.checklist_collapsed_sections.setdefault(active_task["task_id"], set())

    def _set_checklist_section_collapsed(self, section_index, collapsed):
        if section_index is None:
            return

        collapsed_sections = self._checklist_section_collapsed_set()
        if collapsed:
            collapsed_sections.add(section_index)
        else:
            collapsed_sections.discard(section_index)

        for row in range(self.task_checklist.count()):
            item = self.task_checklist.item(row)
            if not item:
                continue
            if item.data(ROLE_CHECKLIST_KIND) == "section" and item.data(ROLE_CHECKLIST_SECTION_INDEX) == section_index:
                item.setData(ROLE_CHECKLIST_SECTION_COLLAPSED, collapsed)
            if item.data(ROLE_CHECKLIST_KIND) == "item" and item.data(ROLE_CHECKLIST_SECTION_INDEX) == section_index:
                item.setHidden(collapsed)

        self.task_checklist.viewport().update()
        self._update_checklist_collapse_button()

    def _toggle_checklist_section(self, section_index):
        collapsed_sections = self._checklist_section_collapsed_set()
        self._set_checklist_section_collapsed(section_index, section_index not in collapsed_sections)

    def _checklist_section_indices(self):
        indices = []
        for row in range(self.task_checklist.count()):
            item = self.task_checklist.item(row)
            if item and item.data(ROLE_CHECKLIST_KIND) == "section":
                indices.append(item.data(ROLE_CHECKLIST_SECTION_INDEX))
        return indices

    def _toggle_all_checklist_sections(self):
        indices = self._checklist_section_indices()
        if not indices:
            return

        collapsed_sections = self._checklist_section_collapsed_set()
        collapse_all = not all(index in collapsed_sections for index in indices)
        for section_index in indices:
            self._set_checklist_section_collapsed(section_index, collapse_all)

        self._update_checklist_collapse_button()

    def _update_checklist_collapse_button(self):
        if not hasattr(self, "task_checklist") or self.task_checklist.count() == 0:
            self._reset_checklist_sticky_header()
            return

        indices = self._checklist_section_indices()
        if not indices:
            self._reset_checklist_sticky_header()
            return

        collapsed_sections = self._checklist_section_collapsed_set()
        all_collapsed = all(index in collapsed_sections for index in indices)

        self.checklist_sticky_header.blockSignals(True)
        self.checklist_sticky_header.setText("Развернуть чек-лист" if all_collapsed else "Свернуть чек-лист")
        self.checklist_sticky_header.setChecked(all_collapsed)
        self.checklist_sticky_header.setArrowType(Qt.DownArrow if all_collapsed else Qt.UpArrow)
        self.checklist_sticky_header.blockSignals(False)
        self._set_checklist_sticky_visible(True)

    def _update_sticky_header(self):
        """Обновляет приклеенный заголовок на основе позиции прокрутки."""
        if not hasattr(self, '_section_cards') or not self._section_cards:
            self._set_sticky_header_visible(False)
            return

        scroll_y = self.instruction_scroll.verticalScrollBar().value()
        sticky_title = None
        sticky_index = None

        for idx, card in enumerate(self._section_cards):
            # Позиция кнопки-заголовка относительно контейнера
            toggle_pos = card.toggle.mapTo(self.instruction_container, card.toggle.pos())
            toggle_top = toggle_pos.y()

            if toggle_top <= scroll_y:
                sticky_title = card.toggle.text()
                sticky_index = idx
            else:
                break  # дальше заголовки ещё не видны

        if sticky_title:
            self.sticky_header_label.setText(sticky_title)
            self.sticky_header_section_index = sticky_index
            card = self._section_cards[sticky_index]
            self.sticky_header_label.blockSignals(True)
            self.sticky_header_label.setChecked(card.toggle.isChecked())
            self.sticky_header_label.setArrowType(Qt.DownArrow if card.toggle.isChecked() else Qt.RightArrow)
            self.sticky_header_label.blockSignals(False)
            self._set_sticky_header_visible(True)
        else:
            self.sticky_header_section_index = None
            self._set_sticky_header_visible(False)

    def _set_sticky_header_visible(self, visible: bool):
        if hasattr(self, "sticky_header_label"):
            self.sticky_header_label.setVisible(visible)

    def _toggle_sticky_header_section(self):
        idx = self.sticky_header_section_index
        if idx is None or not getattr(self, "_section_cards", None):
            return
        if idx < 0 or idx >= len(self._section_cards):
            return

        card = self._section_cards[idx]
        card.toggle.setChecked(not card.toggle.isChecked())
        self._update_sticky_header()

    def _reset_sticky_header(self, reset_scroll=False):
        self.sticky_header_section_index = None
        if hasattr(self, "sticky_header_label"):
            self.sticky_header_label.setText("")
            self.sticky_header_label.setArrowType(Qt.RightArrow)
        self._set_sticky_header_visible(False)

        if reset_scroll and hasattr(self, "instruction_scroll"):
            bar = self.instruction_scroll.verticalScrollBar()
            bar.blockSignals(True)
            bar.setValue(0)
            bar.blockSignals(False)

    def _task_checklist_items(self):
        """Возвращает только чекбокс-пункты, без заголовков секций."""
        if not hasattr(self, "task_checklist"):
            return []

        items = []
        for i in range(self.task_checklist.count()):
            item = self.task_checklist.item(i)
            if item.data(ROLE_CHECKLIST_KIND) == "item":
                items.append(item)
        return items

    def _set_checklist_sidebar_visible(self, visible: bool):
        self._checklist_content_available = bool(visible)
        if not visible:
            self._checklist_panel_user_hidden = False
        self._update_checklist_panel_visibility()

    def _hide_checklist_sidebar(self):
        self.current_checklist_task = None
        self._set_checklist_sidebar_visible(False)

        if hasattr(self, "task_checklist"):
            self.task_checklist.blockSignals(True)
            self.task_checklist.clear()
            self.task_checklist.blockSignals(False)
            self.task_checklist.setEnabled(False)
            self._reset_checklist_sticky_header()

        if hasattr(self, "edit_checklist_button"):
            self.edit_checklist_button.setEnabled(False)

        if hasattr(self, "prev_step_button"):
            self.prev_step_button.setVisible(False)
            self.prev_step_button.setEnabled(False)
        if hasattr(self, "next_step_button"):
            self.next_step_button.setVisible(False)
            self.next_step_button.setEnabled(False)
        if hasattr(self, "reset_checklist_button"):
            self.reset_checklist_button.setVisible(False)
            self.reset_checklist_button.setEnabled(False)

    def _instruction_section_index_by_title(self, instruction, section_title):
        """Ищет секцию инструкции по её заголовку."""
        if not instruction or not section_title:
            return None

        needle = section_title.casefold().strip()
        for idx, section in enumerate(instruction.get("sections", [])):
            title = (section.get("title") or "").casefold().strip()
            if title == needle:
                return idx
        return None

    def _instruction_section_index_by_id(self, instruction, section_id):
        """Ищет секцию инструкции по стабильному id."""
        return section_index_by_id(instruction, section_id)

    def _open_checklist_target(self, item):
        """Открывает инструкцию и нужную секцию для пункта чек-листа."""
        if not item or item.data(ROLE_CHECKLIST_KIND) != "item":
            return

        target_task_id = item.data(ROLE_CHECKLIST_TARGET_TASK_ID)
        try:
            target_task_id = int(target_task_id) if target_task_id else None
        except (TypeError, ValueError):
            target_task_id = None

        target_instruction_id = item.data(ROLE_CHECKLIST_TARGET_INSTRUCTION_ID)
        try:
            target_instruction_id = int(target_instruction_id) if target_instruction_id else None
        except (TypeError, ValueError):
            target_instruction_id = None

        target_section_id = (item.data(ROLE_CHECKLIST_TARGET_SECTION_ID) or "").strip()
        target_instruction_title = (item.data(ROLE_CHECKLIST_TARGET_INSTRUCTION) or "").strip()
        target_section_title = (item.data(ROLE_CHECKLIST_TARGET_SECTION) or "").strip()
        step_index = int(item.data(ROLE_STEP_INDEX) or 0)
        checklist_text = item.text()

        if target_task_id:
            current_task_id = self.current_task.get("task_id") if self.current_task else None
            if current_task_id != target_task_id:
                self.search_timer.stop()
                self.search_edit.blockSignals(True)
                self.search_edit.clear()
                self.search_edit.blockSignals(False)
                self.reload_nav_tree()
                self.show_task(target_task_id, force=True)
                self.tabs.setCurrentWidget(self.instruction_tab)
            else:
                self._select_in_nav_tree(task_id=target_task_id)
        elif target_instruction_id:
            current_instruction_id = self.current_instruction.get("instruction_id") if self.current_instruction else None
            if current_instruction_id != target_instruction_id:
                self.open_instruction_by_id(target_instruction_id)
        elif target_instruction_title:
            if not self.current_instruction or self.current_instruction["instruction_title"] != target_instruction_title:
                self.open_instruction_by_title(target_instruction_title)

        if not self.current_instruction:
            self.statusBar().showMessage("Для этого пункта не найдена инструкция.")
            return

        section_index = self._instruction_section_index_by_id(self.current_instruction, target_section_id)

        if section_index is None:
            section_index = self._instruction_section_index_by_title(self.current_instruction, target_section_title)

        if section_index is None:
            section_index = self._resolve_instruction_section_index(
                step_index,
                checklist_text,
                target_section_title
            )

        if section_index is None:
            self.statusBar().showMessage("Не удалось найти нужный фрагмент в инструкции.")
            return

        self.open_instruction_section(section_index)

    def show_instruction_only(self, instruction_id):
        instruction = self.db.instruction_by_id(instruction_id)
        if not instruction:
            return

        self.current_task = None
        self.current_instruction = instruction
        self.current_section_titles = [
            s.get("title", "")
            for s in instruction.get("sections", [])
        ]

        self.current_checklist_task = None
        self.task_checklist.blockSignals(True)
        self.task_checklist.clear()
        self.task_checklist.blockSignals(False)
        self.task_checklist.setEnabled(False)
        if hasattr(self, "prev_step_button"):
            self.prev_step_button.setVisible(False)
            self.prev_step_button.setEnabled(False)
        if hasattr(self, "next_step_button"):
            self.next_step_button.setVisible(False)
            self.next_step_button.setEnabled(False)
        if hasattr(self, "reset_checklist_button"):
            self.reset_checklist_button.setVisible(False)
            self.reset_checklist_button.setEnabled(False)
        self.refresh_instruction_tab()
        self.refresh_feedback_tab()
        self.tabs.setCurrentWidget(self.instruction_tab)
        self.statusBar().showMessage(f"Открыта инструкция: {instruction['instruction_title']}")

    def move_tree_item_up(self):
        """Перемещает выбранный элемент дерева вверх."""
        if not self.is_admin:
            return
        kind, entity_id = self.selected_tree_entity()
        if kind == "category":
            self.db.move_category_up(entity_id)
        elif kind == "task":
            self.db.move_task_up(entity_id)
        else:
            return
        self.reload_nav_tree()

    def move_tree_item_down(self):
        """Перемещает выбранный элемент дерева вниз."""
        if not self.is_admin:
            return
        kind, entity_id = self.selected_tree_entity()
        if kind == "category":
            self.db.move_category_down(entity_id)
        elif kind == "task":
            self.db.move_task_down(entity_id)
        else:
            return
        self.reload_nav_tree()

    # ================== Вкладка задачи ==================

    def open_task_editor(self):
        """Открывает отдельное окно для редактирования задачи."""
        if not self.is_admin:
            return

        active_task = self.current_task or self.current_checklist_task
        if not active_task:
            return

        saved_task_id = active_task["task_id"]
        saved_category_id = self.current_category["id"] if self.current_category else None
        opened_from_category = self.current_task is None and saved_category_id is not None

        active_instruction = self.current_instruction
        instruction_id = active_task.get("instruction_real_id") or active_task.get("instruction_id")
        if active_instruction is None and instruction_id:
            active_instruction = self.db.instruction_by_id(instruction_id)

        dialog = TaskEditorDialog(
            db=self.db,
            task=active_task,
            instruction=active_instruction,
            parent=self
        )

        if dialog.exec() != QDialog.Accepted:
            return

        data = dialog.get_data()
        try:
            self.db.update_task_view_data(
                saved_task_id,
                data["short_desc"],
                data["instruction_title"],
                data["checklist_sections"]
            )
        except Exception as exc:
            QMessageBox.warning(self, "Ошибка", f"Не удалось сохранить изменения: {exc}")
            return

        self.reload_nav_tree()
        if opened_from_category and saved_category_id:
            self.show_category(saved_category_id)
        else:
            self.show_task(saved_task_id, force=True)
        self.remember_db_state()
        self.statusBar().showMessage("Задача обновлена")

    def refresh_task_tab(self):
        if self.current_task:
            self.task_title_label.setText(self.current_task["task_title"])
            self.task_category_label.setText(f"Раздел: {self.current_task['category_name']}")
            if hasattr(self, 'task_desc_group'):
                self.task_desc_group.setVisible(True)
                self.task_desc_label.setText(render_markdown(self.current_task["short_desc"]))
            if self.current_instruction:
                self.task_instruction_label.setText(f"Инструкция: {self.current_instruction['instruction_title']}")
            else:
                self.task_instruction_label.setText("Инструкция: не привязана")
            self.task_checklist.setEnabled(True)
            if hasattr(self, "edit_checklist_button"):
                self.edit_checklist_button.setEnabled(True)
            if hasattr(self, "prev_step_button"):
                self.prev_step_button.setEnabled(self.current_instruction is not None and self.task_checklist.count() > 0)
            if hasattr(self, "next_step_button"):
                self.next_step_button.setEnabled(self.current_instruction is not None and self.task_checklist.count() > 0)
            if hasattr(self, "reset_checklist_button"):
                self.reset_checklist_button.setEnabled(self.task_checklist.count() > 0)

            self.task_instruction_label.setVisible(True)
            if hasattr(self, "prev_step_button"):
                self.prev_step_button.setEnabled(self.current_instruction is not None and self.task_checklist.count() > 0)
            if hasattr(self, "next_step_button"):
                self.next_step_button.setEnabled(self.current_instruction is not None and self.task_checklist.count() > 0)
            if hasattr(self, "reset_checklist_button"):
                self.reset_checklist_button.setEnabled(self.task_checklist.count() > 0)
            if hasattr(self, "category_tasks_group"):
                self.category_tasks_group.setVisible(False)
            return

        if self.current_instruction and not self.current_task:
            self.task_title_label.setText(self.current_instruction["instruction_title"])
            self.task_category_label.setText(f"Раздел: {self.current_instruction['category_name']}")
            if hasattr(self, 'task_desc_group'):
                self.task_desc_group.setVisible(True)
                self.task_desc_label.setText(render_markdown(self.current_instruction["short_desc"]))
            self.task_instruction_label.setText("Открыто по ссылке без отдельной задачи")
            self.task_checklist.blockSignals(True)
            self.task_checklist.clear()
            self.task_checklist.blockSignals(False)
            self.task_checklist.setEnabled(False)
            if hasattr(self, "edit_checklist_button"):
                self.edit_checklist_button.setEnabled(False)

            self.task_instruction_label.setVisible(True)
            if hasattr(self, "category_tasks_group"):
                self.category_tasks_group.setVisible(False)
            return

        if self.current_category:
            self.task_title_label.setText(self.current_category["name"])
            self.task_category_label.setText(f"Раздел: {self.current_category['name']}")
            if hasattr(self, 'task_desc_group'):
                self.task_desc_group.setVisible(True)
                self.task_desc_label.setText(render_markdown(
                    f"В этом разделе задач: {self.current_category['task_count']}. "
                    "Выбери конкретную задачу слева или ниже по кнопке."
                ))
            self.task_instruction_label.setText("Выберите конкретную задачу слева")
            self.task_checklist.blockSignals(True)
            self.task_checklist.clear()
            self.task_checklist.blockSignals(False)
            self.task_checklist.setEnabled(False)

            self.task_instruction_label.setVisible(False)
            self.task_checklist_group.setVisible(False)
            if hasattr(self, "category_tasks_group"):
                self.category_tasks_group.setVisible(True)
                self.refresh_category_task_buttons()
            return

        self.task_title_label.setText("Выберите задачу слева")
        self.task_category_label.setText("")
        if hasattr(self, 'task_desc_group'):
            self.task_desc_group.setVisible(False)
        self.task_instruction_label.setText("")
        self.task_checklist.setEnabled(False)
        self.open_instruction_button.setEnabled(False)

    def on_checklist_item_changed(self, item):
        active_task = self.current_checklist_task or self.current_task
        if not active_task:
            return

        if item and item.data(ROLE_CHECKLIST_KIND) == "item":
            self.selected_checklist_task_id = active_task["task_id"]
            self.selected_checklist_step_index = int(item.data(ROLE_STEP_INDEX) or 0)

        states = [check_item.checkState() for check_item in self._task_checklist_items()]
        self.task_state_cache[active_task["task_id"]] = states

    def reset_checklist_states(self):
        active_task = self.current_checklist_task or self.current_task
        items = self._task_checklist_items()
        if not active_task or not items:
            return

        self.task_checklist.blockSignals(True)
        for item in items:
            item.setCheckState(Qt.Unchecked)
        self.task_checklist.blockSignals(False)

        self.task_state_cache[active_task["task_id"]] = [Qt.Unchecked for _ in items]
        self.selected_checklist_task_id = None
        self.selected_checklist_step_index = None
        self.task_checklist.clearSelection()
        self.statusBar().showMessage("Отметки чек-листа очищены")

    def on_checklist_item_clicked(self, item):
        if not item:
            self.task_checklist.clearSelection()
            return

        if item.data(ROLE_CHECKLIST_KIND) == "section":
            self.task_checklist.clearSelection()
            self._toggle_checklist_section(item.data(ROLE_CHECKLIST_SECTION_INDEX))
            return

        if item.data(ROLE_CHECKLIST_KIND) != "item":
            self.task_checklist.clearSelection()
            return

        active_task = self.current_checklist_task or self.current_task
        self.selected_checklist_task_id = active_task["task_id"] if active_task else None
        self.selected_checklist_step_index = int(item.data(ROLE_STEP_INDEX) or 0)
        self.task_checklist.setCurrentItem(item)
        item.setSelected(True)
        self._open_checklist_target(item)

    def _resolve_instruction_section_index(self, checklist_index, checklist_text, target_section_title=None):
        sections = self.current_instruction.get("sections", []) if self.current_instruction else []
        if not sections:
            return None

        if target_section_title:
            idx = self._instruction_section_index_by_title(self.current_instruction, target_section_title)
            if idx is not None:
                return idx

        needle = (checklist_text or "").casefold().strip()

        for idx, section in enumerate(sections):
            title = (section.get("title") or "").casefold().strip()
            if title and needle and (title in needle or needle in title):
                return idx

        if checklist_index < 0:
            checklist_index = 0
        return min(checklist_index, len(sections) - 1)

    def open_instruction_section(self, section_index):
        if not self.current_instruction:
            return

        if not getattr(self, "_section_cards", None):
            self.tabs.setCurrentWidget(self.instruction_tab)
            return

        section_index = max(0, min(section_index, len(self._section_cards) - 1))
        self.tabs.setCurrentWidget(self.instruction_tab)
        QTimer.singleShot(0, lambda idx=section_index: self._highlight_instruction_section(idx))

    def _highlight_instruction_section(self, section_index):
        if not getattr(self, "_section_cards", None):
            return

        section_index = max(0, min(section_index, len(self._section_cards) - 1))

        for idx, card in enumerate(self._section_cards):
            card.set_active(idx == section_index)

        card = self._section_cards[section_index]
        card.toggle.setChecked(True)
        self.sticky_header_section_index = section_index
        QTimer.singleShot(0, lambda c=card: self._scroll_instruction_card_to_top(c))
        self._update_sticky_header()
        self.statusBar().showMessage(f"Открыт шаг инструкции: {card.toggle.text()}")

    def _scroll_instruction_card_to_top(self, card):
        if not card or not hasattr(self, "instruction_scroll"):
            return

        bar = self.instruction_scroll.verticalScrollBar()
        top = card.mapTo(self.instruction_container, card.rect().topLeft()).y()
        bar.setValue(max(0, top))
        self._update_sticky_header()

    def _sync_instruction_progress_highlight(self):
        if not self.current_instruction or not getattr(self, "_section_cards", None):
            return

        items = self._task_checklist_items()
        if not items:
            return

        next_item = None
        for item in items:
            if item.checkState() != Qt.Checked:
                next_item = item
                break
        if next_item is None:
            next_item = items[-1]

        section_index = self._instruction_section_index_by_id(
            self.current_instruction,
            (next_item.data(ROLE_CHECKLIST_TARGET_SECTION_ID) or "").strip()
        )

        if section_index is None:
            section_index = self._instruction_section_index_by_title(
                self.current_instruction,
                (next_item.data(ROLE_CHECKLIST_TARGET_SECTION) or "").strip()
            )

        if section_index is None:
            section_index = int(next_item.data(ROLE_STEP_INDEX) or 0)

        section_index = max(0, min(section_index, len(self._section_cards) - 1))

        for idx, card in enumerate(self._section_cards):
            card.set_active(idx == section_index)

    def open_current_instruction(self):
        if self.current_instruction:
            self.tabs.setCurrentWidget(self.instruction_tab)
        else:
            QMessageBox.warning(self, "Внимание", "Для этой задачи пока не привязана инструкция.")

    def _navigate_checklist_item(self, item):
        if item is None:
            return
        active_task = self.current_checklist_task or self.current_task
        self.selected_checklist_task_id = active_task["task_id"] if active_task else None
        self.selected_checklist_step_index = int(item.data(ROLE_STEP_INDEX) or 0)
        self.task_checklist.setCurrentItem(item)
        item.setSelected(True)
        self._open_checklist_target(item)

    def go_to_previous_step(self):
        items = self._task_checklist_items()
        if not items:
            QMessageBox.information(self, "Внимание", "В чек-листе нет пунктов.")
            return

        current_item = self.task_checklist.currentItem()
        prev_item = None

        if current_item in items:
            current_index = items.index(current_item)
            if current_index > 0:
                prev_item = items[current_index - 1]
            else:
                QMessageBox.information(self, "Внимание", "Это первый пункт чек-листа.")
                prev_item = current_item
        else:
            prev_item = items[-1]

        self._navigate_checklist_item(prev_item)

    def go_to_next_step(self):
        items = self._task_checklist_items()
        if not items:
            QMessageBox.information(self, "Внимание", "В чек-листе нет пунктов.")
            return

        next_item = None
        current_item = self.task_checklist.currentItem()
        current_index = -1

        if current_item in items:
            current_index = items.index(current_item)
            if current_item.checkState() != Qt.Checked:
                current_item.setCheckState(Qt.Checked)

            for item in items[current_index + 1:]:
                if item.checkState() != Qt.Checked:
                    next_item = item
                    break
        else:
            for item in items:
                if item.checkState() != Qt.Checked:
                    next_item = item
                    break

        if next_item is None:
            QMessageBox.information(
                self,
                "Готово",
                "Все пункты чек-листа уже отмечены."
            )
            if current_item in items:
                next_item = current_item
            else:
                next_item = items[-1]

        self._navigate_checklist_item(next_item)

    def refresh_category_task_buttons(self):
        if not hasattr(self, "category_tasks_layout"):
            return

        clear_layout(self.category_tasks_layout)

        if not self.current_category:
            return

        tasks = self.db.tasks_for_category(self.current_category["id"])
        if not tasks:
            lbl = QLabel("В этом разделе пока нет задач.")
            lbl.setWordWrap(True)
            lbl.setStyleSheet("color: #5b6577;")
            self.category_tasks_layout.addWidget(lbl)
            return

        for task in tasks:
            btn = QPushButton(short_button_text(task["task_title"]))
            btn.setToolTip(task["task_title"])
            btn.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
            btn.clicked.connect(lambda _, tid=task["task_id"]: self.show_task(tid))
            self.category_tasks_layout.addWidget(btn)

    def apply_admin_mode(self):
        if hasattr(self, "tree_admin_group"):
            self.tree_admin_group.setVisible(self.is_admin)
        if hasattr(self, "edit_task_button"):
            self.edit_task_button.setVisible(self.is_admin)
        if hasattr(self, "export_feedback_button"):
            self.export_feedback_button.setVisible(self.is_admin)
        if hasattr(self, "edit_checklist_button"):
            self.edit_checklist_button.setVisible(self.is_admin)
        if hasattr(self, "add_subcategory_button"):
            self.add_subcategory_button.setVisible(self.is_admin)

        if self.is_admin:
            self.statusBar().showMessage(f"Администратор: {self.username}")

        if hasattr(self, "manage_feedback_btn"):
            self.manage_feedback_btn.setVisible(self.is_admin)

        if hasattr(self, "move_up_button"):
            self.move_up_button.setVisible(self.is_admin)
        if hasattr(self, "move_down_button"):
            self.move_down_button.setVisible(self.is_admin)

        self.update_tree_admin_controls()

    def resizeEvent(self, event):
        super().resizeEvent(event)
        overlay = getattr(self, "_onboarding_overlay", None)
        if overlay and overlay.isVisible():
            overlay._fit_to_parent()
            overlay.refresh_current_step()

    def _find_first_nav_task_item(self):
        def walk(item):
            if item.data(0, ROLE_KIND) == "task":
                return item
            for i in range(item.childCount()):
                found = walk(item.child(i))
                if found:
                    return found
            return None

        for i in range(self.nav_tree.topLevelItemCount()):
            found = walk(self.nav_tree.topLevelItem(i))
            if found:
                return found
        return None

    def _nav_tree_item_rect(self, item, overlay):
        if item is None or overlay is None:
            return None
        parent = item.parent()
        while parent:
            parent.setExpanded(True)
            parent = parent.parent()
        self.nav_tree.scrollToItem(item, QAbstractItemView.EnsureVisible)
        QApplication.processEvents()
        rect = self.nav_tree.visualItemRect(item)
        top_left = self.nav_tree.viewport().mapToGlobal(rect.topLeft())
        return QRect(overlay.mapFromGlobal(top_left), rect.size())

    def _select_demo_task_for_onboarding(self):
        item = self._find_first_nav_task_item()
        if not item:
            return
        task_id = item.data(0, ROLE_ID)
        self.show_task(task_id)

    def _onboarding_step_buttons_rect(self, overlay):
        buttons = [b for b in (self.prev_step_button, self.next_step_button) if b.isVisible()]
        if not buttons:
            return None
        united = None
        for btn in buttons:
            top_left = btn.mapToGlobal(QPoint(0, 0))
            rect = QRect(overlay.mapFromGlobal(top_left), btn.size())
            united = rect if united is None else united.united(rect)
        return united.adjusted(-2, -2, 2, 2)

    def _prepare_onboarding_comments_demo(self):
        """Открывает демо-задачу и раскрывает первый блок инструкции для шага про комментарии."""
        self._select_demo_task_for_onboarding()
        cards = getattr(self, "_section_cards", None) or []
        if not cards:
            return
        first_card = cards[0]
        if not first_card.toggle.isChecked():
            first_card.toggle.setChecked(True)
        QTimer.singleShot(80, lambda c=first_card: self._scroll_instruction_card_to_top(c))

    def _onboarding_section_comment_button_rect(self, overlay):
        cards = getattr(self, "_section_cards", None) or []
        for card in cards:
            btn = getattr(card, "comment_button", None)
            if btn is not None and btn.isVisible():
                top_left = btn.mapToGlobal(QPoint(0, 0))
                return QRect(overlay.mapFromGlobal(top_left), btn.size()).adjusted(-6, -4, 6, 4)
        if cards:
            card = cards[0]
            footer = getattr(card, "footer", None)
            if footer is not None and footer.isVisible():
                top_left = footer.mapToGlobal(QPoint(0, 0))
                return QRect(overlay.mapFromGlobal(top_left), footer.size()).adjusted(-4, -4, 4, 4)
        return None

    def _onboarding_help_menu_rect(self, overlay):
        menu_bar = self.menuBar()
        action = getattr(self, "help_menu_bar_action", None)
        if menu_bar is None or not menu_bar.isVisible() or action is None:
            return None

        geo = menu_bar.actionGeometry(action)
        if geo.isNull():
            action = self.help_menu.menuAction()
            geo = menu_bar.actionGeometry(action)
        if geo.isNull():
            return None

        top_left = menu_bar.mapToGlobal(geo.topLeft())
        bottom_right = menu_bar.mapToGlobal(geo.bottomRight())
        overlay_top_left = overlay.mapFromGlobal(top_left)
        return QRect(
            overlay_top_left.x(),
            overlay_top_left.y(),
            bottom_right.x() - top_left.x() + 1,
            bottom_right.y() - top_left.y() + 1,
        )

    def _expand_first_nav_branch(self):
        if self.nav_tree.topLevelItemCount() <= 0:
            return
        section = self.nav_tree.topLevelItem(0)
        section.setExpanded(True)
        if section.childCount() > 0:
            section.child(0).setExpanded(True)

    def start_onboarding(self, mark_completed=False):
        if self._onboarding_overlay and isValid(self._onboarding_overlay):
            self._onboarding_overlay.close()
            self._onboarding_overlay.deleteLater()
            self._onboarding_overlay = None

        steps = self._build_onboarding_steps()

        def on_completed():
            self.menuBar().setEnabled(True)
            if mark_completed:
                self.db.set_onboarding_completed(self.username, True)
            self._onboarding_overlay = None

        self._onboarding_overlay = OnboardingOverlay(
            self,
            steps,
            on_completed=on_completed,
            parent=self,
        )
        self.menuBar().setEnabled(False)
        self._onboarding_overlay.start()

    def _build_onboarding_steps(self):
        return [
            {
                "title": "Дерево задач",
                "text": (
                    "Слева находится дерево задач. Раскройте раздел, выберите задачу, "
                    "затем подзадачу — по ней откроется инструкция и чек-лист."
                ),
                "widget_getter": lambda w: w.nav_tree,
                "placement": "right",
                "on_enter": lambda w: w._expand_first_nav_branch(),
            },
            {
                "title": "Выбор подзадачи",
                "text": (
                    "Подсвечена первая доступная подзадача. Кликните по названию в дереве, "
                    "чтобы открыть её инструкцию."
                ),
                "rect_getter": lambda w, o: w._nav_tree_item_rect(w._find_first_nav_task_item(), o),
                "widget_getter": lambda w: w.nav_tree,
                "on_enter": lambda w: w._select_demo_task_for_onboarding(),
                "placement": "right",
                "padding": 0,
                "geometry_delay": 15,
            },
            {
                "title": "Поиск",
                "text": (
                    "Поле поиска помогает быстро найти задачу, инструкцию или фрагмент текста "
                    "внутри блоков."
                ),
                "widget_getter": lambda w: w.search_edit,
                "placement": "bottom",
            },
            {
                "title": "Инструкция",
                "text": (
                    "В центре — подробная инструкция. Блоки можно раскрывать и сворачивать; "
                    "изображения увеличиваются по кнопке «Увеличить»."
                ),
                "widget_getter": lambda w: w.instruction_scroll,
                "placement": "left",
                "on_enter": lambda w: w._select_demo_task_for_onboarding(),
            },
            {
                "title": "Комментарии к блокам",
                "text": (
                    "В каждом раскрытом блоке инструкции есть кнопка «Комментарий». "
                    "Нажмите её, чтобы оценить блок по шкале 1–10 и оставить замечание. "
                    "Комментарий можно добавить к любому пункту отдельно — не только ко всей инструкции. "
                    "Если ставите оценку ниже 7, обязательно опишите, что непонятно или что улучшить."
                ),
                "rect_getter": lambda w, o: w._onboarding_section_comment_button_rect(o),
                "widget_getter": lambda w: w.instruction_scroll,
                "placement": "left",
                "on_enter": lambda w: w._prepare_onboarding_comments_demo(),
                "geometry_delay": 150,
            },
            {
                "title": "Чек-лист",
                "text": (
                    "Справа — чек-лист выбранной задачи. Отмечайте выполненные пункты; "
                    "клик по пункту открывает нужный фрагмент инструкции."
                ),
                "widget_getter": lambda w: w.task_checklist_group,
                "placement": "left",
                "on_enter": lambda w: w._select_demo_task_for_onboarding(),
            },
            {
                "title": "Навигация по шагам",
                "text": (
                    "Кнопки «Предыдущий шаг» и «Следующий шаг» помогают двигаться по чек-листу "
                    "без ручного поиска пунктов."
                ),
                "rect_getter": lambda w, o: w._onboarding_step_buttons_rect(o),
                "widget_getter": lambda w: w.prev_step_button,
                "placement": "left",
                "on_enter": lambda w: w._select_demo_task_for_onboarding(),
            },
            {
                "title": "Скрыть дерево",
                "text": (
                    "Дерево задач можно скрыть этой кнопкой. Чтобы вернуть панель, "
                    "нажмите стрелку «▶» у левого края окна."
                ),
                "widget_getter": lambda w: w.toggle_nav_panel_btn,
                "placement": "bottom",
            },
            {
                "title": "Повтор обучения",
                "text": (
                    "В любой момент можно снова пройти обучение через меню "
                    "«Справка → Обучение»."
                ),
                "rect_getter": lambda w, o: w._onboarding_help_menu_rect(o),
                "widget_getter": lambda w: w.menuBar(),
                "placement": "bottom",
                "padding": 0,
                "geometry_delay": 120,
            },
        ]

    def selected_tree_entity(self):
        item = self.nav_tree.currentItem()
        if not item:
            return None, None
        return item.data(0, ROLE_KIND), item.data(0, ROLE_ID)

    def update_tree_admin_controls(self):
        if not hasattr(self, "add_category_button"):
            return
        if not self.is_admin:
            return

        kind, entity_id = self.selected_tree_entity()
        selected_category = self.db.category_by_id(entity_id) if kind == "category" else None
        selected_task = self.db.task_bundle(entity_id) if kind == "task" else None

        can_modify = kind in {"category", "task"}
        can_move = kind in {"category", "task"}

        # Раздел — всегда можно добавить
        self.add_category_button.setEnabled(True)

        # Подкатегория — если выбрана категория или задача
        self.add_subcategory_button.setEnabled(kind in {"category", "task"})

        # Задача — только внутри подкатегории
        can_add_task = False
        if selected_category:
            can_add_task = selected_category.get("parent_id") is not None
        elif selected_task:
            task_category = self.db.category_by_id(selected_task["category_id"])
            can_add_task = bool(task_category and task_category.get("parent_id") is not None)

        self.add_task_button.setEnabled(can_add_task)

        # Инструкция — только для задачи без инструкции
        self.add_instruction_button.setEnabled(kind == "task")

        self.edit_tree_button.setEnabled(can_modify)
        self.delete_tree_button.setEnabled(can_modify)

        if hasattr(self, "move_up_button"):
            self.move_up_button.setEnabled(can_move)
        if hasattr(self, "move_down_button"):
            self.move_down_button.setEnabled(can_move)

        self.save_tree_button.setEnabled(True)

    def add_tree_category(self):
        if not self.is_admin:
            return

        name, ok = QInputDialog.getText(self, "Добавить раздел", "Название раздела:")
        if not ok:
            return

        name = name.strip()
        if not name:
            QMessageBox.warning(self, "Внимание", "Название раздела не может быть пустым.")
            return

        try:
            self.db.add_category(name, parent_id=None)
        except Exception as exc:
            QMessageBox.warning(self, "Ошибка", f"Не удалось добавить раздел: {exc}")
            return

        self.reload_nav_tree()

    def add_tree_subcategory(self):
        if not self.is_admin:
            return

        kind, entity_id = self.selected_tree_entity()
        parent_id = None

        if kind == "category":
            category = self.db.category_by_id(entity_id)
            if not category:
                return

            # Если выбрана подкатегория — добавляем в её верхний раздел
            parent_id = category["id"] if category.get("parent_id") is None else category["parent_id"]

        elif kind == "task":
            task = self.db.task_bundle(entity_id)
            if not task:
                return

            category = self.db.category_by_id(task["category_id"])
            if not category:
                return

            parent_id = category["id"] if category.get("parent_id") is None else category["parent_id"]

        else:
            QMessageBox.warning(self, "Внимание", "Сначала выберите раздел или задачу.")
            return

        name, ok = QInputDialog.getText(self, "Добавить задачу", "Название задачи:")
        if not ok:
            return

        name = name.strip()
        if not name:
            QMessageBox.warning(self, "Внимание", "Название задачи не может быть пустым.")
            return

        try:
            self.db.add_category(name, parent_id=parent_id)
        except Exception as exc:
            QMessageBox.warning(self, "Ошибка", f"Не удалось добавить подкатегорию: {exc}")
            return

        self.reload_nav_tree()

    def add_tree_task(self):
        if not self.is_admin:
            return

        kind, entity_id = self.selected_tree_entity()

        category_id = None

        if kind == "category":
            category = self.db.category_by_id(entity_id)
            if not category:
                return

            if category.get("parent_id") is None:
                QMessageBox.warning(
                    self,
                    "Внимание",
                    "Выберите задачу, а не сам раздел."
                )
                return

            category_id = entity_id

        elif kind == "task":
            task = self.db.task_bundle(entity_id)
            if not task:
                return

            category = self.db.category_by_id(task["category_id"])
            if not category or category.get("parent_id") is None:
                QMessageBox.warning(
                    self,
                    "Внимание",
                    "Задачи можно добавлять только в верхний раздел."
                )
                return

            category_id = task["category_id"]

        else:
            QMessageBox.warning(self, "Внимание", "Сначала выберите задачу или подзадачу.")
            return

        title, ok = QInputDialog.getText(self, "Добавить подзадачу", "Название подзадачи:")
        if not ok:
            return

        instruction_id = None
        instructions = self.db.all_instructions()

        if instructions:
            items = ["(без инструкции)"]
            instruction_id_by_label = {}

            for inst in instructions:
                label = f"#{inst['instruction_id']} • {inst['category_name']} — {inst['title']}"
                items.append(label)
                instruction_id_by_label[label] = inst["instruction_id"]

            chosen_label, ok = QInputDialog.getItem(
                self,
                "Добавить задачу",
                "Выберите существующую инструкцию или оставьте без инструкции:",
                items,
                0,
                False
            )
            if not ok:
                return

            if chosen_label != "(без инструкции)":
                instruction_id = instruction_id_by_label[chosen_label]

        try:
            self.db.add_task(category_id, title, instruction_id=instruction_id)
        except Exception as exc:
            QMessageBox.warning(self, "Ошибка", f"Не удалось добавить задачу: {exc}")
            return

        self.reload_nav_tree()

    def add_tree_instruction(self):
        if not self.is_admin:
            return

        kind, entity_id = self.selected_tree_entity()
        if kind != "task":
            QMessageBox.warning(self, "Внимание", "Выберите подзадачу, к которой нужно добавить инструкцию.")
            return

        task = self.db.task_bundle(entity_id)
        if not task:
            return

        instructions = self.db.all_instructions()
        items = ["➕ Создать новую инструкцию"]
        instruction_id_by_label = {}

        for inst in instructions:
            label = f"#{inst['instruction_id']} • {inst['category_name']} — {inst['title']}"
            items.append(label)
            instruction_id_by_label[label] = inst["instruction_id"]

        chosen_label, ok = QInputDialog.getItem(
            self,
            "Добавить инструкцию",
            "Выберите существующую инструкцию или создайте новую:",
            items,
            0,
            False
        )
        if not ok:
            return

        # 1) Привязать существующую инструкцию
        if chosen_label != "➕ Создать новую инструкцию":
            instruction_id = instruction_id_by_label[chosen_label]
            instruction = self.db.instruction_by_id(instruction_id)
            if not instruction:
                QMessageBox.warning(self, "Ошибка", "Инструкция не найдена.")
                return

            if task.get("instruction_id") == instruction["instruction_id"]:
                QMessageBox.information(self, "Внимание", "Эта инструкция уже привязана к подзадаче.")
                return

            if task.get("instruction_id"):
                reply = QMessageBox.question(
                    self,
                    "Подтверждение",
                    "У этой подзадачи уже есть инструкция. Для изменения используй кнопку «Редактировать инструкцию»."
                )
                if reply != QMessageBox.Yes:
                    return

            try:
                self.db.set_task_instruction(task["task_id"], instruction["instruction_id"])
            except Exception as exc:
                QMessageBox.warning(self, "Ошибка", f"Не удалось привязать инструкцию: {exc}")
                return

            self.reload_nav_tree()
            self.show_task(task["task_id"], force=True)
            return

        # 2) Создать новую инструкцию
        categories = self.db.conn.execute("""
            SELECT id, name
            FROM categories
            ORDER BY sort_order, name
        """).fetchall()

        if not categories:
            QMessageBox.warning(self, "Внимание", "Сначала добавь хотя бы один раздел.")
            return

        dialog = InstructionEditorDialog(
            db=self.db,
            categories=[dict(row) for row in categories],
            default_category_id=task["category_id"],
            default_task_id=task["task_id"],
            parent=self
        )

        if dialog.exec() != QDialog.Accepted:
            return

        data = dialog.get_data()
        category_id = data["category_id"]
        task_id = data["task_id"]
        instruction_title = data["title"]
        short_desc = data["short_desc"]
        sections = data["sections"]
        related_ids = data["related_ids"]

        if not task_id:
            QMessageBox.warning(self, "Ошибка", "Не указана задача для привязки инструкции.")
            return

        if not instruction_title:
            QMessageBox.warning(self, "Внимание", "Название инструкции не может быть пустым.")
            return

        if not short_desc:
            short_desc = instruction_title

        try:
            instruction_id, missing_related = self.db.add_instruction(
                category_id,
                task_id,
                instruction_title,
                short_desc,
                sections,
                related_ids
            )
        except Exception as exc:
            QMessageBox.warning(self, "Ошибка", f"Не удалось добавить инструкцию: {exc}")
            return

        self.reload_nav_tree()
        self.show_task(task_id, force=True)

        if missing_related:
            QMessageBox.information(
                self,
                "Внимание",
                "Не найдены связанные инструкции:\n" + "\n".join(missing_related)
            )

    def edit_tree_item(self):
        if not self.is_admin:
            return

        kind, entity_id = self.selected_tree_entity()
        if kind == "category":
            category = self.db.category_by_id(entity_id)
            if not category:
                return

            new_name, ok = QInputDialog.getText(
                self,
                "Редактировать категорию",
                "Новое название категории:",
                QLineEdit.Normal,
                category["name"]
            )
            if not ok:
                return

            new_name = new_name.strip()
            if not new_name:
                QMessageBox.warning(self, "Внимание", "Название категории не может быть пустым.")
                return

            try:
                self.db.rename_category(entity_id, new_name)
            except Exception as exc:
                QMessageBox.warning(self, "Ошибка", f"Не удалось изменить категорию: {exc}")
                return

        elif kind == "task":
            task = self.db.task_bundle(entity_id)
            if not task:
                return

            new_name, ok = QInputDialog.getText(
                self,
                "Редактировать задачу",
                "Новое название задачи:",
                QLineEdit.Normal,
                task["task_title"]
            )
            if not ok:
                return

            new_name = new_name.strip()
            if not new_name:
                QMessageBox.warning(self, "Внимание", "Название задачи не может быть пустым.")
                return

            try:
                self.db.rename_task(entity_id, new_name)
            except Exception as exc:
                QMessageBox.warning(self, "Ошибка", f"Не удалось изменить задачу: {exc}")
                return
        else:
            QMessageBox.warning(self, "Внимание", "Сначала выберите категорию или задачу.")
            return

        self.reload_nav_tree()

    def delete_tree_item(self):
        if not self.is_admin:
            return

        kind, entity_id = self.selected_tree_entity()
        if kind == "category":
            category = self.db.category_by_id(entity_id)
            if not category:
                return

            reply = QMessageBox.question(
                self,
                "Подтверждение",
                f"Удалить категорию «{category['name']}» вместе со всеми её задачами и инструкциями?"
            )
            if reply != QMessageBox.Yes:
                return

            try:
                self.db.delete_category(entity_id)
            except Exception as exc:
                QMessageBox.warning(self, "Ошибка", f"Не удалось удалить категорию: {exc}")
                return

        elif kind == "task":
            task = self.db.task_bundle(entity_id)
            if not task:
                return

            reply = QMessageBox.question(
                self,
                "Подтверждение",
                f"Удалить задачу «{task['task_title']}»?"
            )
            if reply != QMessageBox.Yes:
                return

            try:
                self.db.delete_task(entity_id)
            except Exception as exc:
                QMessageBox.warning(self, "Ошибка", f"Не удалось удалить задачу: {exc}")
                return
        else:
            QMessageBox.warning(self, "Внимание", "Сначала выберите категорию или задачу.")
            return

        self.reload_nav_tree()

    def save_tree_changes(self):
        if not self.is_admin:
            return

        self.reload_nav_tree()
        self.statusBar().showMessage("Дерево обновлено")

    # ================== Вкладка инструкции ==================

    def refresh_instruction_navigation_panel(self):
        if not hasattr(self, "instruction_nav_group"):
            return

        clear_layout(self.instruction_nav_layout)
        self.instruction_nav_group.setVisible(False)

        # Показываем навигацию только в разделе/подкатегории без открытой инструкции
        if self.current_instruction or not self.current_category or self.current_task:
            return

        child_categories = self.db.categories_by_parent(self.current_category["id"])
        if child_categories:
            self.instruction_nav_group.setTitle("Задачи")

            for cat in child_categories:
                btn = QPushButton(short_button_text(cat["name"]))
                btn.setToolTip(cat["name"])
                btn.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
                btn.clicked.connect(lambda _, cid=cat["id"]: self.show_category(cid))
                self.instruction_nav_layout.addWidget(btn)

            self.instruction_nav_group.setVisible(True)
            return

        tasks = self.db.tasks_for_category(self.current_category["id"])
        if tasks:
            self.instruction_nav_group.setTitle("Подзадачи")

            for task in tasks:
                btn = QPushButton(short_button_text(task["task_title"]))
                btn.setToolTip(task["task_title"])
                btn.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
                btn.clicked.connect(lambda _, tid=task["task_id"]: self.show_task(tid))
                self.instruction_nav_layout.addWidget(btn)

            self.instruction_nav_group.setVisible(True)

    def refresh_instruction_tab(self, force=False):
        instruction_id = (
            self.current_instruction.get("instruction_id")
            if self.current_instruction else None
        )
        if (
                not force
                and instruction_id is not None
                and instruction_id == self._rendered_instruction_id
                and getattr(self, "_section_cards", None)
        ):
            if hasattr(self, 'edit_instruction_btn'):
                self.edit_instruction_btn.setVisible(
                    self.current_instruction is not None and self.is_admin
                )
            self.refresh_instruction_navigation_panel()
            self._fit_instruction_content_width()
            return

        self._reset_sticky_header(reset_scroll=True)
        self._section_cards = []
        clear_layout(self.instruction_container_layout)
        self._rendered_instruction_id = instruction_id
        # Показываем кнопку редактирования, только если есть инструкция и пользователь админ
        if hasattr(self, 'edit_instruction_btn'):
            self.edit_instruction_btn.setVisible(
                self.current_instruction is not None and self.is_admin
            )
        self.current_section_titles = []

        if not self.current_instruction:
            if self.current_task:
                # Задача есть, но инструкция не привязана
                self.instruction_title_label.setText(self.current_task["task_title"])
                self.instruction_desc_label.setText("Инструкция пока не добавлена.")
                placeholder = QLabel("Подробная инструкция появится после её добавления.")
                placeholder.setWordWrap(True)
                placeholder.setStyleSheet("color: #5b6577; font-style: italic;")
                self.instruction_container_layout.addWidget(placeholder)

            elif self.current_category:
                # Раздел или подкатегория
                self.instruction_title_label.setText(self.current_category["name"])

                if self.current_category.get("parent_id") is None:
                    self.instruction_desc_label.setText("Выберите задачу ниже")
                else:
                    self.instruction_desc_label.setText("Выберите подзадачу ниже")

                placeholder = QLabel()
                placeholder.setWordWrap(True)
                self.instruction_container_layout.addWidget(placeholder)

                self.refresh_instruction_navigation_panel()

            else:
                # Вообще ничего не выбрано
                self.instruction_title_label.setText("Инструкция")
                self.instruction_desc_label.setText("Выберите задачу слева, чтобы открыть подробную инструкцию.")
                placeholder = QLabel("Здесь будут главы, блоки, картинки и внутренние ссылки.")
                placeholder.setWordWrap(True)
                self.instruction_container_layout.addWidget(placeholder)

            self._fit_instruction_content_width()
            return

        self.instruction_title_label.setText(self.current_instruction["instruction_title"])
        self.instruction_desc_label.setText(render_markdown(self.current_instruction["short_desc"]))

        self.current_section_titles = [s.get("title", "") for s in self.current_instruction["sections"]]

        self._section_cards = []          # сохраняем для sticky-заголовка
        for index, sec in enumerate(self.current_instruction["sections"]):
            card = CollapsibleSection(
                sec.get("title", ""),
                sec.get("blocks", []),
                self.handle_link_activated,
                section_index=index,
                feedback_handler=self.open_section_feedback_dialog,
            )
            self.instruction_container_layout.addWidget(card)
            self._section_cards.append(card)

        if self.current_instruction["related_ids"]:
            related_group = QGroupBox("Связанные инструкции")
            related_layout = QVBoxLayout(related_group)
            related_layout.setSpacing(6)

            for related_id in self.current_instruction["related_ids"]:
                rel = self.db.instruction_by_id(related_id)
                if not rel:
                    continue

                btn = QPushButton(f"{rel['instruction_title']} • {rel['category_name']}")
                btn.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
                btn.clicked.connect(lambda _, rid=rel["instruction_id"]: self.open_instruction_by_id(rid))
                related_layout.addWidget(btn)

            self.instruction_container_layout.addWidget(related_group)
        self.instruction_container_layout.addStretch(1)

        # Обновляем список блоков для комментариев
        self.comment_anchor_box.blockSignals(True)
        self.comment_anchor_box.clear()
        self.comment_anchor_box.addItem("Вся инструкция", "")
        for title in self.current_section_titles:
            if title:
                self.comment_anchor_box.addItem(title, title)
        self.comment_anchor_box.setCurrentIndex(0)
        self.comment_anchor_box.blockSignals(False)
        self.refresh_instruction_navigation_panel()

        # Подсвечиваем поисковый текст в описании инструкции
        search_text = self.search_edit.text().strip()
        if search_text:
            QTimer.singleShot(100, lambda: self._highlight_search_in_widget(self.instruction_desc_label, search_text))

        # Подсвечиваем поисковый текст в секциях инструкции
        if search_text:
            QTimer.singleShot(150, lambda: self._highlight_search_in_sections(search_text))

        self._fit_instruction_content_width()

    def edit_current_instruction(self):
        """Открывает редактор для изменения текущей инструкции (только админ)."""
        if not self.is_admin or not self.current_instruction:
            return

        categories = self.db.conn.execute(
            "SELECT id, name FROM categories ORDER BY sort_order, name"
        ).fetchall()
        categories = [dict(row) for row in categories]

        # Подготавливаем данные для предзаполнения
        editor_task = self.current_task
        if editor_task is None or editor_task.get("instruction_id") != self.current_instruction["instruction_id"]:
            editor_task = self.db.task_by_instruction_id(self.current_instruction["instruction_id"])
        data = {
            "instruction_id": self.current_instruction["instruction_id"],
            "category_id": self.current_instruction["category_id"],
            "task_id": editor_task["task_id"] if editor_task else None,
            "task_category_id": editor_task["category_id"] if editor_task else None,
            "title": self.current_instruction["instruction_title"],
            "short_desc": self.current_instruction["short_desc"],
            "sections": self.current_instruction["sections"],
            "related_ids": self.current_instruction["related_ids"]
        }

        dialog = InstructionEditorDialog(
            db=self.db,
            categories=categories,
            instruction_data=data,
            parent=self
        )

        if dialog.exec() != QDialog.Accepted:
            return

        new_data = dialog.get_data()
        try:
            missing = self.db.update_instruction(
                self.current_instruction["instruction_id"],
                new_data["category_id"],
                new_data["title"],
                new_data["short_desc"],
                new_data["sections"],
                new_data["related_ids"],
                new_data.get("task_id")
            )
        except Exception as exc:
            QMessageBox.warning(self, "Ошибка", f"Не удалось обновить инструкцию: {exc}")
            return

        self.reload_nav_tree()
        reopen_task_id = self.current_task["task_id"] if self.current_task else new_data.get("task_id")
        if reopen_task_id:
            self.show_task(reopen_task_id, force=True)
        self.statusBar().showMessage("Инструкция обновлена")
        if missing:
            QMessageBox.information(
                self,
                "Внимание",
                "Не найдены связанные инструкции:\n" + "\n".join(missing)
            )

    def open_checklist_editor(self):
        """Открывает редактор общего чек-листа, показанного в боковой панели."""
        if not self.is_admin:
            return

        active_task = self.current_checklist_task or self.current_task
        if not active_task:
            return

        saved_task_id = active_task["task_id"]
        selected_task_id = self.current_task["task_id"] if self.current_task else None
        saved_category_id = self.current_category["id"] if self.current_category else None

        instruction_id = active_task.get("instruction_real_id") or active_task.get("instruction_id")
        active_instruction = self.db.instruction_by_id(instruction_id) if instruction_id else None

        dialog = TaskEditorDialog(
            db=self.db,
            task=active_task,
            instruction=active_instruction,
            parent=self
        )

        if dialog.exec() != QDialog.Accepted:
            return

        data = dialog.get_data()
        try:
            self.db.update_task_view_data(
                saved_task_id,
                data["short_desc"],
                data["instruction_title"],
                data["checklist_sections"]
            )
        except Exception as exc:
            QMessageBox.warning(self, "Ошибка", f"Не удалось сохранить изменения: {exc}")
            return

        self.reload_nav_tree()
        if selected_task_id:
            self.show_task(selected_task_id, force=True)
        elif saved_category_id:
            self.show_category(saved_category_id)
        else:
            self.show_task(saved_task_id, force=True)
        self.statusBar().showMessage("Чек-лист обновлён")

    def handle_link_activated(self, link):
        """Все клики по ссылкам – внутренние или через диалог."""
        if link.startswith("instruction://id:"):
            instruction_id = int(unquote(link.replace("instruction://id:", "", 1)))
            self.open_instruction_by_id(instruction_id)
        elif link.startswith("instruction://"):
            title = unquote(link.replace("instruction://", "", 1))
            self.open_instruction_by_title(title)
        else:
            self.show_link_dialog(link)

    def open_instruction_by_title(self, title):
        instruction = self.db.instruction_by_title(title)
        if not instruction:
            QMessageBox.warning(self, "Внимание", f"Инструкция не найдена: {title}")
            return

        task = self.db.task_by_instruction_id(instruction["instruction_id"])

        # Снимаем поисковый фильтр, чтобы целевая инструкция точно была видна в дереве
        self.search_timer.stop()
        self.search_edit.blockSignals(True)
        self.search_edit.clear()
        self.search_edit.blockSignals(False)

        if task:
            self.reload_nav_tree()
            self.show_task(task["task_id"])
            self.tabs.setCurrentWidget(self.instruction_tab)
        else:
            self.show_instruction_only(instruction["instruction_id"])

    def open_instruction_by_id(self, instruction_id):
        instruction = self.db.instruction_by_id(instruction_id)
        if not instruction:
            QMessageBox.warning(self, "Внимание", f"Инструкция не найдена: {instruction_id}")
            return

        task = self.db.task_by_instruction_id(instruction["instruction_id"])

        self.search_timer.stop()
        self.search_edit.blockSignals(True)
        self.search_edit.clear()
        self.search_edit.blockSignals(False)

        if task:
            self.reload_nav_tree()
            self.show_task(task["task_id"])
            self.tabs.setCurrentWidget(self.instruction_tab)
        else:
            self.show_instruction_only(instruction["instruction_id"])

    def show_link_dialog(self, url: str):
        """Показывает диалог со ссылкой и кнопками 'Скопировать' и 'Закрыть'."""
        # Не показываем диалог для внутренних ссылок
        if url.startswith("instruction://id:"):
            self.open_instruction_by_id(int(unquote(url.replace("instruction://id:", "", 1))))
            return

        if url.startswith("instruction://"):
            self.open_instruction_by_title(unquote(url.replace("instruction://", "", 1)))
            return

        from PySide6.QtWidgets import QApplication

        dlg = QDialog(self)
        dlg.setWindowTitle("Ссылка")
        dlg.setMinimumWidth(480)

        layout = QVBoxLayout(dlg)

        hint = QLabel(
            "Если это ссылка на папку или файл на сервере — скопируй и вставь "
            "в строку поиска проводника.\n"
            "Если это ссылка на сайт — вставь в браузер."
        )
        hint.setWordWrap(True)
        layout.addWidget(hint)

        url_edit = QLineEdit(url)
        url_edit.setReadOnly(True)
        layout.addWidget(url_edit)

        btn_layout = QHBoxLayout()
        copy_btn = QPushButton("Скопировать")
        close_btn = QPushButton("Закрыть")

        def copy():
            QApplication.clipboard().setText(url)
            dlg.accept()

        copy_btn.clicked.connect(copy)
        close_btn.clicked.connect(dlg.reject)

        btn_layout.addStretch()
        btn_layout.addWidget(copy_btn)
        btn_layout.addWidget(close_btn)
        layout.addLayout(btn_layout)

        dlg.exec()

    # ================== Комментарии и оценка ==================

    def refresh_feedback_tab(self):
        """Обновляет всю вкладку: статистику, список блоков, комментарии."""
        if not hasattr(self, 'feedback_stats_label'):
            return

        if not self.current_instruction:
            self.feedback_stats_label.setText("Выберите инструкцию слева")
            self.submit_feedback_btn.setEnabled(False)
            self.feedback_comments_browser.setHtml("<i>Выберите инструкцию слева.</i>")
            return

        # Статистика
        avg_rating, cnt = self.db.rating_stats(self.current_instruction["instruction_id"])
        if cnt == 0:
            self.feedback_stats_label.setText("Пока нет оценок")
        else:
            self.feedback_stats_label.setText(
                f"Средняя оценка: {avg_rating:.1f} / 10 | Оценок: {cnt}"
            )

        # Обновляем список блоков для комментариев
        self.comment_anchor_box.blockSignals(True)
        self.comment_anchor_box.clear()
        self.comment_anchor_box.addItem("Вся инструкция", "")
        for title in self.current_section_titles:
            if title:
                self.comment_anchor_box.addItem(title, title)
        self.comment_anchor_box.setCurrentIndex(0)
        self.comment_anchor_box.blockSignals(False)

        # Комментарии
        comments = self.db.comments_for_instruction(self.current_instruction["instruction_id"])
        self.feedback_comments_browser.setHtml(self._render_comments_html(comments))

        # Подсвечиваем поисковый текст в комментариях
        search_text = self.search_edit.text().strip()
        if search_text:
            QTimer.singleShot(100, lambda: self._highlight_search_in_widget(self.feedback_comments_browser, search_text))

    def _render_comments_html(self, comments):
        if not comments:
            return "<p><i>Комментариев пока нет.</i></p>"

        blocks = []
        for c in comments:
            anchor = ""
            if c["anchor"]:
                anchor = f" <span style='color:#6b7280'>[блок: {escape_html(c['anchor'])}]</span>"

            rating = ""
            if c.get("rating") is not None:
                rating = (
                    " <span style='color:#b45309;font-weight:600;'>"
                    f"Оценка: {escape_html(str(c['rating']))}/10</span>"
                )

            blocks.append(
                "<div style='padding:10px 6px;border-bottom:1px solid #e5e7eb;'>"
                f"<div><b>{escape_html(c['author'])}</b>{anchor}{rating} "
                f"<span style='color:#6b7280'>{escape_html(c['created_at'])}</span></div>"
                f"<div style='margin-top:5px;line-height:1.45;'>{escape_html(c['text'])}</div>"
                "</div>"
            )

        return "".join(blocks)

    def open_section_feedback_dialog(self, anchor_title):
        """Открывает диалог комментария и оценки к конкретному блоку инструкции."""
        if not self.current_instruction:
            QMessageBox.warning(self, "Внимание", "Сначала выберите инструкцию.")
            return

        dialog = SectionFeedbackDialog(anchor_title, parent=self)
        if dialog.exec() != QDialog.Accepted:
            return

        data = dialog.get_data()
        if self._save_instruction_feedback(
            data["rating"],
            data["text"],
            data["author"],
            anchor_title,
        ):
            self.statusBar().showMessage(f"Отзыв по блоку «{anchor_title}» сохранён")

    def _save_instruction_feedback(self, rating, comment_text, author, anchor=""):
        """Сохраняет оценку и/или комментарий к инструкции или блоку."""
        if not self.current_instruction:
            return False

        if rating < 7 and not comment_text:
            QMessageBox.warning(
                self,
                "Требуется комментарий",
                "Для оценки ниже 7 необходимо оставить комментарий.\n"
                "Опиши, что именно непонятно или требует доработки.",
            )
            return False

        anchor = anchor or ""

        created_at = ts()
        comment_rating = int(rating) if (rating > 0 and comment_text) else None

        if rating > 0:
            try:
                self.db.add_rating(
                    self.current_instruction["instruction_id"],
                    rating,
                    anchor,
                    created_at,
                )
            except Exception as exc:
                QMessageBox.warning(self, "Ошибка сохранения оценки", str(exc))
                return False

        if comment_text:
            success, err_msg = self.db.add_comment(
                self.current_instruction["instruction_id"],
                anchor,
                author,
                False,
                comment_text,
                rating=comment_rating,
                created_at=created_at,
            )
            if not success:
                QMessageBox.warning(
                    self,
                    "Ошибка сохранения комментария",
                    f"Не удалось сохранить комментарий.\nПричина: {err_msg}\n\n"
                    "Проверьте права доступа к файлу базы данных и повторите попытку.",
                )
                return False

        self.refresh_feedback_tab()
        self.remember_db_state()
        return True

    def submit_feedback(self):
        """Ставит оценку и/или добавляет комментарий."""
        if not self.current_instruction:
            QMessageBox.warning(self, "Внимание", "Сначала выберите инструкцию.")
            return

        rating = self.rating_spin.value()
        comment_text = self.comment_text_edit.toPlainText().strip()
        anchor = self.comment_anchor_box.currentData() or ""
        author = self.comment_author_edit.text().strip() or "Пользователь"

        if not self._save_instruction_feedback(rating, comment_text, author, anchor):
            return

        self.comment_text_edit.clear()
        self.rating_spin.setValue(10)
        self.statusBar().showMessage("Оценка и комментарий добавлены")

    def open_feedback_manager(self):
        """Открывает диалог управления оценками и комментариями (только админ)."""
        if not self.is_admin or not self.current_instruction:
            return

        dialog = FeedbackManagerDialog(
            db=self.db,
            instruction_id=self.current_instruction["instruction_id"],
            parent=self
        )
        dialog.exec()
        self.refresh_feedback_tab()
        self.remember_db_state()

    def export_feedback_to_excel(self):
        if not self.is_admin:
            return

        default_name = f"feedback_export_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Экспорт в Excel",
            default_name,
            "Excel files (*.xlsx)"
        )
        if not file_path:
            return

        try:
            self.db.export_feedback_xlsx(file_path)
        except Exception as exc:
            QMessageBox.warning(self, "Ошибка", f"Не удалось экспортировать данные: {exc}")
            return

        self.statusBar().showMessage(f"Экспорт выполнен: {file_path}")

    # ================== Подсветка поиска ==================

    def _highlight_search_in_widget(self, widget, search_text):
        """Подсвечивает search_text в QLabel или QTextBrowser жёлтым фоном"""
        if not search_text or not widget:
            return

        needle = search_text.strip()
        if not needle:
            return

        import re

        if isinstance(widget, QLabel):
            # Сохраняем оригинальный HTML если ещё не сохранён
            if not widget.property("_original_html"):
                widget.setProperty("_original_html", widget.text())

            original = widget.property("_original_html")
            if not original:
                return

            # Ищем слова игнорируя регистр, но не внутри HTML-тегов
            # Разбиваем на части: текст вне тегов и внутри тегов
            def highlight_text_part(text):
                pattern = re.compile(r'(' + re.escape(needle) + r')', re.IGNORECASE)
                return pattern.sub(
                    r'<span style="background-color: #FFE7B8; padding: 1px 3px; border-radius: 3px;">\1</span>',
                    text
                )

            # Обрабатываем только текстовые участки (не внутри <...>)
            result = []
            i = 0
            while i < len(original):
                if original[i] == '<':
                    # Нашли начало тега — копируем до конца тега
                    end = original.find('>', i)
                    if end == -1:
                        result.append(original[i:])
                        break
                    result.append(original[i:end + 1])
                    i = end + 1
                else:
                    # Текст — ищем до начала следующего тега
                    end = original.find('<', i)
                    if end == -1:
                        result.append(highlight_text_part(original[i:]))
                        break
                    result.append(highlight_text_part(original[i:end]))
                    i = end

            widget.setText(''.join(result))

        elif isinstance(widget, QTextBrowser):
            # Сбрасываем предыдущую подсветку
            cursor = widget.textCursor()
            cursor.setPosition(0)
            cursor.movePosition(QTextCursor.End, QTextCursor.KeepAnchor)
            default_fmt = QTextCharFormat()
            default_fmt.setBackground(QColor("transparent"))
            cursor.mergeCharFormat(default_fmt)

            # Новая подсветка
            fmt = QTextCharFormat()
            fmt.setBackground(QColor("#FFE7B8"))

            cursor.setPosition(0)
            doc = widget.document()
            while True:
                cursor = doc.find(needle, cursor)
                if cursor.isNull():
                    break
                cursor.mergeCharFormat(fmt)

    def _highlight_search_in_tree(self):
        """Подсвечивает поисковый текст в элементах дерева"""
        search_text = self.search_edit.text().strip()

        def process_item(item):
            if not item:
                return

            # Получаем оригинальный чистый текст
            original = item.data(0, Qt.UserRole + 10)
            if not original:
                import re
                clean = re.sub(r'<[^>]+>', '', item.text(0))
                item.setData(0, Qt.UserRole + 10, clean)
                original = clean

            # Всегда сначала восстанавливаем оригинал
            item.setText(0, original)

            if search_text and search_text.casefold() in original.casefold():
                import re
                escaped = re.escape(search_text)
                pattern = re.compile(r'(' + escaped + r')', re.IGNORECASE)
                highlighted = pattern.sub(
                    r'<span style="background-color: #FFE7B8;">\1</span>',
                    original
                )
                item.setText(0, highlighted)

                font = item.font(0)
                font.setBold(True)
                item.setFont(0, font)
            else:
                kind = item.data(0, ROLE_KIND)
                font = item.font(0)
                if kind == "category":
                    font.setBold(item.parent() is None)
                else:
                    font.setBold(False)
                item.setFont(0, font)

            for i in range(item.childCount()):
                process_item(item.child(i))

        for i in range(self.nav_tree.topLevelItemCount()):
            process_item(self.nav_tree.topLevelItem(i))

        def highlight_items(parent):
            for i in range(parent.childCount() if hasattr(parent, 'childCount') else self.nav_tree.topLevelItemCount()):
                if hasattr(parent, 'topLevelItem'):
                    item = parent.topLevelItem(i)
                else:
                    item = parent.child(i)

                if not item:
                    continue

                text = item.text(0)

                # Восстанавливаем оригинальный текст
                original = item.data(0, Qt.UserRole + 10)
                if not original:
                    item.setData(0, Qt.UserRole + 10, text)
                    original = text

                # Всегда сначала ставим оригинал
                item.setText(0, original)

                if search_text and search_text.casefold() in original.casefold():
                    # Подсвечиваем жёлтым фоном через HTML
                    import re
                    pattern = re.compile(r'(' + re.escape(search_text) + r')', re.IGNORECASE)
                    highlighted = pattern.sub(
                        r'<span style="background-color: #FFE7B8; padding: 1px 2px; border-radius: 2px;">\1</span>',
                        original
                    )
                    item.setText(0, highlighted)

                    # Делаем шрифт жирным
                    font = item.font(0)
                    font.setBold(True)
                    item.setFont(0, font)
                else:
                    # Восстанавливаем обычный шрифт
                    kind = item.data(0, ROLE_KIND)
                    font = item.font(0)
                    if kind == "category":
                        parent_item = item.parent()
                        font.setBold(parent_item is None)
                    else:
                        font.setBold(False)
                    item.setFont(0, font)

                # Рекурсивно для дочерних
                if item.childCount() > 0:
                    highlight_items(item)

        highlight_items(self.nav_tree)

    def _clear_search_highlight(self):
        """Сбрасывает всю подсветку поиска"""
        search_text = self.search_edit.text().strip()
        if search_text:
            return  # Не сбрасываем если есть текст поиска

        # Сбрасываем QLabel'ы
        for widget in [self.task_desc_label, self.instruction_desc_label]:
            if hasattr(widget, 'property') and widget.property("_original_html"):
                original = widget.property("_original_html")
                widget.setText(original)
                widget.setProperty("_original_html", "")

        # Сбрасываем feedback_comments_browser
        if hasattr(self, 'feedback_comments_browser'):
            cursor = self.feedback_comments_browser.textCursor()
            cursor.setPosition(0)
            cursor.movePosition(QTextCursor.End, QTextCursor.KeepAnchor)
            default_fmt = QTextCharFormat()
            default_fmt.setBackground(QColor("transparent"))
            cursor.mergeCharFormat(default_fmt)

    def _highlight_search_in_sections(self, search_text):
        """Подсвечивает поисковый текст во всех секциях инструкции"""
        if not search_text or not hasattr(self, '_section_cards'):
            return

        needle = search_text.strip()
        if not needle:
            return

        import re

        for card in self._section_cards:
            # === Заголовок секции (QToolButton) ===
            # QToolButton НЕ поддерживает HTML в тексте!
            # Сохраняем оригинальный текст заголовка
            original_title = card.property("_original_title")
            if not original_title:
                clean = re.sub(r'<[^>]+>', '', card.toggle.text())
                card.setProperty("_original_title", clean)
                original_title = clean

            # Восстанавливаем оригинал
            card.toggle.setText(original_title)

            # Меняем фон кнопки через stylesheet если есть совпадение
            if needle.casefold() in original_title.casefold():
                card.toggle.setStyleSheet(card.toggle.styleSheet() + """
                    QToolButton {
                        background-color: #FFE7B8;
                    }
                """)
            else:
                # Сбрасываем стиль (он будет унаследован от родителя)
                card.toggle.setStyleSheet("")

            # === Содержимое секции (QLabel) ===
            self._highlight_search_in_children(card.content, needle)

    def _highlight_search_in_children(self, parent_widget, needle):
        """Рекурсивно подсвечивает текст во всех дочерних QLabel"""
        import re

        for child in parent_widget.findChildren(QLabel):
            if not child.text():
                continue

            # Сохраняем оригинальный HTML
            if not child.property("_original_section_html"):
                child.setProperty("_original_section_html", child.text())

            original = child.property("_original_section_html")

            # Восстанавливаем оригинал
            child.setText(original)

            # Если нечего подсвечивать — пропускаем
            if needle.casefold() not in original.casefold():
                continue

            # Подсвечиваем только текст вне HTML-тегов
            escaped = re.escape(needle)
            pattern = re.compile(r'(' + escaped + r')', re.IGNORECASE)
            replacement = r'<span style="background-color: #FFE7B8;">\1</span>'

            result = []
            i = 0
            while i < len(original):
                if original[i] == '<':
                    end = original.find('>', i)
                    if end == -1:
                        result.append(original[i:])
                        break
                    result.append(original[i:end + 1])
                    i = end + 1
                else:
                    end = original.find('<', i)
                    if end == -1:
                        result.append(pattern.sub(replacement, original[i:]))
                        break
                    result.append(pattern.sub(replacement, original[i:end]))
                    i = end

            child.setText(''.join(result))

    def closeEvent(self, event):
        try:
            if hasattr(self, "refresh_timer"):
                self.refresh_timer.stop()
            if hasattr(self, "update_check_timer"):
                self.update_check_timer.stop()
            self._backup_database()
            self.db.close()
        except Exception:
            pass
        super().closeEvent(event)

    def _backup_database(self):
        """Создаёт резервную копию базы данных в папке backups рядом с БД."""
        backup_dir = Path(DB_PATH).parent / "backups"
        backup_dir.mkdir(parents=True, exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_path = backup_dir / f"knowledge_base_{timestamp}.db"
        try:
            self.db.conn.commit()
            cleanup_sqlite_sidecar_files(DB_PATH)
            shutil.copy2(DB_PATH, backup_path)
            existing = sorted(
                backup_dir.glob("knowledge_base_*.db"),
                key=lambda p: p.stat().st_mtime,
                reverse=True,
            )
            for old in existing[BACKUP_MAX_COUNT:]:
                try:
                    old.unlink()
                except OSError:
                    pass
        except Exception as exc:
            print(f"Backup error: {exc}")


# ================== СЛУЖЕБНОЕ ==================

def clear_layout(layout):
    while layout.count():
        item = layout.takeAt(0)
        widget = item.widget()
        child_layout = item.layout()

        if widget is not None:
            widget.deleteLater()
        elif child_layout is not None:
            clear_layout(child_layout)


def main():
    app = QApplication(sys.argv)
    app.setStyle("Fusion")
    apply_light_palette(app)
    app.setApplicationName("База знаний по задачам")
    window = MainWindow()
    window.show()
    sys.exit(app.exec())


if __name__ == "__main__":
    main()
